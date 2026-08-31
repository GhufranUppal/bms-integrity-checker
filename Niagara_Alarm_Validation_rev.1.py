"""
Niagara_Alarm_Validation_rev.1.py
=================================
BMS point-integrity checker for Niagara N4 as-built exports.

This is a refactored, self-contained revision of ``Niagara_Alarm_Validation.py``.
The legacy, hard-coded ``pointcheckoutSiemens`` / ``pointcheckoutSchneider``
routines live on in that original file; this revision replaces them with a
single, vendor-agnostic validation engine plus a modern Tkinter interface.

What it does
------------
Compares an as-built alarm/trend configuration (exported from Niagara to CSV)
against the Control Point List design intent (``Controls_Design_Engineering.xlsx``)
and produces a colour-coded ``Validation_Report.xlsx``:

    * green  - the as-built value matches the design intent
    * red    - the as-built value differs from the design intent
    * yellow - point present in the field but not found in the design intent

The matcher is vendor agnostic: Siemens point names contain no separators
(``AHU1BMSSFSPDALM``) while Schneider names use underscores
(``AHU1_BMS_SF_SPD_ALM``); both normalise to the same key, so one engine
validates either export. Supported point families include the evaporative
cooler / AHU points, the Industrial Water System, the Electrical Room and the
CAT environmental sensors.

Usage
-----
    python Niagara_Alarm_Validation_rev.1.py                    # launch the GUI
    python Niagara_Alarm_Validation_rev.1.py validate FOLDER VENDOR

The Alarm and Trend CSVs are exported from the Niagara Workbench (Report Service
/ BQL); the CPL is the project design-intent Excel workbook.
"""

from __future__ import annotations

import os
import re
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill

try:  # matplotlib powers the optional results chart in the GUI.
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure

    _HAS_MPL = True
except Exception:  # pragma: no cover - chart is a nice-to-have, not required.
    _HAS_MPL = False


# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #
APP_TITLE = "N4 Point Validation Tool"
CPL_SHEET = "Evap Cooler Points"
CPL_FILENAME = "Controls_Design_Engineering.xlsx"
REPORT_FILENAME = "Validation_Report.xlsx"
VENDORS = ("schneider", "siemens")
_ENCODINGS = ("latin_1", "utf_8", "utf_8_sig")

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FONT_HEADER = Font(color="FFFFFF", bold=True)

REPORT_HEADERS = [
    "Equipment", "Point_Name", "Alarm_Class", "CPL_Level", "Class_Result",
    "Delay", "CPL_Delay", "Delay_Result",
    "High_Limit", "CPL_High", "High_Result",
    "Low_Limit", "CPL_Low", "Low_Result",
    "Dead_Band", "CPL_DB", "DB_Result",
    "Trend_Interval", "Overall",
]


# --------------------------------------------------------------------------- #
# Normalisation helpers
# --------------------------------------------------------------------------- #
def norm_token(value) -> str:
    """Uppercase and strip every non-alphanumeric character.

    Makes matching independent of separators so the Siemens style
    ``AHU1EVAP1STG1FLOWALM`` and the Schneider style ``AHU1_EVAP1_STG1_FLOW_ALM``
    reduce to the same key.
    """
    return re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()


def normalize_delay(value):
    """Convert a delay expression to integer seconds.

    ``'5 sec'`` / ``'2 minute'`` / ``5`` -> seconds; blank -> ``None``.
    """
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ("", "nan", "none"):
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    if not match:
        return None
    seconds = float(match.group(1))
    if "min" in text:
        seconds *= 60
    return int(seconds)


def normalize_number(value):
    """Extract the first numeric token as a float.

    Handles values such as ``'78'``, ``'78.0'`` or ``'78 degF'``; blank or
    non-numeric input returns ``None``. Used for the CAT sensor high/low limit
    and dead-band thresholds.
    """
    if value is None:
        return None
    text = str(value).strip()
    if text.lower() in ("", "nan", "none"):
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def _fmt_num(value):
    """Format a number for display, dropping a redundant ``.0`` on integers."""
    if value is None:
        return ""
    return str(int(value)) if float(value).is_integer() else str(value)


def read_csv_any(path):
    """Read a CSV trying a few common encodings."""
    for encoding in _ENCODINGS:
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception:
            continue
    raise RuntimeError(f"Could not read {path}")


# --------------------------------------------------------------------------- #
# Control Point List (design intent)
# --------------------------------------------------------------------------- #
def load_cpl(cpl_path):
    """Load the Control Point List keyed by normalised bracket point name."""
    try:
        df = pd.read_excel(cpl_path, sheet_name=CPL_SHEET, header=0)
    except Exception:
        df = pd.read_excel(cpl_path, sheet_name=0, header=0)
    df = df.rename(columns={c: str(c).strip() for c in df.columns})

    cpl = {}
    for _, row in df.iterrows():
        name = row.get("Point_Name")
        if not isinstance(name, str) or "[" not in name:
            continue  # skip section headings ("AHU", "CAT SENSORS") and blanks
        key = norm_token(name)
        high = normalize_number(row.get("HIGH LIMIT"))
        low = normalize_number(row.get("LOW LIMIT"))
        dead = normalize_number(row.get("DEAD BAND"))
        cpl[key] = {
            "description": row.get("Point_Description", ""),
            "delay": normalize_delay(row.get("ALARM DELAY")),
            "level": norm_token(row.get("NOTIFICATION LEVEL")),
            "high": high,
            "low": low,
            "dead": dead,
            # CAT sensors are analog OutOfRange points: they carry high/low
            # limits and a dead-band that must be validated numerically.
            "is_cat": "CAT" in key or high is not None or low is not None or dead is not None,
        }
    return cpl


def match_cpl(equipment, point_name, cpl):
    """Return ``(key, design)`` for the CPL entry matching an as-built point.

    The as-built name carries an equipment prefix (e.g. ``AHU1``); the CPL name
    is equipment-relative (e.g. ``[EVAP1][STG1][FLOW][ALM]``). Strip the
    equipment instance, then try two candidates: with the equipment family kept
    (for CAT/IWS/ELEC style points) and without it (for the generic evap points).
    """
    norm_point = norm_token(point_name)
    norm_equip = norm_token(equipment)
    if norm_equip and norm_point.startswith(norm_equip):
        suffix = norm_point[len(norm_equip):]
    else:
        suffix = norm_point
    family = norm_token(re.sub(r"\d+$", "", str(equipment)))  # AHU1 -> AHU, CAT1 -> CAT
    for candidate in (family + suffix, suffix):
        if candidate in cpl:
            return candidate, cpl[candidate]
    return None, None


# --------------------------------------------------------------------------- #
# Validation engine
# --------------------------------------------------------------------------- #
def validate_configuration(alarm_file1, alarm_file2, trend_file, cpl_path,
                           output_folder, vendor="schneider"):
    """Validate an as-built configuration against the CPL design intent.

    Compares, for every matched point, the Alarm Class against the CPL
    Notification Level and the Alarm Delay against the CPL Alarm Delay, writes a
    colour-coded ``Validation_Report.xlsx`` to ``output_folder`` and returns a
    result dict with ``stats``, display ``rows`` and the ``report_path``.
    """
    vendor = (vendor or "schneider").lower()

    alarms = pd.concat(
        [read_csv_any(alarm_file1), read_csv_any(alarm_file2)], ignore_index=True)
    alarms = alarms.drop_duplicates(subset=["Point_Name"], keep="first")

    try:
        trend = read_csv_any(trend_file).drop_duplicates(subset=["Point_Name"], keep="first")
        trend_lookup = dict(zip(trend["Point_Name"], trend["Trend_Interval"]))
    except Exception:
        trend_lookup = {}

    cpl = load_cpl(cpl_path)

    stats = {"evaluated": 0, "compliant": 0, "mismatch": 0, "not_found": 0}
    matched_keys = set()
    rows = []

    for _, row in alarms.iterrows():
        equipment = row.get("Equipment", "")
        point = row.get("Point_Name", "")
        cls = norm_token(row.get("Alarm_Class"))
        delay = normalize_delay(row.get("Delay"))
        trend_val = trend_lookup.get(point, row.get("Trend_Interval", ""))

        key, design = match_cpl(equipment, point, cpl)
        stats["evaluated"] += 1

        if design is None:
            rows.append({
                "equipment": equipment, "point": point,
                "alarm_class": row.get("Alarm_Class", ""), "cpl_level": "",
                "class_result": "REVIEW", "delay": row.get("Delay", ""),
                "cpl_delay": "", "delay_result": "REVIEW",
                "trend": trend_val, "overall": "NOT IN CPL",
            })
            stats["not_found"] += 1
            continue

        matched_keys.add(key)
        class_ok = cls == design["level"]
        delay_ok = delay == design["delay"]
        cpl_delay = "" if design["delay"] is None else f"{design['delay']} sec"

        row_out = {
            "equipment": equipment, "point": point,
            "alarm_class": row.get("Alarm_Class", ""), "cpl_level": design["level"],
            "class_result": "MATCH" if class_ok else "MISMATCH",
            "delay": row.get("Delay", ""), "cpl_delay": cpl_delay,
            "delay_result": "MATCH" if delay_ok else "MISMATCH",
            "high_limit": "", "cpl_high": "", "high_result": "",
            "low_limit": "", "cpl_low": "", "low_result": "",
            "dead_band": "", "cpl_db": "", "db_result": "",
            "trend": trend_val,
        }
        checks_ok = [class_ok, delay_ok]

        # CAT sensors additionally validate the numeric alarm thresholds
        # (high limit, low limit) and the dead-band against the design intent.
        if design["is_cat"]:
            limits = (
                ("High_Limit", design["high"], "high_limit", "cpl_high", "high_result"),
                ("Low_Limit", design["low"], "low_limit", "cpl_low", "low_result"),
                ("Dead_Band", design["dead"], "dead_band", "cpl_db", "db_result"),
            )
            for column, design_val, asb_key, cpl_key, res_key in limits:
                if design_val is None:
                    continue
                asb_val = normalize_number(row.get(column))
                ok = asb_val is not None and abs(asb_val - design_val) < 1e-9
                row_out[asb_key] = _fmt_num(asb_val)
                row_out[cpl_key] = _fmt_num(design_val)
                row_out[res_key] = "MATCH" if ok else "MISMATCH"
                checks_ok.append(ok)

        overall_ok = all(checks_ok)
        row_out["overall"] = "COMPLIANT" if overall_ok else "NON-COMPLIANT"
        rows.append(row_out)
        stats["compliant" if overall_ok else "mismatch"] += 1

    not_built = [(k, d) for k, d in cpl.items() if k not in matched_keys]
    report_path = _write_report(output_folder, vendor, rows, not_built, stats)

    print(f"[{vendor}] evaluated={stats['evaluated']} compliant={stats['compliant']} "
          f"mismatch={stats['mismatch']} not_found={stats['not_found']} -> {report_path}")
    return {"stats": stats, "rows": rows, "report_path": str(report_path),
            "not_built": len(not_built), "vendor": vendor}


def _color_result(cell):
    """Fill a result cell green (MATCH), red (MISMATCH) or yellow (REVIEW)."""
    value = cell.value
    if value == "MATCH":
        cell.fill = FILL_GREEN
    elif value == "MISMATCH":
        cell.fill = FILL_RED
    elif value == "REVIEW":
        cell.fill = FILL_YELLOW


def _write_report(output_folder, vendor, rows, not_built, stats):
    """Write the colour-coded validation workbook and return its path."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(REPORT_HEADERS)
    for col in range(1, len(REPORT_HEADERS) + 1):
        ws.cell(row=1, column=col).fill = FILL_HEADER
        ws.cell(row=1, column=col).font = FONT_HEADER

    for r in rows:
        ws.append([
            r["equipment"], r["point"], r["alarm_class"], r["cpl_level"], r["class_result"],
            r["delay"], r["cpl_delay"], r["delay_result"],
            r.get("high_limit", ""), r.get("cpl_high", ""), r.get("high_result", ""),
            r.get("low_limit", ""), r.get("cpl_low", ""), r.get("low_result", ""),
            r.get("dead_band", ""), r.get("cpl_db", ""), r.get("db_result", ""),
            r["trend"], r["overall"],
        ])
        row_idx = ws.max_row
        if r["overall"] == "NOT IN CPL":
            for col in (5, 8, 19):
                ws.cell(row=row_idx, column=col).fill = FILL_YELLOW
        else:
            for col in (5, 8, 11, 14, 17):  # class / delay / high / low / dead
                _color_result(ws.cell(row=row_idx, column=col))
            ws.cell(row=row_idx, column=19).fill = (
                FILL_GREEN if r["overall"] == "COMPLIANT" else FILL_RED)

    ws2 = wb.create_sheet("Design_Not_Built")
    ws2.append(["CPL_Key", "Description", "CPL_Level", "CPL_Delay"])
    for col in range(1, 5):
        ws2.cell(row=1, column=col).fill = FILL_HEADER
        ws2.cell(row=1, column=col).font = FONT_HEADER
    for key, design in not_built:
        cpl_delay = "" if design["delay"] is None else f"{design['delay']} sec"
        ws2.append([key, design["description"], design["level"], cpl_delay])
        for col in range(1, 5):
            ws2.cell(row=ws2.max_row, column=col).fill = FILL_YELLOW

    ws3 = wb.create_sheet("Stats")
    ws3.append(["Vendor", vendor])
    ws3.append(["Points evaluated", stats["evaluated"]])
    ws3.append(["Fully compliant", stats["compliant"]])
    ws3.append(["Non-compliant (rule mismatch)", stats["mismatch"]])
    ws3.append(["Not found in CPL (manual review)", stats["not_found"]])
    ws3.append(["Designed but not built", len(not_built)])

    # Native Excel bar chart summarising the outcome breakdown. Built from a
    # small category/value block so it renders without matplotlib.
    chart_start = ws3.max_row + 2
    ws3.cell(row=chart_start, column=1, value="Category")
    ws3.cell(row=chart_start, column=2, value="Count")
    # (label, value, bar colour) - colours mirror the Summary cell fills.
    breakdown = [
        ("Compliant", stats["compliant"], "70AD47"),      # green
        ("Rule mismatch", stats["mismatch"], "C00000"),   # red
        ("Manual review", stats["not_found"], "FFC000"),  # amber
        ("Not built", len(not_built), "808080"),          # grey
    ]
    for offset, (label, value, _colour) in enumerate(breakdown, start=1):
        ws3.cell(row=chart_start + offset, column=1, value=label)
        ws3.cell(row=chart_start + offset, column=2, value=value)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Validation Outcome - {vendor.title()}"
    chart.y_axis.title = "Number of points"
    chart.x_axis.title = "Outcome category"
    chart.legend = None
    chart.gapWidth = 60
    data = Reference(ws3, min_col=2, min_row=chart_start,
                     max_row=chart_start + len(breakdown))
    cats = Reference(ws3, min_col=1, min_row=chart_start + 1,
                     max_row=chart_start + len(breakdown))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Colour each bar individually and show the count above it.
    series = chart.series[0]
    for idx, (_label, _value, colour) in enumerate(breakdown):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties = GraphicalProperties(solidFill=colour)
        series.data_points.append(pt)
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.showLegendKey = False
    series.dLbls.showCatName = False
    series.dLbls.showSerName = False

    # Keep both axes and their labels visible.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.majorGridlines = None
    chart.height = 9
    chart.width = 17
    ws3.add_chart(chart, "D2")

    for sheet in (ws, ws2, ws3):
        for col_cells in sheet.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(width + 2, 60)

    Path(output_folder).mkdir(parents=True, exist_ok=True)
    report_path = Path(output_folder) / REPORT_FILENAME
    wb.save(report_path)
    return report_path


# --------------------------------------------------------------------------- #
# GUI
# --------------------------------------------------------------------------- #
def is_valid_path(filepath):
    """Return True if the path exists, otherwise show an error dialog."""
    if filepath and Path(filepath).exists():
        return True
    messagebox.showerror(APP_TITLE, f"Path not found:\n{filepath}")
    return False


class ValidationApp:
    """A compact Tkinter front-end for :func:`validate_configuration`."""

    _INPUTS = (
        ("alarm1", "Alarm File 1 (Boolean):", "csv"),
        ("alarm2", "Alarm File 2 (Numeric):", "csv"),
        ("trend", "Trend File:", "csv"),
        ("cpl", "Control Point List (CPL):", "xlsx"),
        ("output", "Output Folder:", "folder"),
    )

    def __init__(self, root):
        self.root = root
        self.entries = {}
        self.vendor = tk.StringVar(value="schneider")
        self.status = tk.StringVar(value="Ready.")
        self._result = None
        self._error = None
        self._chart_canvas = None

        root.title(APP_TITLE)
        root.geometry("1040x680")
        root.minsize(900, 600)

        self._build_style()
        self._build_header()
        self._build_inputs()
        self._build_controls()
        self._build_results()
        self._build_statusbar()

    # -- layout ----------------------------------------------------------- #
    def _build_style(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"), foreground="#1f3864")
        style.configure("Sub.TLabel", font=("Segoe UI", 9), foreground="#555555")
        style.configure("Run.TButton", font=("Segoe UI", 10, "bold"))
        style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))

    def _build_header(self):
        frame = ttk.Frame(self.root, padding=(14, 10, 14, 4))
        frame.pack(fill="x")
        ttk.Label(frame, text="Niagara N4 Point Validation",
                  style="Header.TLabel").pack(anchor="w")
        ttk.Label(frame, style="Sub.TLabel",
                  text="Compare an as-built configuration against the Control Point "
                       "List design intent.").pack(anchor="w")

    def _build_inputs(self):
        frame = ttk.LabelFrame(self.root, text="Input Files", padding=10)
        frame.pack(fill="x", padx=14, pady=6)
        frame.columnconfigure(1, weight=1)

        for row, (key, label, kind) in enumerate(self._INPUTS):
            ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", padx=4, pady=4)
            entry = ttk.Entry(frame)
            entry.grid(row=row, column=1, sticky="ew", padx=4, pady=4)
            ttk.Button(frame, text="Browse...",
                       command=lambda k=key, t=kind: self._browse(k, t)
                       ).grid(row=row, column=2, padx=4, pady=4)
            self.entries[key] = entry

    def _build_controls(self):
        frame = ttk.Frame(self.root, padding=(14, 2, 14, 2))
        frame.pack(fill="x")

        # Vendor-specific validation buttons (see README interface section).
        ttk.Button(frame, text="Point Validation (Schneider)", style="Run.TButton",
                   command=lambda: self._run_validation("schneider")).pack(side="left", padx=(0, 6))
        ttk.Button(frame, text="Point Validation (Siemens)", style="Run.TButton",
                   command=lambda: self._run_validation("siemens")).pack(side="left", padx=6)

        display_box = ttk.Frame(frame)
        display_box.pack(side="left", padx=6)
        self.display_choice = tk.StringVar(value="Alarm File 1")
        ttk.Combobox(display_box, textvariable=self.display_choice, width=16,
                     state="readonly",
                     values=["Alarm File 1", "Alarm File 2", "Trend File", "CPL"]
                     ).pack(side="left")
        ttk.Button(display_box, text="Preview",
                   command=self._preview_file).pack(side="left", padx=4)

        ttk.Button(frame, text="Exit", command=self.root.destroy).pack(side="right")

        self.progress = ttk.Progressbar(self.root, mode="indeterminate")
        self.progress.pack(fill="x", padx=14, pady=(6, 0))

    def _build_results(self):
        outer = ttk.Frame(self.root, padding=(14, 6, 14, 6))
        outer.pack(fill="both", expand=True)

        # Summary metrics + optional chart on the left.
        left = ttk.LabelFrame(outer, text="Summary", padding=10)
        left.pack(side="left", fill="y")
        self.summary_vars = {
            "evaluated": tk.StringVar(value="-"),
            "compliant": tk.StringVar(value="-"),
            "mismatch": tk.StringVar(value="-"),
            "not_found": tk.StringVar(value="-"),
        }
        labels = [
            ("Evaluated", "evaluated", "#1f3864"),
            ("Compliant", "compliant", "#2e7d32"),
            ("Mismatch", "mismatch", "#c62828"),
            ("Not in CPL", "not_found", "#f9a825"),
        ]
        for text, key, color in labels:
            row = ttk.Frame(left)
            row.pack(fill="x", pady=2)
            ttk.Label(row, text=text + ":").pack(side="left")
            tk.Label(row, textvariable=self.summary_vars[key], width=6, anchor="e",
                     font=("Segoe UI", 11, "bold"), fg=color).pack(side="right")

        if _HAS_MPL:
            self.figure = Figure(figsize=(3.0, 2.2), dpi=100)
            self.axes = self.figure.add_subplot(111)
            self._chart_canvas = FigureCanvasTkAgg(self.figure, master=left)
            self._chart_canvas.get_tk_widget().pack(fill="both", expand=True, pady=(8, 0))
            self._draw_chart(0, 0, 0)

        # Detailed results table on the right.
        right = ttk.LabelFrame(outer, text="Results", padding=6)
        right.pack(side="left", fill="both", expand=True, padx=(10, 0))
        columns = ("equipment", "point", "class_result", "delay_result", "overall")
        self.tree = ttk.Treeview(right, columns=columns, show="headings", height=12)
        headings = {
            "equipment": "Equipment", "point": "Point Name",
            "class_result": "Class", "delay_result": "Delay", "overall": "Overall",
        }
        widths = {"equipment": 90, "point": 220, "class_result": 80,
                  "delay_result": 80, "overall": 130}
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], anchor="w")
        self.tree.tag_configure("ok", background="#e7f5e9")
        self.tree.tag_configure("bad", background="#fdecea")
        self.tree.tag_configure("review", background="#fff8e1")
        vbar = ttk.Scrollbar(right, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vbar.pack(side="right", fill="y")

    def _build_statusbar(self):
        bar = ttk.Frame(self.root, relief="sunken", padding=(8, 3))
        bar.pack(fill="x", side="bottom")
        ttk.Label(bar, textvariable=self.status, style="Sub.TLabel").pack(side="left")

    # -- behaviour -------------------------------------------------------- #
    def _browse(self, key, kind):
        if kind == "folder":
            path = filedialog.askdirectory()
        elif kind == "xlsx":
            path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        else:
            path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if path:
            self.entries[key].delete(0, "end")
            self.entries[key].insert(0, path)

    def _paths(self):
        return {key: entry.get().strip() for key, entry in self.entries.items()}

    def _run_validation(self, vendor):
        paths = self._paths()
        required = ("alarm1", "alarm2", "trend", "cpl", "output")
        if not all(is_valid_path(paths[key]) for key in required):
            return

        self._result = None
        self._error = None
        self.vendor.set(vendor)
        self.status.set(f"Validating ({vendor})...")
        self.progress.start(12)

        def work():
            try:
                self._result = validate_configuration(
                    paths["alarm1"], paths["alarm2"], paths["trend"],
                    paths["cpl"], paths["output"], vendor)
            except Exception as exc:  # surfaced on the UI thread in _poll
                self._error = str(exc)

        thread = threading.Thread(target=work, daemon=True)
        thread.start()
        self._poll(thread)

    def _poll(self, thread):
        if thread.is_alive():
            self.root.after(150, lambda: self._poll(thread))
            return
        self.progress.stop()
        if self._error:
            self.status.set("Validation failed.")
            messagebox.showerror(APP_TITLE, f"Validation failed:\n{self._error}")
            return
        self._render_result(self._result)

    def _render_result(self, result):
        stats = result["stats"]
        for key, var in self.summary_vars.items():
            var.set(str(stats[key]))
        self._draw_chart(stats["compliant"], stats["mismatch"], stats["not_found"])

        self.tree.delete(*self.tree.get_children())
        for row in result["rows"]:
            if row["overall"] == "COMPLIANT":
                tag = "ok"
            elif row["overall"] == "NOT IN CPL":
                tag = "review"
            else:
                tag = "bad"
            self.tree.insert("", "end", tags=(tag,), values=(
                row["equipment"], row["point"], row["class_result"],
                row["delay_result"], row["overall"]))

        self.status.set(
            f"Done ({result['vendor']}): {stats['compliant']} compliant, "
            f"{stats['mismatch']} mismatch, {stats['not_found']} not in CPL. "
            f"Report: {result['report_path']}")
        messagebox.showinfo(
            APP_TITLE,
            f"Validation complete ({result['vendor']}).\n\n"
            f"Evaluated: {stats['evaluated']}\n"
            f"Compliant: {stats['compliant']}\n"
            f"Mismatches: {stats['mismatch']}\n"
            f"Not in CPL: {stats['not_found']}\n\n"
            f"Report saved to:\n{result['report_path']}")

    def _draw_chart(self, compliant, mismatch, not_found):
        if not _HAS_MPL:
            return
        self.axes.clear()
        self.axes.bar(
            ["Compliant", "Mismatch", "Not in CPL"],
            [compliant, mismatch, not_found],
            color=["#2e7d32", "#c62828", "#f9a825"])
        self.axes.set_ylabel("Points")
        self.axes.set_title("Validation Result", fontsize=9)
        self.axes.tick_params(axis="x", labelsize=8)
        self.figure.tight_layout()
        self._chart_canvas.draw()

    def _preview_file(self):
        choice = self.display_choice.get()
        key, reader = {
            "Alarm File 1": ("alarm1", pd.read_csv),
            "Alarm File 2": ("alarm2", pd.read_csv),
            "Trend File": ("trend", pd.read_csv),
            "CPL": ("cpl", pd.read_excel),
        }[choice]
        path = self.entries[key].get().strip()
        if not is_valid_path(path):
            return
        try:
            df = reader(path)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Could not read file:\n{exc}")
            return
        self._show_dataframe(df, Path(path).name)

    def _show_dataframe(self, df, title):
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.geometry("900x600")
        text = scrolledtext.ScrolledText(popup, wrap="none", font=("Consolas", 9))
        text.pack(fill="both", expand=True)
        text.insert("end", f"{df.dtypes}\n{'=' * 60}\n{df.to_string()}")
        text.configure(state="disabled")


def build_gui():
    root = tk.Tk()
    ValidationApp(root)
    root.mainloop()


# --------------------------------------------------------------------------- #
# Command-line entry point
# --------------------------------------------------------------------------- #
def _cli(argv):
    command = argv[1].lower() if len(argv) > 1 else ""

    if command == "validate" and len(argv) >= 3:
        folder = argv[2]
        vendor = argv[3] if len(argv) > 3 else "schneider"
        validate_configuration(
            os.path.join(folder, "Alarm_File_1.csv"),
            os.path.join(folder, "Alarm_File_2.csv"),
            os.path.join(folder, "Trend_File.csv"),
            os.path.join(folder, CPL_FILENAME),
            folder, vendor)
        return

    build_gui()


if __name__ == "__main__":
    import sys

    _cli(sys.argv)
