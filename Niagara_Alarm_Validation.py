import PySimpleGUI as sg
import pandas as pd
import xlwings as xw
from pathlib import Path
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill
import numpy as np
import os
import warnings
warnings.filterwarnings("ignore")
from pathlib import Path
import threading
import time as time
# import tkthread as tk
import matplotlib.pyplot as plt
import re



def pointcheckoutSiemens(alarmFile1,alarmFile2,trendFile1,cdepointList,output_folder):
    # Reading Input Files,Currently DefiningtheFiePath on C Drive
    #path ='C:'
    #alarm_File1_path= os.path.join(path,'\Projects\Script'+alarmFile1)
    #alarm_File2_path= os.path.join(path,'\Projects\Script'+alarmFile2)
    #trend_File_path= os.path.join(path,'\Projects\Script'+trendFile1)
    #cde_point_list_path= os.path.join(path,'\Projects\Script'+cdepointList)
    #output_File_path= os.path.join(path,'\Projects\Script'+outputFile)
    # Creating Data Frames 
    
    try:
        output_File_path = Path(output_folder) /"Reporttest.xlsx"
        alarm_File1_path= Path(alarmFile1)
        alarm_File2_path= Path(alarmFile2)
        trend_File_path = Path(trendFile1)
        cde_point_list_path= Path(cdepointList)
        report_file_path = Path (output_folder)/"Overview_Report.xlsx"
        chart_path = Path(output_folder) /"myplot1.png"

        global tot_point_eval, tot_point_mtach, tot_point_mismatch

        print(output_File_path)
        print( alarm_File1_path)
        print ( alarm_File2_path)
        print(trend_File_path)
        print(cde_point_list_path)
        print(chart_path)

        encoding_list = [ 'latin_1','utf_8', 'utf_8_sig']

        #path = 'C:'
        #path= os.path.join(path,'\Projects\Script'+'\Alarm_report_2_PDX.csv')
        #print(path)
        for encoding in encoding_list:
            worked = True
            try:
                df_Alarms_1 = pd.read_csv(alarm_File1_path, encoding=encoding)
            except:
                worked = False
            if worked:
                break
        
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[0]: "Equipment"}) 
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[1]: "Point_Name"}) 
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[2]: "Alarm_Extension"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[3]: "Alarm_Class"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[4]: "Alarm_Type"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[5]: "Delay"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[6]: "offNormal_Text"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[7]: "Normal_Text"})



        #df_Alarms_1=pd.read_csv(alarm_File1_path)
            
        df_Alarms_1['Point_Name_1'] =  df_Alarms_1['Equipment'] +'_'+ df_Alarms_1['Point_Name']
        df_Alarms_1.to_excel('C:\Projects\Script\Area\AlarmsAllBool.xlsx')
        df_merge_CAT_1 = df_Alarms_1[(df_Alarms_1['Point_Name'].notna()) & ((df_Alarms_1.Point_Name.str.contains ('Cat')) | (df_Alarms_1['Point_Name'].notna() & df_Alarms_1.Point_Name.str.contains ('CAT')))]
        df_merge_CAT_1 = df_merge_CAT_1[ df_merge_CAT_1.Alarm_Extension.str.contains ('OutOfRangeAlarmExt')]
        df_merge_CAT_1 = df_merge_CAT_1[['Point_Name','Alarm_Class','Delay' ]]

        df_merge_CAT_1 = df_merge_CAT_1.reset_index(drop = True)
        print(df_merge_CAT_1)
        df_merge_CAT_1['Point_Name'] = df_merge_CAT_1['Point_Name'] + df_merge_CAT_1.index.astype(str)
        


        #df_merge_CAT_1['Point_Name']= df_merge_CAT_1['Point_Name_1']
        df_merge_CAT_1.to_excel('C:\Projects\Script\Area\AlarmCATBool.xlsx')
        #df_CAT_Siemens = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("Cat") |df_Alrm_Trend.Point_Name.str.contains("CAT")) & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))| ( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('HardwareFault')))]
        for encoding in encoding_list:
            worked = True
            try:
                df_Alarms_2 = pd.read_csv(alarm_File2_path, encoding=encoding)
            except:
                worked = False
            if worked:
                break

        
        df_Alarms_2 = df_Alarms_2.rename(columns={list(df_Alarms_2.columns)[0]: "Equipment"})
        df_Alarms_2 = df_Alarms_2.rename(columns={list(df_Alarms_2.columns)[1]: "Point_Name"}) 
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[2]: "high_Limit"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[3]: "low_Limit"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[4]: "Dead_Band"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[5]: "high_Limit_Text"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[6]: "low_Limit_Text"})

        


        #df_Alarms_2=pd.read_csv(alarm_File2_path)
            
        df_merge_CAT = df_Alarms_2[(df_Alarms_2['Point_Name'].notna() & df_Alarms_2.Point_Name.str.contains ('Cat')) | (df_Alarms_2['Point_Name'].notna() & df_Alarms_2.Point_Name.str.contains ('CAT')) ]
        df_merge_CAT.to_excel('C:\Projects\Script\Area\AlarmCAT.xlsx')

        df_Alarms_2['Point_Name_1'] =  df_Alarms_2['Equipment'] +'_'+ df_Alarms_2['Point_Name']
        df_Alarms_2 =df_Alarms_2.drop(columns =['Equipment'])

        # Filtering for not cat 
        #df[~df.Point_Name.str.contains('Cat')]
        #df_Alarms_2 = df_Alarms_2[df_Alarms_2['Point_Name'].notna() & ~(df_Alarms_2.Point_Name.str.contains('Cat'))]
        df_Alarms_2.to_excel('C:\Projects\Script\Area\DropCAT.xlsx')

        #df_Alarms_2.to_excel ('C:\Projects\PDX\Report\NotMergedCATPoints.xlsx')

        #df_Alarms_3 = df_Alarms_2 [df_Alarms_2.Point_Name.str.contains('Cat')]
        #df_Alarms_3.to_excel ('C:\Projects\PDX\Report\MergedCATPoints.xlsx')


        # Merging Alarm DataFrames
        df_Alarms_All = pd.merge(df_Alarms_1, df_Alarms_2, left_on = 'Point_Name_1', right_on ='Point_Name_1', how = 'outer')
        #df_Alarms_All= df_Alarms_All[df_Alarms_All['Point_Name_1'].notna() & ~df_Alarms_All.Point_Name_1.astype(str).str.contains('Cat')]
        #df_CAT_bool =df_Alarms_All[df_Alarms_All.Point_Name_1.str.contains ('Cat')]
        #print (df_CAT_bool)
        
        #df_CAT_bool.to_excel('C:\Projects\PDX\CATALarm.xlsx')

        mask = df_Alarms_All['Point_Name_1'].notna() & ~df_Alarms_All['Point_Name_1'].astype(str).str.contains('Cat')
        df_Alarms_All = df_Alarms_All[mask]

        mask1 = df_Alarms_All['Point_Name_1'].notna() & ~df_Alarms_All['Point_Name_1'].astype(str).str.contains('CAT')
        df_Alarms_All = df_Alarms_All[mask1]

        df_Alarms_All.to_excel('C:\Projects\Script\Area\AlarmALL.xlsx')


        # Let us merge CAT Sensors

        #df_merge_CAT['Point_Name_1'] =  df_merge_CAT['Point_Name']
        #df_merge_CAT.drop(columns = ['Equipment','Point_Name'])
        df_merge_CAT['Point_Name_1'] = df_merge_CAT['Point_Name']
        df_merge_CAT = df_merge_CAT[['Point_Name_1','high_Limit','low_Limit','Dead_Band']]
        df_merge_CAT.to_excel('C:\Projects\Script\Area\AlarmCATBEFOREMERGE.xlsx')
        df_merge_CAT['Point_Name'] = df_merge_CAT['Point_Name_1']
        df_merge_CAT = df_merge_CAT.reset_index(drop = True)
        df_merge_CAT['Point_Name'] = df_merge_CAT['Point_Name'] + df_merge_CAT.index.astype(str)
        df_merge_CAT.to_excel('C:\Projects\Script\Area\AlarmsNumeric.xlsx')

        df_merge_CAT_2 = pd.merge(df_merge_CAT_1,df_merge_CAT,on = ['Point_Name' ], how = 'inner')
        
        df_merge_CAT_2['Alarm_Extension'] = 'Alarms'

        df_merge_CAT_2.to_excel('C:\Projects\Script\Area\MergedcatCOMPLETE.xlsx')

        df_Alarms_All = pd.merge(df_Alarms_All, df_merge_CAT_2, on = ['Point_Name_1','Alarm_Extension', 'Alarm_Class','Delay','high_Limit','low_Limit','Dead_Band' ], how = 'outer')
        df_Alarms_All.to_excel('C:\Projects\Script\Area\AlarmCATAFTERMERGE.xlsx')
        print(df_Alarms_All.columns)
        

    
        #df_Alarms_All = df_Alarms_All.append(df_Alarms_3)

        #df_Alarms_All = df_Alarms_All.append(df_Alarms_2.Point_Name_1.str.contains('Cat'))
        
        df_Alarms_All =df_Alarms_All[['Point_Name_1','Equipment','Alarm_Extension', 'Alarm_Class','Alarm_Type', 'Delay', 'offNormal_Text', 
                                    'Normal_Text',  'Point_Name_y', 'high_Limit', 'low_Limit', 'Dead_Band','High_Limit_Text', 
                                    'low_Limit_Text']]
        df_Alarms_All.to_excel('C:\Projects\Script\Area\AlarmAll.xlsx')

        for encoding in encoding_list:
            worked = True
            try:
                df_Trends_1 = pd.read_csv(trend_File_path, encoding=encoding)
            except:
                worked = False

            if worked:
              break

        try:
            df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[0]: "Equipment"})
            df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[1]: "Point_Name"})
            df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[2]: "Trend_Type"}) 
            df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[3]: "Trend_Interval"})

        except:
            print  ("Check the Trend File !")

       

        # Combining with Trend Data
        #df_Trends_1 = pd.read_csv(trend_File_path)
        #list1 = ['BooleanCov','NumericInterval','EnumCov','NumericCov', 'NumericInteval']
        list1 = ['BooleanCov','NumericInterval','EnumCov','NumericCov', 'NumericInteval','NumericInterval_05','NumericInterval_01','NumericInterval_15']
        df_Trends_1 = df_Trends_1.loc[df_Trends_1['Trend_Type'].isin(list1)]
        df_Trends_1['Point_Name_1'] = df_Trends_1['Equipment'] +'_'+ df_Trends_1['Point_Name']
        # Merging Alarm and Trend Data
        df_Alrm_Trend = pd.merge(df_Alarms_All, df_Trends_1, left_on = 'Point_Name_1', right_on ='Point_Name_1', how = 'outer')
        df_Alrm_Trend= df_Alrm_Trend.drop (columns = ['Point_Name_y','Point_Name','Equipment_y'])
        df_Alrm_Trend =df_Alrm_Trend.rename(columns={"Point_Name_1": "Point_Name",'Equipment_x':'Equipment' })
        #Dealing with B-Formatting
        ### A commonly encountered B-Formatting is %alarmData.highlimit% and %alarmData.lowlimit%, checking for B- Formatting 
        ###to see if it is correctly formatted
        df_Alrm_Trend_1 = df_Alrm_Trend[['high_Limit','low_Limit','High_Limit_Text','low_Limit_Text' ]]
        list7=list(df_Alrm_Trend_1['low_Limit_Text'].str.contains('%alarmData.lowLimit%') == True)
        list8 = list(enumerate(list7))
        list9=[]
        for i in range (len(list8)):
            if(list8[i][1] == True):
                list9.append(list8[i][0])
        first_row_in_excel=2
        list10=[x+ first_row_in_excel for x in list9]
        string="N"
        list11= ["{}{}".format(string,i) for i in list10]
        ## Highlighting the Cell Containing B-Formatting as %alarmData.highLimit%, highlighting the rows 
        ##which contains this B-Formatted String
        list17=list(df_Alrm_Trend_1['High_Limit_Text'].astype(str).str.contains('%alarmData.highLimit%') == True)
        list18 = list(enumerate(list7))
        list19=[]
        for i in range (len(list18)):
            if(list18[i][1] == True):
                list19.append(list8[i][0])
        first_row_in_excel=2
        list20=[x+ first_row_in_excel for x in list19]
        string="M"
        list21= ["{}{}".format(string,i) for i in list20]
        ## Transfering combined data to Excel
        writer = pd.ExcelWriter(output_File_path, engine="xlsxwriter")
        df_Alrm_Trend.to_excel(writer, sheet_name='Summary')
        writer.save()
        writer.close()
        df_Alrm_Trend.to_excel('C:\Projects\Script\Area\TestRaw.xlsx')
        ## Highlighting the Cells Containing B- formatting of %alarmData.lowLimit%
        wb = openpyxl.load_workbook(output_File_path)
        ws = wb['Summary'] #Name of the working sheet
        fill_cell1 = PatternFill(patternType='solid', fgColor='ffff00')
        for cell in list11:
            ws[cell].fill = fill_cell1
        wb.save(output_File_path)
        ## Highlighting the Cells Containing B- formatting of %alarmData.highLimit%
        wb = openpyxl.load_workbook(output_File_path)
        ws = wb['Summary'] #Name of the working sheet
        fill_cell1 = PatternFill(patternType='solid', fgColor='ffff00')
        for cell in list21:
            ws[cell].fill = fill_cell1
        wb.save(output_File_path)
        ## Adding Functionality of Uploading CDE Point List that can be used to compare Alarms
        ##importing Point List and renaming Columns
        df_Compare = pd.read_excel(cde_point_list_path,sheet_name ='BMS Points',skiprows=[0])
        df_Compare = df_Compare.rename(columns ={"FUNCTION":'Point_Description',df_Compare.columns[1]:'Point_Name', "DEFINITION":"Trend_interval",df_Compare.columns[17]:"Alarm",df_Compare.columns[18]:"Alarm LO/HI"}).drop(columns=['JITDC','OPTDC','Site Specific','Relinquish Default','NOTES'])
        df_Compare_1 = df_Compare[['Point_Description','Point_Name']]
        ## Splitting Point_Name in the DataFrame of the Point List
        df_Compare_2 =df_Compare_1['Point_Name'].str.split(r'\[(.*?)\]', expand = True).astype(object).mask(lambda x: x.isna(), None)
        ##Creating a Data Frame to start comparing the DAHU Points in combined Report with CDE Point List 
        
        ## Creating an Index in Consolidated Alarm Trend Data Frame to track rows
        df_Alrm_Trend = df_Alrm_Trend.reset_index()
        df_Alrm_Trend.to_excel('C:\Projects\Script\Area\TestRaw1.xlsx')

        ## Filtering the Consolidated Alarm Data Frame for DAHU key
        df_Compare_11 = df_Alrm_Trend[df_Alrm_Trend.Point_Name.str.contains("DAHU") & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]

        global len_list_DAHU_1
        len_list_DAHU_1 =0  # This parameter tracks the points which are being checked 
        global len_list_DAHU_2 # Not used for now
        len_list_DAHU_2 =0
        global len_list_DAHU_3 # This parameter tracks the points which are not Compliant
        len_list_DAHU_3 =0 
        global len_list_DAHU_4 # Not used for now
        len_list_DAHU_4=0

        # df_CAT_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('CAT') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        ## This Line of Code is Simens Specific, it is because Siemens uses both the Small and Cap Letter 
        ## for Point Name
        ##Also need to add rules to delete the numbers that come in between the alphabets
        
        if(df_Compare_11.shape[0]> 1):
            # This Code looks at the CDE Point List and will Execuete Only if there is a 'DAHU' in the Database  
            list_ids=[]
            # This code block Loops through the 'CDE Point List' and find the points associated with Equipment Type DAHU or DAHU-Evap
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'DAHU - Evap'):
                    idx_start =index
                    list_ids.append(idx_start)
                elif (df_Compare.loc[index,'Point_Description'] == 'DAHU'):
                    idx_start =index
                    list_ids.append(idx_start)
            
            # Create a Data Frame for the Points associated with 'DAHU' or 'DAHU - Evap' 

            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            df_Compare_8.fillna("",inplace=True)
            df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
            df_Compare_9.to_excel('C:\Projects\Script\DAHUSiemens.xlsx') # This is a Check Point Only 
            for index, rows in df_Compare_8.iterrows():
                df_Compare_9.loc[index,'new_col'] = df_Compare_9.loc[list_ids[0],'new_col'] 
                df_Compare_10 = df_Compare_9.loc[list_ids[0]+1:list_ids[1], ]
            # Create a Dictionary 
            
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_10.iterrows():
                key = df_Compare_9.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_9.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]

            df_Compare_11 = df_Compare_11.assign(Point_Name=df_Compare_11.Point_Name.str.upper())
            df = df_Compare_11
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\df.xlsx') # This is a checkpoint only and not part of the final code 
            print(df)
            df[4] = df[4].fillna('').astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            # Creating a List of Dictionaries to iterate over
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            # Creating the list of Data Frames
            list_df_DAHU=[]
            for lst in list_pts:
                list1 = ''.join(lst)
                df_DAHU=df[df.Point_Name.str.contains(list1)]
                list_df_DAHU.append(df_DAHU)
            # Creating Keys to Compare the Data Frames
            # First for the CDE Point List
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            # Creating a key to Match in Siemens Report Generated
            list_merge_1=[]
            for x in range(0,len(list_df_DAHU),1):
                df_DAHU = list_df_DAHU[x]
                df_Compare_DAHU_1 =df_DAHU.assign( key1='')
            # For the 1st iteration of loop, match should be the first og pts, for second iteration second of points
                match=''
                list_match=list_pts[x]
                for y in range(len(list_match)):
                    match=match+list_match[y]
                #print(match,x)
                list_match_1=[]
                list_match_1.append(match)
                for i in list_match_1:
                    df_Compare_DAHU_1['key1']=df_Compare_DAHU_1['Point_Name'].str.contains(i).map({True:i,False:np.nan})
            # Merging the keys between dataframes
                df_DAHU_Point = pd.merge(df_Compare_DAHU_1 ,df_Compare_Point,left_on='key1',right_on='key2', how='inner')
                df_DAHU_Point_1=df_DAHU_Point[['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                if (df_DAHU_Point.empty == False):
                     list_merge_1.append(df_DAHU_Point)
                #print(df_DAHU_Point)
                #print(df_DAHU_Point_1)
                #print(list_merge_1[46][['key1','key2','Alarm_Class','NOTIFICATION LEVEL']])
            ## Doing the Comparision of Differet Features
            ## Alarm Class - Comparing Alarm Class with Notification class
            for x in range(0,len(list_merge_1),1):
                #print(list_merge_1[x])
                df_DAHU_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                #print(df_DAHU_Alm_Class)
                df_diff_7=  df_DAHU_Alm_Class [df_DAHU_Alm_Class['Alarm_Class'].astype(str).values ==df_DAHU_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                #print(df_diff_7)
                list_wtht_diff=  df_DAHU_Alm_Class['index'].to_list()
                len_list_DAHU_1 = len_list_DAHU_1 + len(list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                len_list_DAHU_3 = len_list_DAHU_3 + len(list_smlr)
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                #print(list_pts)
                ## Alarm Delay - Comparing Alarm Delay between the Point List and Report
                df_DAHU_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_DAHU_delay_Alarm['Delay'] = df_DAHU_delay_Alarm['Delay'].apply(str)
                df_DAHU_delay_Alarm['ALARM DELAY'] = df_DAHU_delay_Alarm['ALARM DELAY'].apply(str)
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('seconds', 'sec')
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('minutes', 'min')
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('1 minute', '60 sec')
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('1min', '60 sec')
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('5secs', '5 sec')



                #df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('1 minute', '60 sec')
                df_diff_8 = df_DAHU_delay_Alarm [df_DAHU_delay_Alarm['Delay'].astype(str).values ==df_DAHU_delay_Alarm['ALARM DELAY'].astype(str).values]
                list_wtht_diff=  df_DAHU_delay_Alarm['index'].to_list()
                len_list_DAHU_1 = len_list_DAHU_1 + len(list_wtht_diff)
                list_smlr = df_diff_8['index'].to_list()
                len_list_DAHU_3 = len_list_DAHU_3 + len(list_smlr)
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                #print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                '''
                ## Alarm Text - Comparing Alarm Delay between the Point List and Report    
                df_Compare_text = list_merge_1[x][['index','Point_Name_x','Point_Name_y','ALARM TXT','offNormal_Text']]
                df_diff_9 = df_Compare_text [df_Compare_text['ALARM TXT'].astype(str).values ==df_Compare_text ['offNormal_Text'].astype(str).values]
                list_wtht_diff=  df_Compare_text['index'].to_list()
                list_smlr = df_diff_9['index'].to_list()
                list12 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                first_row_in_excel=2
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                highlight_rows_alarm_text=[x+ first_row_in_excel for x in list12]
                string="H"
                Cell_numbers_Difference_Highlighted_2= ["{}{}".format(string,i) for i in highlight_rows_alarm_text]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_2:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                ## High Low Limit Test  - Not Required for DAHU 
                ## Checking for Trends 
                df_Compare_Eqpt_Trnd = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Trend_Interval','Trend_interval']]
                df_Compare_Eqpt_Trnd['Trend_Interval']= df_Compare_Eqpt_Trnd['Trend_Interval'].str.replace('irregular', 'COV')
                df_diff_10 = df_Compare_Eqpt_Trnd [df_Compare_Eqpt_Trnd['Trend_Interval'].astype(str).values ==df_Compare_Eqpt_Trnd ['Trend_interval'].astype(str).values]
                list_wtht_diff=  df_Compare_Eqpt_Trnd['index'].to_list()
                list_smlr = df_diff_10['index'].to_list()
                first_row_in_excel=2
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                list13 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                highlight_rows_alarm_trend=[x+ first_row_in_excel for x in list12]
                string="P"
                Cell_numbers_Difference_Highlighted_3= ["{}{}".format(string,i) for i in highlight_rows_alarm_trend]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_3:
                    ws[cell].fill = fill_cell1
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb.save(output_File_path)
                '''
                # Highlighting Rows Which were not Validated
            DAHU_cons = pd.concat(list_df_DAHU)
            DAHU_cons_duplicate = DAHU_cons[DAHU_cons.index.duplicated()]
            DAHU_cons_1=DAHU_cons.drop_duplicates(subset = 'index')
            list_DAHU_1= df_Compare_11['index'].to_list()
            list_all_DAHU_found = DAHU_cons_1['index'].to_list()
            list_all_DAHU_not_found = [int(x) for x in list_DAHU_1 if x not in list_all_DAHU_found]
            first_row_in_excel=2
            highlight_rows_DAHU=[x+ first_row_in_excel for x in list_all_DAHU_not_found]
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_DAHU:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)
        else :

            print (" This site does not have DAHUs ? ")
        
        df_Alrm_Trend.to_excel('C:\Projects\Script\Area\TotalAlarmIW.xlsx')

        df_IW = df_Alrm_Trend[(df_Alrm_Trend.Equipment.str.contains("IW"))& (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')) |( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt')))]
        if (df_IW.empty == True):
            df_IW= df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('_Flow', case=False, regex=True) | df_Alrm_Trend['Point_Name'].str.contains('_Valve', case=False, regex=True) | df_Alrm_Trend['Point_Name'].str.contains('_CityPressure', case=False, regex=True) | df_Alrm_Trend['Point_Name'].str.contains('_SysPressure', case=False, regex=True) ]
        df_IW_Shape= df_IW
        df_IW_all_found = df_IW
        df_IW.to_excel('C:\Projects\Script\Area\BeforeFilteringIW.xlsx')

        dataIW = [{'Point_Description': 'Evap Cooler', 'Point_Name': '[CONDUCTIVITY]', 'ALARM DELAY' :'10 sec', 'NOTIFICATION LEVEL' : 'Medium'},
                  {'Point_Description': 'Evap Cooler', 'Point_Name': '[FILLALARM]', 'ALARM DELAY' :'1 min', 'NOTIFICATION LEVEL' : 'Medium'},]
        df_dict_IW = pd.DataFrame.from_dict(dataIW, orient='columns')
        df_dict_IW_1 =df_dict_IW['Point_Name'].str.split(r'\[(.*?)\]', expand = True).astype(object).mask(lambda x: x.isna(), None)
        print('df_dict_IW')
        print (df_dict_IW)
        df_dict_IW.to_excel('C:\Projects\Script\Area\ExtraIw.xlsx')




        #df_Compare_IW = df_Alrm_Trend[((df_Alrm_Trend.Point_Name.str.contains("Flow")) | df_Alrm_Trend.Point_Name.str.contains(" Flow")| (df_Alrm_Trend.Point_Name.str.contains("Conductivity")))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        #df_Compare_IW. to_excel('C:\Projects\Script\IWSiemens.xlsx')
        
        print('df_Compare')
        print(df_Compare)
        df_Compare.to_excel('C:\Projects\Script\Area\CompareFresh.xlsx')
        
        if (df_IW.empty == False):
            global len_list_IW_Smns_1
            global len_list_IW_Smns_3
            len_list_IW_Smns_1 =0
            len_list_IW_Smns_3 =0
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'INDUSTRIAL WATER SYSTEM'):
                    idx_start =index
                    list_ids.append(idx_start)
            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:] 
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_IW_CDE =df_Compare.loc[list_ids[0]:list_ids[1], :]
            print('df_Compare_IW_CDE')
            print(df_Compare_IW_CDE)
            df_Compare_IW_CDE.to_excel('C:\Projects\Script\Area\ExtractIW.xlsx')
            df_Compare_IW_CDE  = pd.concat([df_Compare_IW_CDE, df_dict_IW],axis=0)
            print('df_Compare_IW_CDE') 
            print(df_Compare_IW_CDE) 
            df_Compare_IW_CDE.to_excel('C:\Projects\Script\Area\AppendedIW.xlsx')

            df_Comp_IW_CDE_pat = df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            df_Comp_IW_CDE_pat  = pd.concat([df_Comp_IW_CDE_pat, df_dict_IW_1],axis=0)
            df_Comp_IW_CDE_pat.fillna("",inplace=True)
            df_Compare_IW_CDE_1 = df_Comp_IW_CDE_pat.assign(new_col=df_Comp_IW_CDE_pat[3].str.replace('##',''))
            df_Compare_IW_CDE_1['new_col'] = 'IW'
            df_Compare_IW_CDE_1.to_excel('C:\Projects\Script\IWSiemens.xlsx')
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_IW_CDE_1.iterrows():
                key = df_Compare_IW_CDE_1.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_IW_CDE_1.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
            print(list2)
            df = df_IW
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.assign(Point_Name=df.Point_Name.str.upper())
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\dfIW.xlsx') # This is a checkpoint only and not part of the final code 
            print(df)
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            print(list_pts)
            list_pts_1=[]
            for lst in list_pts:
                if not ((len(lst) == 1) & (lst[0] =='PRESS')) |((len(lst) == 1) & (lst[0] =='MODE')) | ((len(lst) == 1) & (lst[0] =='IW')) |(((len(lst) == 1) & (lst[0] =='MAT') )|((len(lst) == 1) & (lst[0] =='SS') |((len(lst) == 1) & (lst[0] =='EVAP')))|((len(lst) == 1) & (lst[0] =='SAT'))):
                    list_pts_1.append(lst)

            print(list_pts_1)

            list_df_IW=[]

            for lst in list_pts_1:
                list1 = ' '.join(lst)
                df_IW=df[df.Point_Name.str.contains(list1)]
                df_IW['Point_Name'] = df_IW['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_IW)
                list_df_IW.append(df_IW)
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :]
            df_Compare_IW_CDE['key2']= df_Compare_IW_CDE.Point_Name.str.replace('[','').str.replace(']','')
            #df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_IW_CDE.to_excel('C:\Projects\Script\df_Compare_Point_IW.xlsx')
            print('list_df_IW')
            print(list_df_IW)
            df_Compare_Point = df_Compare_IW_CDE

            cons = pd.concat(list_df_IW)
            cons.to_excel('C:\Projects\Script\df_Consolidtaed_IW.xlsx')

            list_merge_1=[]
            for x in range(0,len(list_df_IW),1):
                df_IWS = list_df_IW[x]
                if df_IWS.empty:
                    continue
                print('Raw')
                print(df_IWS[['Point_Name']])

                if df_IWS['Point_Name'].str.contains('POINTS').any():
                    print("IW IS THERE")
                    df_IWS['key1'] = df_IWS['Point_Name'].str.replace('POINTS', '')
                    df_IWS.to_excel('C:\Projects\Script\key_problem_IW.xlsx')
                elif df_IWS['Point_Name'].str.contains('IW').any():
                    print("IW IS THERE")
                    df_IWS['key1'] = df_IWS['Point_Name'].str.replace('IW', '')
                    df_IWS.to_excel('C:\Projects\Script\key_problem_IW.xlsx')
        
                merged_dfs = []
                for index, row in df_Compare_Point.iterrows():
                    #print ('Trying to Merge')
                    #print(row['key2'])
                    filtered_df = df_IWS[df_IWS['key1'].str.contains(row['key2'], case=False, regex=False)]
                    if not filtered_df.empty:
                        #print('What are we merging')
                        #print(df_Compare_Point.loc[[index]])
                        df_matched = df_Compare_Point.loc[[index]]
                        df_matched = df_matched.reset_index(drop = True)
                        #print('df_matched')
                        #print(df_matched)
                        df_matched['match1']= ['dummy']
                        #print('df_AHU_before_Key')
                        # print(df_AHU)
                        df_IWS.index = df_IWS.index.astype(str)
                        df_IWS['match2'] = ['dummy'] * len(df_IWS)
                        #df_AHU['match2']= ['dummy']
                        merged_df = pd.merge(df_IWS, df_matched, left_on='match2', right_on='match1', how = 'inner')
                        print(merged_df)
                        merged_df.to_excel('C:\Projects\Script\key_problem_3.xlsx')
                        df_IW_Point = merged_df
                        print(df_IW_Point[['key1','key2']])
                        list_merge_1.append(df_IW_Point)
                        print('The Length of List is')
                        print (len(list_merge_1))

            for x in range(0,len(list_merge_1),1):
                print ('The Length of List is')
                print (len(list_merge_1))
                print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                print(list_merge_1[x])
                

                
                # This Code Checks the Configuration of the Notification Class against Cde Point List

                df_IW_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                print('df_IW_Alm_Class')
                print(df_IW_Alm_Class)
                #print(df_DAHU_Alm_Class)
                len_list_IW_Smns_1 = len_list_IW_Smns_1 + len (df_IW_Alm_Class['index'].to_list())
                df_diff_7=  df_IW_Alm_Class [df_IW_Alm_Class['Alarm_Class'].astype(str).values ==df_IW_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                #print(df_diff_7)
                list_wtht_diff=  df_IW_Alm_Class['index'].to_list()
                # For AHUs, we are evaluating these configurations : 1. Notification Class 2. Alarm Delay
                # Adding the Notification Class Files
                len_list_IW_Smns_3 = len_list_IW_Smns_3 + len (list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                #len_list_AHU_3 =len_list_AHU_3 + len (list10)
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

                # This Code Checks the Configuration of the Notification Class against Cde Point List

                df_IW_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                
                df_IW_delay_Alarm['Delay'] = df_IW_delay_Alarm['Delay'].apply(str)
                df_IW_delay_Alarm['ALARM DELAY'] = df_IW_delay_Alarm['ALARM DELAY'].apply(str)
                print(df_IW_delay_Alarm.dtypes)
                len_list_IW_Smns_1 = len_list_IW_Smns_1 + len (df_IW_delay_Alarm['index'].to_list())
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('1 minute', '1 min')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                df_IW_delay_Alarm['Delay'] = df_IW_delay_Alarm['Delay'].astype(str).str.replace('0 ms', 'nan')
                df_IW_delay_Alarm['Delay'] = df_IW_delay_Alarm['Delay'].astype(str).str.replace('10 seconds', '10 sec')
                print('df_IW_delay_Alarm')
                print(df_IW_delay_Alarm)
                #df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('0 ms', '')
                list_wtht_diff=  df_IW_delay_Alarm['index'].to_list()
                print('list_wtht_diff')
                print(list_wtht_diff)
                df_diff_8 = df_IW_delay_Alarm [df_IW_delay_Alarm['Delay'].astype(str).values ==df_IW_delay_Alarm['ALARM DELAY'].astype(str).values]
                # list_wtht_diff=  df_IW_delay_Alarm['index'].to_list()
                list_smlr = df_diff_8['index'].to_list()
                print('similiar_Delay')
                len_list_IW_Smns_3 =len_list_IW_Smns_3 + len (list_smlr)
                # list_smlr= [x+ first_row_in_excel for x in list_smlr]
                print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('Not_similiar_Delay')
                print(list11)
                #len_list_AHU_4=len_list_AHU_4 + len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

            IW_cons = cons
            IW_cons_duplicate = IW_cons[IW_cons.index.duplicated()]
            IW_cons_1=IW_cons.drop_duplicates(subset = 'index')
            list_IW_1= df_IW_all_found['index'].to_list()
            print ('list_IW_1')
            print(list_IW_1)
            list_all_IW_found = IW_cons_1['index'].to_list()
            list_all_IW_not_found = [int(x) for x in list_IW_1 if x not in list_all_IW_found]
            print ('list_IW_not_found')
            print (list_all_IW_not_found)
            first_row_in_excel=2
            highlight_rows_IW=[x+ first_row_in_excel for x in list_all_IW_not_found]
            print('list_not_found')
            print (highlight_rows_IW)
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_IW:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)
            
            df_IW_ylw = df_Alrm_Trend[(df_Alrm_Trend.Equipment.str.contains("IW"))& ( df_Alrm_Trend['Alarm_Extension'].str.contains('StatusAlarmExt'))]
            if (df_IW_ylw.empty == False):
                list_IW_ylw_1= df_IW_ylw['index'].to_list()
                print(list_IW_ylw_1)
                first_row_in_excel=2
                highlight_rows_IW_ylw=[x+ first_row_in_excel for x in list_IW_ylw_1]
                with xw.App(visible=False)as app:
                    updated_wb= app.books.open(output_File_path)
                    updated_ws = updated_wb.sheets('Summary')
                    rng=updated_ws.used_range
                    print(rng.address) 
                    for row in rng.rows:
                        if row.row in highlight_rows_IW_ylw:
                            row.color ='ffff00'
                    updated_wb.save(output_File_path)

        

        df_Compare_ER = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("Electric") |df_Alrm_Trend.Point_Name.str.contains("D1E2"))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        df_Compare_ER.to_excel('C:\Projects\Script\ElectricRoom.xlsx') # This is a Check Point Only and will not be part of the final implementation  
        if (df_Compare_ER.shape[0] >=1):
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Electric Rm / Catcher Rm  (Area Controller)'):
                    idx_start =index
                    list_ids.append(idx_start)
                
            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            print(df_Compare_8)
            df_Compare_8.fillna("",inplace=True)
            df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
            df_Compare_9['new_col'] = 'ER'
            df_Compare_9.to_excel('C:\Projects\Script\ERSchneider.xlsx')
            # Create a Dictionary 
            print ("The Script is evalauting Electric Room it may take several minutes, please be patient")
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_9.iterrows():
                key = df_Compare_9.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_9.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
                print ("The Script is evalauting Electric Room it may take several minutes, please be patient")
            df = df_Compare_ER
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.assign(Point_Name=df.Point_Name.str.upper())
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\df_ER.xlsx') # This is a checkpoint only and not part of the final code 
            print(df)
            #df[4] = df[4].fillna('').astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            # Need to look at other Siemens sites to check if the above code is required and create a condition accordingly 
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            # Creating a List of Dictionaries to iterate over
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            print(list_pts)
            list_df_ER=[]
            for lst in list_pts:
                list1 = ''.join(lst)
                df_ER=df[df.Point_Name.str.contains(list1)]
                df_ER['Point_Name'] = df_ER['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_ER)
                list_df_ER.append(df_ER)
             # First for the CDE Point List, create a key that will be matched to the Database 
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_Point.to_excel('C:\Projects\Script\df_Compare_Point_ER.xlsx')
            list_merge_1=[]
            for x in range(0,len(list_df_ER),1):
                df_ER = list_df_ER[x]
                if df_ER.empty:
                    continue
                print('Raw')
                #print(df_ER[['Point_Name']])
                df_ER['key1'] = df_ER ['Point_Name']
                print(df_ER)
                merged_dfs =[]
                for index, row in df_Compare_Point.iterrows():
                    filtered_df = df_ER[df_ER['key1'].str.contains(row['key2'], case=False, regex=False)]
                    if not filtered_df.empty:
                        #print('What are we merging')
                        #print(df_Compare_Point.loc[[index]])
                        df_matched = df_Compare_Point.loc[[index]]
                        df_matched = df_matched.reset_index(drop = True)
                        #print('df_matched')
                        #print(df_matched)
                        df_matched['match1']= ['dummy']
                        df_ER.index = df_ER.index.astype(str)
                        df_ER['match2'] = ['dummy'] * len(df_ER)
                        merged_df = pd.merge(df_ER, df_matched, left_on='match2', right_on='match1', how = 'inner')
                        print(merged_df)
                        df_ER_Point = merged_df
                        df_ER_Point.to_excel('C:\Projects\Script\SchneiderER.xlsx')
                        print(df_ER_Point[['key1','key2']])
                        list_merge_1.append(df_ER_Point)
                        print('The Length of List is')
                        print (len(list_merge_1))
            
            global len_list_ER_1,len_list_ER_2,len_list_ER_3
            len_list_ER_1 =0 # Tracks the total number of Electrical Room Points
            len_list_ER_2 =0
            len_list_ER_3 =0  # Traks the number of Electrical Room Points which are similiar 


            for x in range(0,len(list_merge_1),1):
                print ('The Length of List is')
                print (len(list_merge_1))
                print ("The Script is evalauting Electric Rom it may take several minutes, please be patient")
                print(list_merge_1[x])
                list_merge_1[x].to_excel('C:\Projects\Script\Test44.xlsx' )
                df_ER_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                #df_ER_Alm_Class = list_merge_1[x][['index','Point_Name','Alarm_Class','NOTIFICATION LEVEL']]
                print('df_ER_Alm_Class')
                print(df_ER_Alm_Class)
                len_list_ER_1 = len_list_ER_1 + len (df_ER_Alm_Class['index'].to_list())
                df_diff_7=  df_ER_Alm_Class [df_ER_Alm_Class['Alarm_Class'].astype(str).values ==df_ER_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                list_wtht_diff=  df_ER_Alm_Class['index'].to_list()
                len_list_ER_3 = len_list_ER_3 + len (list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print(list10)
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

                df_ER_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_ER_delay_Alarm['Delay'] = df_ER_delay_Alarm['Delay'].apply(str)
                df_ER_delay_Alarm['ALARM DELAY'] = df_ER_delay_Alarm['ALARM DELAY'].apply(str)
                print('df_ER_delay_Alarm')
                print(df_ER_delay_Alarm)
                print(df_ER_delay_Alarm.dtypes)
                len_list_ER_1 = len_list_ER_1 + len (df_ER_delay_Alarm['index'].to_list())
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('1 minute', '60 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('0.0', 'nan')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('30000', '30 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('10000', '10 sec')
                
                
                df_diff_8 = df_ER_delay_Alarm [df_ER_delay_Alarm['Delay'].astype(str).values ==df_ER_delay_Alarm['ALARM DELAY'].astype(str).values]
                list_wtht_diff=  df_ER_delay_Alarm['index'].to_list()

                list_smlr = df_diff_8['index'].to_list()
                print('similiar_Delay')
                len_list_ER_3 =len_list_ER_3 + len (list_smlr)
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('Not_similiar_Delay')
                print(list11)
                #len_list_AHU_4=len_list_AHU_4 + len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
            print ("The Script is evalauting ERs it may take several minutes, please be patient")
            ER_cons = pd.concat(list_df_ER)
            ER_cons_duplicate = ER_cons[ER_cons.index.duplicated()]
            ER_cons_1=ER_cons.drop_duplicates(subset = 'index')
            list_ER_1= df_Compare_ER['index'].to_list()
            list_all_ER_found = ER_cons_1['index'].to_list()
            list_all_ER_not_found = [int(x) for x in list_ER_1 if x not in list_all_ER_found]
            first_row_in_excel=2
            highlight_rows_ER=[x+ first_row_in_excel for x in list_all_ER_not_found]
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_ER:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)    

        df_Compare_CRAH = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("Crahu") | df_Alrm_Trend.Point_Name.str.contains("CRAH")) & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        df_Compare_CRAH.to_excel('C:\Projects\Script\CRAH.xlsx') 
        if (df_Compare_CRAH.shape[0] >=1):
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'CRAH'):
                    idx_start =index
                    list_ids.append(idx_start)
            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            print(df_Compare_8)
            df_Compare_8.fillna("",inplace=True)
            df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
            df_Compare_9['new_col'] = 'CRAH'
            df_Compare_9.to_excel('C:\Projects\Script\CRAHSchneider.xlsx')
            # Create a Dictionary 
            print ("The Script is evalauting Electric Room it may take several minutes, please be patient")
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_9.iterrows():
                key = df_Compare_9.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_9.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            print(list2)

            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
                print ("The Script is evalauting CRAHUs it may take several minutes, please be patient")
            df = df_Compare_CRAH
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.assign(Point_Name=df.Point_Name.str.upper())
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\df_CRAH.xlsx') # This is a checkpoint only and not part of the final codE
            print(df)
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            df.to_excel('C:\Projects\Script\df_CRAH_1.xlsx') # This is a checkpoint only and not part of the final codE
            # Creating a List of Dictionaries to iterate over
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            print(list_pts)
            list_df_CRAH=[]
            for lst in list_pts:
                list1 = ' '.join(lst)
                df_CRAH=df[df.Point_Name.str.contains(list1)]
                df_CRAH['Point_Name'] = df_CRAH['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_CRAH)
                list_df_CRAH.append(df_CRAH)
            # First for the CDE Point List, create a key that will be matched to the Database
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_Point.to_excel('C:\Projects\Script\df_Compare_Point_CRAH.xlsx')
            list_merge_1=[]
            for x in range(0,len(list_df_CRAH),1):
                df_CRAH = list_df_CRAH[x]
                if df_CRAH.empty:
                    continue
                print('Raw')
                #print(df_ER[['Point_Name']])
                df_CRAH['key1'] = df_CRAH ['Point_Name']
                print(df_CRAH)
                merged_dfs =[]
                for index, row in df_Compare_Point.iterrows():
                    filtered_df = df_CRAH[df_CRAH['key1'].str.contains(row['key2'], case=False, regex=False)]
                    if not filtered_df.empty:
                        #print('What are we merging')
                        #print(df_Compare_Point.loc[[index]])
                        df_matched = df_Compare_Point.loc[[index]]
                        df_matched = df_matched.reset_index(drop = True)
                        #print('df_matched')
                        #print(df_matched)
                        df_matched['match1']= ['dummy']
                        df_CRAH.index = df_CRAH.index.astype(str)
                        df_CRAH['match2'] = ['dummy'] * len(df_CRAH)
                        merged_df = pd.merge(df_CRAH, df_matched, left_on='match2', right_on='match1', how = 'inner')
                        print(merged_df)
                        df_CRAH_Point = merged_df
                        df_CRAH_Point.to_excel('C:\Projects\Script\SchneiderCRAH.xlsx')
                        print(df_CRAH_Point[['key1','key2']])
                        list_merge_1.append(df_CRAH_Point)
                        print('The Length of List is')
                        print (len(list_merge_1)) 
            
            global len_list_CRAH_1,len_list_CRAH_2,len_list_CRAH_3
            len_list_CRAH_1 =0 # Total Number of CRAHU Points
            len_list_CRAH_2 =0
            len_list_CRAH_3 =0 # Total Number of CRAHU Points which are similiar

            for x in range(0,len(list_merge_1),1):
                print ('The Length of List is')
                print (len(list_merge_1))
                print ("The Script is evalauting Electric Rom it may take several minutes, please be patient")
                print(list_merge_1[x])
                list_merge_1[x].to_excel('C:\Projects\Script\TestCRAH.xlsx' )
                df_CRAH_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                #df_ER_Alm_Class = list_merge_1[x][['index','Point_Name','Alarm_Class','NOTIFICATION LEVEL']]
                print('df_ER_Alm_Class')
                print(df_CRAH_Alm_Class)
                len_list_CRAH_1 = len_list_CRAH_1 + len (df_CRAH_Alm_Class['index'].to_list())
                df_diff_7=  df_CRAH_Alm_Class [df_CRAH_Alm_Class['Alarm_Class'].astype(str).values ==df_CRAH_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                list_wtht_diff=  df_CRAH_Alm_Class['index'].to_list()
                len_list_CRAH_3 = len_list_CRAH_3 + len (list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print(list10)
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

                df_CRAH_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_CRAH_delay_Alarm['Delay'] = df_CRAH_delay_Alarm['Delay'].apply(str)
                df_CRAH_delay_Alarm['ALARM DELAY'] = df_CRAH_delay_Alarm['ALARM DELAY'].apply(str)
                print('df_CRAH_delay_Alarm')
                print(df_CRAH_delay_Alarm)
                print(df_CRAH_delay_Alarm.dtypes)
                len_list_CRAH_1 = len_list_CRAH_1 + len (df_CRAH_delay_Alarm['index'].to_list())
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('1 minute', '60 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('0.0', 'nan')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('30000', '30 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('10000', '10 sec')

                df_CRAH_delay_Alarm.loc[(df_CRAH_delay_Alarm['Delay'] == '0.0') & (df_CRAH_delay_Alarm['ALARM DELAY'] == 'nan'), 'Delay'] = 'nan'
                #df_CRAH_delay_Alarm['ALARM DELAY'] = np.where((df_CRAH_delay_Alarm['Delay'] == 0.0) & (df_CRAH_delay_Alarm['ALARM DELAY'].isna()), 0.0, df_CRAH_delay_Alarm['ALARM DELAY'])

                df_diff_8 = df_CRAH_delay_Alarm [df_CRAH_delay_Alarm['Delay'].astype(str).values ==df_CRAH_delay_Alarm['ALARM DELAY'].astype(str).values]
                list_wtht_diff=  df_CRAH_delay_Alarm['index'].to_list()

                list_smlr = df_diff_8['index'].to_list()
                print('similiar_Delay')
                len_list_CRAH_3 =len_list_CRAH_3 + len (list_smlr)
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('Not_similiar_Delay')
                print(list11)
                #len_list_AHU_4=len_list_AHU_4 + len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
            print ("The Script is evalauting ERs it may take several minutes, please be patient")
            CRAHU_cons = pd.concat(list_df_CRAH)
            CRAHU_cons_duplicate = CRAHU_cons[CRAHU_cons.index.duplicated()]
            CRAHU_cons_1=CRAHU_cons.drop_duplicates(subset = 'index')
            list_CRAHU_1= df_Compare_CRAH['index'].to_list()
            list_all_CRAHU_found = CRAHU_cons_1['index'].to_list()
            list_all_CRAHU_not_found = [int(x) for x in list_CRAHU_1 if x not in list_all_CRAHU_found]
            first_row_in_excel=2
            highlight_rows_CRAHU=[x+ first_row_in_excel for x in list_all_CRAHU_not_found]
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_CRAHU:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)     
 
 



        # Adding the Code for Sites which have AHUs ( Typical JITDC) - AHUs are identified by the presence of string 'AHU' or 'OAHU' in the Point Name 
        if (df_Compare_11.empty == True):

            global df_Compare_12
            
            df_Compare_12 = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("AHU") |df_Alrm_Trend.Point_Name.str.contains("Ahu"))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            
            #global df_Compare_12
            # This code will execute only if the Data base has AHUs or OAHUs
            if(df_Compare_12.shape[0]>= 1):


                print ("The Script is evalauting AHUs it may take several minutes, please be patient")

                global len_list_AHU_1
                len_list_AHU_1 =0  # This parameter tracks the points which are being checked 
                global len_list_AHU_2 # Not used for now
                len_list_AHU_2 =0
                global len_list_AHU_3 # This parameter tracks the points which are not Compliant
                len_list_AHU_3 =0 
                global len_list_AHU_4 # Not used for now
                len_list_AHU_4=0

                # Adding the code so nothing needs to be added so no manual addition is required 

                data1 = [{'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG1][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG2][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG1][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG2][FLOW][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[PWR][UPS][STATUS]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High' }, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][MODE][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SAT][STPT][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SF][SPD][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][LOWER][ALM]','ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'Medium'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][UPPER][ALM]','ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[ECFAN][FAIL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][CMD]','ALARM DELAY' :'30 secs','NOTIFICATION LEVEL' : 'Medium'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][HIGH][WATER][ALM]', 'ALARM DELAY' :'5 sec','NOTIFICATION LEVEL' : 'High' },
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][LOW][WATER][ALM]', 'ALARM DELAY' :'5 sec','NOTIFICATION LEVEL' : 'High' },
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][LOWER][STS]', 'ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][UPPER][STS]', 'ALARM DELAY' :'2 minute', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FILDP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[PREFILTER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FINALFILTER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][LOWER][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][UPPER][ALM]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][HIGH][ALM]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][LOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T1]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T2]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'},
            ]
            
                
                # Creating the Data Frame 

                # Creating a Data Frame 

                df_dict = pd.DataFrame.from_dict(data1, orient='columns')
                df_dict =df_dict['Point_Name'].str.split(r'\[(.*?)\]', expand = True).astype(object).mask(lambda x: x.isna(), None)
                

                # This Code looks at the CDE Point List and will Execuete Only if there is a 'AHU' in the Database  

                list_ids=[]

                # This code block Loops through the 'CDE Point List' and find the points associated with Equipment Type AHU 

                for index,rows in df_Compare.iterrows():
                    if (df_Compare.loc[index,'Point_Description'] == 'AHU'):
                        idx_start =index
                        list_ids.append(idx_start)

                # Create a Data Frame for the Points associated with 'AHU' 



                df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
                df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
                list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
                df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
                df_Compare_8  = pd.concat([df_Compare_8, df_dict],axis=0)
                print(df_Compare_8)
                df_Compare_8.fillna("",inplace=True)
                df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
                df_Compare_9['new_col'] = 'AHU'
                df_Compare_9.to_excel('C:\Projects\Script\AHUSiemens.xlsx') # This is a Check Point Only and will not be part of the final implementation  
                
                # Create a Dictionary 
                print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
                list1=[]
                list2=[]
                for index,rows in df_Compare_9.iterrows():
                    key = df_Compare_9.loc[index,'new_col']
                    for x in list_val:
                        list1.append(df_Compare_9.loc[index,x])
                    value =list1
                    dicT= {key : value} 
                    list1=[]
                    list2.append(dicT)
                list102=[]
                list101=[]
                for index in range(len(list2)):
                    for i,x in enumerate (list2[index][key]):
                        if (x == ''):
                            if(i <= len(list2[index][key])):
                                list101.append(i)
                    list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                    list101=[]
                    print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                
                #df_Compare_12 = df_Compare_12.assign(Point_Name=df_Compare_12.Point_Name.str.upper())
                df = df_Compare_12
                df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
                df = df.assign(Point_Name=df.Point_Name.str.upper())
                df = df.join(df.Point_Name.str.split(r' ',expand=True))
                df.to_excel('C:\Projects\Script\df.xlsx') # This is a checkpoint only and not part of the final code 
                print(df)
                #df[4] = df[4].fillna('').astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
                # Need to look at other Siemens sites to check if the above code is required and create a condition accordingly 
                df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
                # Creating a List of Dictionaries to iterate over
                list_pts=[]
                for dicts in list2:
                    for eqpt,points in dicts.items():
                        list_pts.append(points) 
                
                list_pts_1=[]
                for lst in list_pts:
                    if not ((len(lst) == 1) & (lst[0] =='OAD')) |((len(lst) == 1) & (lst[0] =='BYPD')) | ((len(lst) == 1) & (lst[0] =='RAD')) |(((len(lst) == 1) & (lst[0] =='MAT') )|((len(lst) == 1) & (lst[0] =='SS') |((len(lst) == 1) & (lst[0] =='EVAP')))|((len(lst) == 1) & (lst[0] =='SAT'))):
                        list_pts_1.append(lst)
        
                print(list_pts_1)

                # Creating the list of Data Frames
                list_df_AHU=[]
                list_not_found =[]
                for lst in list_pts_1:
                    list1 = ''.join(lst)
                    print(list1)
                    df_AHU=df[df.Point_Name.str.contains(list1)]
                    # Adding Functionality for the Points where they could not be a Match 
                    if (df_AHU.empty == True):
                        print("True")
                        list_not_found.append(list1)
                    list_df_AHU.append(df_AHU)
                print('list_not_found')
                print(list_not_found)
                global df_notfound
                
                '''
                df_notfound = pd.DataFrame({'col':list_not_found})
                print(df_notfound)
                df_notfound.to_excel('C:\Projects\Script\Missing.xlsx')
                '''
                list_comp= [('OAFILFDPT','FINALFILTER'),('PAFILPDPT','PREFILTER'),('EVAPHIGHWATERALM','SUMPHILVLALM') ]
                dict2={'Missing_Alarm':''}
                df_notfound = pd.DataFrame()
                dummy_1=0
                for x in list_comp:
                    for y in list_not_found:
                        if(y == x[0]):
                            dummy_1 = dummy_1+1
                            for y in list_not_found:
                                #print (y) 
                                if(y==x[1]):
                                    dummy_1 = dummy_1+1
                    if (dummy_1 ==2):
                        #print(dummy_1)
                        dict2['Missing_Alarm'] = x[1]
                        df_notfound_1 = pd.DataFrame.from_dict(dict2,orient = 'index')
                        df_notfound = pd.concat([df_notfound,df_notfound_1])
                        df_notfound_1 = pd.DataFrame()
                        print( dict2)
                    dummy_1 =0 
                    dict2={'Missing_Alarm':''}


                # writer = pd.ExcelWriter(output_File_path, engine="xlsxwriter")
                # df_notfound.to_excel(writer, sheet_name='Alarms_not_found') 
                # vwriter.save()
                # writer.close()


                # Creating Keys to Compare the Data Frames
                # Adding a Sheet to Excel File, where this data will be stored


                # First for the CDE Point List, create a key that will be matched to the Database 
                df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
                df_dict_1 = pd.DataFrame.from_dict(data1, orient='columns')
                df_Compare_Point  = pd.concat([df_Compare_Point, df_dict_1],axis=0,ignore_index = True)
                df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
                
                # Adding the code to concatentae the extra Points 


                # Creating a key for the Database 
                list_merge_1=[]

                for x in range(0,len(list_df_AHU),1):
                    df_AHU = list_df_AHU[x]
                    print(df_AHU)
                    df_Compare_AHU_1 =df_AHU.assign( key1='')
                    match=''
                    list_match=list_pts_1[x]
                    for y in range(len(list_match)):
                        match=match+list_match[y]
                    print(match,x)
                    list_match_1=[]
                    list_match_1.append(match)
                    for i in list_match_1:
                        df_Compare_AHU_1['key1']=df_Compare_AHU_1['Point_Name'].str.contains(i).map({True:i,False:np.nan})

                    df_AHU_Point = pd.merge(df_Compare_AHU_1 ,df_Compare_Point,left_on='key1',right_on='key2', how='inner')
                    print('Merged DataFrame')
                    print(df_AHU_Point)
                    if (df_AHU_Point.empty == False):
                        list_merge_1.append(df_AHU_Point)
                        
                    
                    print ("The Script is evalauting AHUs it may take several minutes, please be patient")

                # Looping and Checking if the Notification Class is Properly Configured for AHUs 

                for x in range(0,len(list_merge_1),1):
                    print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                    print(list_merge_1[x])

                    # This Code Checks the Configuration of the Notification Class against Cde Point List
                
                    df_AHU_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                    print('df_AHU_Alm_Class')
                    print(df_AHU_Alm_Class)
                    #print(df_DAHU_Alm_Class)
                    len_list_AHU_1 = len_list_AHU_1 + len (df_AHU_Alm_Class['index'].to_list())
                    df_diff_7=  df_AHU_Alm_Class [df_AHU_Alm_Class['Alarm_Class'].astype(str).values ==df_AHU_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                    #print(df_diff_7)
                    list_wtht_diff=  df_AHU_Alm_Class['index'].to_list()
                    # For AHUs, we are evaluating these configurations : 1. Notification Class 2. Alarm Delay
                    # Adding the Notification Class Files
                    len_list_AHU_3 = len_list_AHU_3 + len (list_wtht_diff)
                    list_smlr = df_diff_7['index'].to_list()
                    list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                    #len_list_AHU_3 =len_list_AHU_3 + len (list10)
                    first_row_in_excel=2
                    highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                    list_smlr= [x+ first_row_in_excel for x in list_smlr]
                    string="E"
                    Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                    print(Cell_numbers_Difference_Highlighted)
                    Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                    print(Cell_numbers_similiar)
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid', 
                                    fgColor='FC2C03')
                    for cell in Cell_numbers_Difference_Highlighted:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path) 
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                    for cell in Cell_numbers_similiar:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path) 

                    # This Code Checks the Configuration of the Notification Class against Cde Point List

                    ## Alarm Delay - Comparing Alarm Delay between the Point List and Report
                    df_AHU_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                    len_list_AHU_1 = len_list_AHU_1 + len (df_AHU_delay_Alarm['index'].to_list())
                    df_AHU_delay_Alarm['Delay'] = df_AHU_delay_Alarm['Delay'].apply(str)
                    df_AHU_delay_Alarm['ALARM DELAY'] = df_AHU_delay_Alarm['ALARM DELAY'].apply(str)
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace('seconds', 'sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace('minutes', 'min')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace('1 minute', '60 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace('30secs', '30 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace('5secs', '5 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace('30 seconds', '30 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.str.replace(r'\b5000.0\b', '5 sec')
                

                    #df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('1 minute', '60 sec')
                    df_diff_8 = df_AHU_delay_Alarm [df_AHU_delay_Alarm['Delay'].astype(str).values ==df_AHU_delay_Alarm['ALARM DELAY'].astype(str).values]
                    print('df_Alarm_Delay')
                    print(df_diff_8)
                    list_wtht_diff=  df_AHU_delay_Alarm['index'].to_list()
                    # Adding the Delay Points
                    #len_list_AHU_1 =len_list_AHU_1 + len (list_wtht_diff)

                    list_smlr = df_diff_8['index'].to_list()
                    len_list_AHU_3 =len_list_AHU_3 + len (list_smlr)
                    list_smlr= [x+ first_row_in_excel for x in list_smlr]
                    #print(list_smlr)
                    list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                    #len_list_AHU_4=len_list_AHU_4 + len (list11)
                    #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                    first_row_in_excel=2
                    highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                    #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                    #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                    string="G"
                    Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                    Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                    for cell in Cell_numbers_Difference_Highlighted_1:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path)
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                    for cell in Cell_numbers_similiar:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path)

                print ("The Script is evalauting AHUs it may take several minutes, please be patient")

                # Highlighting rows which are not validated
                
                AHU_cons = pd.concat(list_df_AHU)
                AHU_cons_duplicate = AHU_cons[AHU_cons.index.duplicated()]
                AHU_cons_1=AHU_cons.drop_duplicates(subset = 'index')
                list_AHU_1= df_Compare_12['index'].to_list()
                list_all_AHU_found = AHU_cons_1['index'].to_list()
                list_all_AHU_not_found = [int(x) for x in list_AHU_1 if x not in list_all_AHU_found]
                first_row_in_excel=2
                highlight_rows_AHU=[x+ first_row_in_excel for x in list_all_AHU_not_found]
                with xw.App(visible=False)as app:
                    updated_wb= app.books.open(output_File_path)
                    updated_ws = updated_wb.sheets('Summary')
                    rng=updated_ws.used_range
                    print(rng.address) 
                    for row in rng.rows:
                        if row.row in highlight_rows_AHU:
                            row.color ='ffff00'
                    updated_wb.save(output_File_path) 
                global len_list_tot_AHU, len_list_wtht_diff_AHU,list_tot_AHU,list_dssmlr_AHU

                len_list_tot_AHU =0 
                len_list_wtht_diff_AHU=0
                list_dssmlr_AHU=0



                #len_list_tot_AHU = len_list_AHU_2 + len_list_AHU_4
                len_list_tot_AHU = len_list_AHU_1 
                #len_list_wtht_diff_AHU = len_list_AHU_1 + len_list_AHU_3
                len_list_wtht_diff_AHU =  len_list_AHU_3
                list_tot_AHU=[]
                list_dssmlr_AHU = []
                list_tot_AHU.append(len_list_tot_AHU)
                list_dssmlr_AHU.append(len_list_wtht_diff_AHU)
                print('list_tot')
                print(list_tot_AHU)
                print('list_dssmlr')
                print(list_dssmlr_AHU)


            else:

                print ("This site does not have AHUs")

        # This code will execuete to evaluate the configuration of CAT Sensors 

        # Iw Room 

        #df_Compare_IW = df_Alrm_Trend[((df_Alrm_Trend.Point_Name.str.contains("_Flow")) | df_Alrm_Trend.Point_Name.str.contains(" Flow")| (df_Alrm_Trend.Point_Name.str.contains("_Conductivity")))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        #df_Compare_IW. to_excel('C:\Projects\Script\IWSiemens.xlsx')

        #df_IW= df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('_Flow', case=False, regex=True) | df_Alrm_Trend['Point_Name'].str.contains('_Valve', case=False, regex=True) | df_Alrm_Trend['Point_Name'].str.contains('_CityPressure', case=False, regex=True) | df_Alrm_Trend['Point_Name'].str.contains('_SysPressure', case=False, regex=True) ]
        #df_IW.to_excel('C:\Projects\Script\Area\BeforeFilteringIW.xlsx')



        
        # Let us create an Excel File Again
        df_Alrm_Trend.to_excel('C:\Projects\Script\Area\BeforeFilteringCAT.xlsx')

        #df_CAT_Siemens = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("Cat") |df_Alrm_Trend.Point_Name.str.contains("CAT")) & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('HardwareFault')))]
        
       
        # df_CAT_Siemens = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("CatA")) | (df_Alrm_Trend.Point_Name.str.contains("CatB")) | (df_Alrm_Trend.Point_Name.str.contains("CatC")) | (df_Alrm_Trend.Point_Name.str.contains("CatD")) | (df_Alrm_Trend.Point_Name.str.contains("CatE")) | ((df_Alrm_Trend.Point_Name.str.contains("CAT") & (df_Alrm_Trend['Alarm_Extension'].str.contains('Alarms'))))]
        
        df_CAT_Siemens = df_Alrm_Trend [((df_Alrm_Trend.Point_Name.str.contains("Cat")) & (df_Alrm_Trend['Alarm_Extension'].str.contains('Alarms'))) |((df_Alrm_Trend.Point_Name.str.contains("CAT") & (df_Alrm_Trend['Alarm_Extension'].str.contains('Alarms'))))]
        
        
        df_CAT_Siemens.to_excel('C:\Projects\Script\Area\PDXCAT.xlsx')

        
        if (df_CAT_Siemens.shape[0] >= 1 ):
            print ('The Script is evaluating CAT Sensors')
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Cold Aisle Temp'):
                    idx_start =index
                    list_ids.append(idx_start)
            
            # Getting the required data for CAT Sensors from the Point List and computing the High and low limits
            
            df_Compare_final_ids = df_Compare.iloc[[list_ids[0]]]
            df_CAT =df_Compare_final_ids [['Point_Description','Point_Name','Alarm LO/HI','ALARM DELAY', 'NOTIFICATION LEVEL', 'ALARM TXT', 'DEAD BAND' ]]
            df_CAT = df_CAT.rename(columns ={df_CAT.columns[2]:'Alarm_Range'})
            df_CAT['Alarm_Range'] = df_CAT['Alarm_Range'].str.replace(r"\(.*\)","")
            df_CAT [['Alarm_Low','Alarm_High']] = df_CAT ['Alarm_Range'].str.split ( '/', expand = True, regex=False)
            df_CAT.to_excel('C:\Projects\Script\Area\CATsensors2.xlsx')
            df_CAT_1 = df_CAT
            df_CAT_1 = df_CAT_1.drop (columns = 'Point_Name')
            df_CAT_1 ['key2'] = 'CAT'
            df_CAT_Siemens['key1'] = 'CAT'

            # Merging the CAT Sensors 
        
            global len_list_tot_CAT, len_list_not_smlr_CAT
            len_list_not_smlr_CAT = 0

            # Thid will create a table to compare the total list

            df_CAT_compare = pd.merge(df_CAT_Siemens,df_CAT_1,left_on = 'key1', right_on ='key2', how = 'inner')
           

            # Computing the total number of CAT Sensors
            list_tot_CAT =  df_CAT_compare['index'].to_list()
            len_list_tot_CAT = len(list_tot_CAT)
            print('Total CAT Sensors Evaluated')
            global tot_CAT_attr
            tot_CAT_attr = 4 * len_list_tot_CAT
            print(tot_CAT_attr)

            
            
            # Comapring the time Delay
            df_CAT_compare['Delay'] = df_CAT_compare['Delay'].apply(str)
                    
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('seconds', 'sec')
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('minutes', 'min')
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('1 minute', '60 sec')
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('30secs', '30 sec')
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('30000.0', '30 sec')

            df_CAT_compare['DEAD BAND'] = df_CAT_compare['DEAD BAND'].str.replace(r"\(.*\)","")

            list_CAT=  df_CAT_Siemens['index'].to_list()
            list_CAT= [ x for x in list_CAT if ~np.isnan(x)]

            # df_CAT_compare['low_limit'] = df_CAT_compare['low_Limit'].apply(str)
            # df_CAT_compare['high_limit'] = df_CAT_compare['high_Limit'].apply(str)
            # df_CAT_compare['Dead_Band'] = df_CAT_compare['Dead_Band'].apply(str)

            df_CAT_compare['low_Limit'] = df_CAT_compare['low_Limit'].astype(str).str.split(' ').str[0]
            df_CAT_compare['high_Limit'] =df_CAT_compare['high_Limit'].astype(str).str.split(' ').str[0]
            df_CAT_compare['Dead_Band'] =df_CAT_compare['Dead_Band'].astype(str).str.split(' ').str[0]

            # This Code evaluates the Low Level Alarm Threshold 
            df_CAT_compare['Delay'] = df_CAT_compare['Delay'].apply(str)

            # Comparing Low limits for CAT Sensors
            df_CAT_Comp_Low =   df_CAT_compare [['index', 'Equipment','low_Limit','Alarm_Low' ]]
            df_CAT_Comp_Low_1 = df_CAT_Comp_Low[df_CAT_Comp_Low['low_Limit'].astype(float).values == df_CAT_Comp_Low['Alarm_Low'].astype(float).values]
            print(df_CAT_Comp_Low_1)
            list_smlr_CAT= df_CAT_Comp_Low_1['index'].to_list()
            list_smlr_CAT= [ x for x in list_smlr_CAT if ~np.isnan(x)]
           #print(list_smlr_CAT)
            list_not_smlr_CAT_Low = [int(x) for x in list_CAT if x not in list_smlr_CAT]
            len_list_not_smlr_CAT = len_list_not_smlr_CAT + len(list_not_smlr_CAT_Low)

            print('list_not_smlr_CAT')
            print (len_list_not_smlr_CAT)
            
            first_row_in_excel=2
            highlight_rows_CAT_low_mismatch=[x+ first_row_in_excel for x in list_not_smlr_CAT_Low]
            string="K"
            Cell_numbers_Difference_Highlighted_Low_CAT= ["{}{}".format(string,i) for i in highlight_rows_CAT_low_mismatch]
            print(Cell_numbers_Difference_Highlighted_Low_CAT)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Low_CAT:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # This Code evaluates the High Level Alarm Threshold 

            df_CAT_Comp_High=   df_CAT_compare [['index', 'Equipment','high_Limit','Alarm_High' ]]
            print (df_CAT_Comp_High)
            df_CAT_compare_High_1 = df_CAT_Comp_High[df_CAT_Comp_High['high_Limit'].astype(float).values == df_CAT_Comp_High['Alarm_High'].astype(float).values]
            print (df_CAT_compare_High_1)
            # Rename the variable list_smlr_CAT_1 to list_smlr_CAT_High
            list_smlr_CAT_1 = df_CAT_compare_High_1['index'].to_list()
            list_smlr_CAT_1= [ x for x in list_smlr_CAT_1 if ~np.isnan(x)]
            print(list_smlr_CAT_1)
            # Rename the variable list_not_smlr_CAT_1 to list_not_smlr_CAT_High

            list_not_smlr_CAT_High = [int(x) for x in list_CAT if x not in list_smlr_CAT_1]
            len_list_not_smlr_CAT = len_list_not_smlr_CAT+ len (list_not_smlr_CAT_High)

            print('list_not_smlr_CAT')
            print (len_list_not_smlr_CAT)
        

            first_row_in_excel=2
            highlight_rows_alarm_CAT_high_mismatch=[x+ first_row_in_excel for x in list_not_smlr_CAT_High]
            string="J"
            Cell_numbers_Difference_Highlighted_High_CAT= ["{}{}".format(string,i) for i in highlight_rows_alarm_CAT_high_mismatch]
            print(Cell_numbers_Difference_Highlighted_High_CAT)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_High_CAT:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Comparing the Dead Band between the CDE Point List and configured as shown by N4 Report 

            df_CAT_DdBnd = df_CAT_compare [['index', 'Equipment','Dead_Band','DEAD BAND' ]]
            df_CAT_DdBnd_1 = df_CAT_DdBnd[df_CAT_DdBnd['Dead_Band'].astype(float).values == df_CAT_DdBnd['DEAD BAND'].astype(float).values]
            lst_smlr_CAT_DdBnd =df_CAT_DdBnd_1 ['index'].to_list()
            lst_smlr_CAT_DdBnd_1=[ x for x in lst_smlr_CAT_DdBnd   if ~np.isnan(x)]

            # Creating a list where the deadband does not mtach 
            lst_smlr_CAT_DdBnd_2=[int(x) for x in list_CAT if x not in lst_smlr_CAT_DdBnd_1]

            # Updating the number of attributes which are mismatching 
            len_list_not_smlr_CAT = len_list_not_smlr_CAT+ len (lst_smlr_CAT_DdBnd_2)

            print('list_not_smlr_CAT')
            print (len_list_not_smlr_CAT)

            
            first_row_in_excel=2
            highlight_rows_alarm_CAT_dbnd=[x+ first_row_in_excel for x in lst_smlr_CAT_DdBnd_2]
            string="L"
            Cell_numbers_Difference_Highlighted_Dnd_CAT= ["{}{}".format(string,i) for i in highlight_rows_alarm_CAT_dbnd]
            print(Cell_numbers_Difference_Highlighted_Dnd_CAT)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Dnd_CAT:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Comparing the Notification Class of the Cold Aisle Sensors
            
            df_diff_7_CAT=  df_CAT_compare [df_CAT_compare['Alarm_Class'].astype(str).values ==df_CAT_compare ['NOTIFICATION LEVEL'].astype(str).values]

            # Highlighlighting Cells where there is a diffwnce in the Notificatiob clasas

            

            list_smlr = df_diff_7_CAT['index'].to_list()
            list_smlr= [ x for x in list_smlr if ~np.isnan(x)]

            list100 = [int(x) for x in list_CAT if x not in list_smlr]

            len_list_not_smlr_CAT = len_list_not_smlr_CAT+ len (list100)

            print('list_not_smlr_CAT')
            print (len_list_not_smlr_CAT)


            #cleanedList = [x for x in countries if x != 'nan']
            #print(list100)

            # Writing back to excel file to highlight where CAT notification class do not match 

            
            first_row_in_excel=2
            highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list100]
            list_smlr= [x+ first_row_in_excel for x in list_smlr]
            string="E"
            Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
            print(Cell_numbers_Difference_Highlighted)
            Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]


            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Alarm Delay - Comparing Alarm Delay between the Point List and Report
            df_diff_8_CAT=  df_CAT_compare [df_CAT_compare['Delay'].astype(str).values ==df_CAT_compare ['ALARM DELAY'].astype(str).values]
            df_diff_8_CAT.to_excel('C:\Projects\Script\CATDelay.xlsx')
            list_smlr_1 = df_diff_8_CAT['index'].to_list()
            list_smlr_1= [ x for x in list_smlr_1 if ~np.isnan(x)]
            list200 = [int(x) for x in list_CAT if x not in list_smlr_1]
            print('list200')
            print(list200)

            len_list_not_smlr_CAT = len_list_not_smlr_CAT+ len (list200)

            print('list_not_smlr_CAT')
            print (len_list_not_smlr_CAT)


            # Writing back to excel sheet

            first_row_in_excel=2
            highlight_rows_alarm_delay_mismatch=[x+ first_row_in_excel for x in list200]
            list_smlr_1= [x+ first_row_in_excel for x in list_smlr_1]
            string="G"
            Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay_mismatch]
            print(Cell_numbers_Difference_Highlighted_1)
            Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr_1]

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_1:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            '''
            # Creating an Exception here - writing all the rows back to white, this is because some previous codfe have made these rows yellow  

            first_row_in_excel=2
            list_all_DP_1= [x+ first_row_in_excel for x in list_all_DP]


            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_all_DP_1:
                        row.color ='ffffff'
                updated_wb.save(output_File_path)


            # Highlighting Rows for SAPDP where the comparision was not made 

            list_DP_not_comp = [int(x) for x in list_all_DP if x not in list_DP_Comp ]
            first_row_in_excel=2
            list_DP_not_comp_1= [x+ first_row_in_excel for x in list_DP_not_comp]

            print(list_DP_not_comp_1)

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_DP_not_comp_1:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)
                '''
            
            #writer = pd.ExcelWriter(output_File_path, engine="xlsxwriter")
            '''writer = pd.ExcelWriter(r 'output_File_path', engine="openpyxl", mode= 'a')
            writer.save()
            writer.close()'''

        # Computing TOTALs similiar, dissimiliar and mismatch
        
        
        #global tot_point_eval, tot_point_mtach, tot_point_mismatch

        try:

            if (df_Compare_12.empty == True) :
                len_list_tot_AHU =0 # Tracks the total number of Electrical Room Points
                len_list_wtht_diff_AHU =0
                #len_list_ER_3 =0 
        except:
                len_list_tot_AHU =0 # Tracks the total number of Electrical Room Points
                len_list_wtht_diff_AHU =0
        '''
        if(df_Compare_12.empty == True ): 
            len_list_tot_AHU =0
            len_list_wtht_diff_AHU=0
            print('len_list_tot_AHU')
            print(len_list_tot_AHU)
            print('len_list_wtht_diff_AHU')
            print(len_list_wtht_diff_AHU)
            df_notfound= pd.DataFrame()
        '''
        
        if(df_Compare_11.empty == True ): 
            len_list_DAHU_1 =0
            len_list_DAHU_3 =0 
            len_list_DAHU_2 =0 
            # df_notfound= pd.DataFrame()


        
        if (df_CAT_Siemens.empty == True):
            len_list_tot_CAT =0
            len_list_not_smlr_CAT =0
        
         # Resetting the totals to zeros for ER 

        if (df_Compare_ER.empty == True) :
            len_list_ER_1 =0 # Tracks the total number of Electrical Room Points
            len_list_ER_2 =0
            len_list_ER_3 =0 

               
        # Resetting the totals to zeros for CRAHs

        if (df_Compare_CRAH.empty == True) :
            len_list_CRAH_1 =0 # Total Number of CRAHU Points
            len_list_CRAH_2 =0
            len_list_CRAH_3 =0 # Total Number of CRAHU Points which are similiar
    
        if(df_IW_Shape.empty == True):
            len_list_IW_Smns_1 =0
            len_list_IW_Smns_3 =0
        
            
        tot_point_eval = len_list_tot_AHU + 4*len_list_tot_CAT +len_list_DAHU_1+len_list_ER_1+len_list_CRAH_1+len_list_IW_Smns_1
        tot_point_mismtach = len_list_not_smlr_CAT+ (len_list_tot_AHU -len_list_wtht_diff_AHU) + (len_list_DAHU_1 -len_list_DAHU_3 ) +(len_list_ER_1 -len_list_ER_3 )+ (len_list_CRAH_1 -len_list_CRAH_3 ) + (len_list_IW_Smns_1 -len_list_IW_Smns_3)
        tot_point_match = len_list_wtht_diff_AHU + (len_list_tot_CAT*4 -len_list_not_smlr_CAT)+ len_list_DAHU_3+len_list_ER_3 +len_list_CRAH_3+len_list_IW_Smns_3

        print('total points evaluated')
        print(tot_point_eval)
        print('total points matched')
        print(tot_point_match)
        print('total points not matched')
        print(tot_point_mismtach)
        
        df_notfound= pd.DataFrame()

        try:

            list_tot_pts=[]
            list_smlr_pts = []
            list_tot_pts.append(tot_point_eval)
            list_smlr_pts.append(tot_point_match)
            list_report = list(zip(list_tot_pts,list_smlr_pts))
            df_report = pd.DataFrame (list_report, columns = ['Total_Points','Compliant_Points'])
            df_report['Non_Compliant_Points'] = df_report['Total_Points'] - df_report['Compliant_Points']
            df_report['Percent_Compliant'] = (df_report['Compliant_Points']/df_report['Total_Points'])*100

        except:
            
            df_report = pd. DataFrame()
            # df_notfound= pd.DataFrame()
            print (' No Points are evaluated!')


        '''

        if(df_Compare_12.shape[0]>= 1):  # This Data
        
            list_tot_AHU=[]
            list_dssmlr_AHU = []
            list_tot_AHU.append(len_list_tot_AHU)
            list_dssmlr_AHU.append(len_list_wtht_diff_AHU)
            lst_report = list(zip(list_tot_AHU,list_dssmlr_AHU))
            # Creating a Data Frame fromn the lists created above
            df_report = pd.DataFrame (lst_report, columns = ['Total_Points','Compliant_Points'])
            df_report['Non_Compliant_Points'] = df_report['Total_Points'] - df_report['Compliant_Points']
            df_report['Percent_Compliant'] = (df_report['Compliant_Points']/df_report['Total_Points'])*100

            print(df_report)
            #df_report.to_excel(report_file_path)
        else :
            df_report = pd. DataFrame()
            df_notfound= pd.DataFrame()

        '''

        ''''

        df_report = pd.read_excel (report_file_path)
    df_report = pd.melt(df_report, id_vars=df_report.columns[0], value_vars=['Total', 'Mismatch'])[['variable','value']]
    df_report =df_report.set_index(df_report.columns[0])
    pd.DataFrame(df_report.value).plot.pie(y = 'value', figsize = (4,4), ylabel = ' ')
    plt.title("Alarm check out report")
    plt.savefig(plot_file_path, dpi=150)

    '''

        print ('df_report')  
        print(df_report)
        print(df_report.shape[0])
        print(type(df_report.shape[0]))
        print (df_report.shape[0] >0)
        workbook = openpyxl.load_workbook(output_File_path)
        workbook.create_sheet('Alarms_not_found')
        workbook.create_sheet('Chart')
        #df_notfound.to_excel(writer, sheet_name='Alarms_not_found') 
        workbook.save(output_File_path)

        #load workbook
        app = xw.App(visible=False)
        wb = xw.Book(output_File_path)  
        ws = wb.sheets['Alarms_not_found']

        #Update workbook at specified range
        if (df_notfound.empty  == False):
            ws.range('A2').options(index=False).value = df_notfound
        ws = wb.sheets['Chart']
        if (df_report.empty == False):
            print ("Is it Execueting??")
            ws.range('A2').options(index=False).value = df_report
        # Adding the code for cgart 

    
        #Close workbook
        wb.save()
        wb.close()
        app.quit()

        ## Creating Chart 

        #df_report_1 = df_report.T

        print (' Again df_report')
        print(df_report.shape[0] >0)

        
        if (df_report.empty== False): 
            print ('It should draw the Chart')

        try:

            if (df_report.empty == False): 
                print ('Drawing the Chart')

                df_report_1= df_report[['Compliant_Points','Non_Compliant_Points']]
                df_report_2 = df_report_1.T
                df_report_3 = df_report_2.rename(columns = {df_report_2.columns[0]:"Point_Distribution"})
                df_report_3.to_excel('C:\Projects\Script\Area\FinalReport11.xlsx')
                fig = plt.figure()
                plt.pie(df_report_3['Point_Distribution'], labels = ['Compliant_Points','Non_Compliant_Points'], colors =['Green','Red'])

                plt.savefig('C:\Projects\Script\myplot1.png', dpi=150)
                plt.savefig(chart_path)

                wb = openpyxl.load_workbook(output_File_path)
                ws=wb['Chart']
                img = openpyxl.drawing.image.Image(chart_path)
                ws.add_image(img, "A5")
                wb.save(output_File_path)
        
        except :
             print ('There is a NaN Value!')
       
        # Creating a Pie Chart
    
        sg.popup(" Done! :)")


    except RuntimeError :
        pass

    pass

def pointcheckoutSchneider(alarmFile1,alarmFile2,trendFile1,cdepointList,output_folder):

    len_list_tot_CAT_Schndr =0
    len_list_smlr_CAT_Schndr =0
    len_list_dissmlr_CAT_Schndr =0

    len_list_tot_CAT_Schndr_High =0
    len_list_smlr_CAT_Schndr_High =0
    len_list_dissmlr_CAT_Schndr_High =0

    len_list_tot_CAT_Schndr_Low =0
    len_list_smlr_CAT_Schndr_Low =0
    len_list_dissmlr_CAT_Schndr_Low=0

    len_list_tot_CAT_Schndr_Dbnd=0
    len_list_smlr_CAT_Schndr_Dbnd=0
    len_list_dissmlr_CAT_Schndr_Dbnd=0

    
    len_list_tot_CAT_Schndr_Class=0
    len_list_smlr_CAT_Schndr_Class=0
    len_list_dissmlr_CAT_Schndr_Class=0

        
    len_list_tot_CAT_Schndr_Delay=0
    len_list_smlr_CAT_Schndr_Delay=0
    len_list_dissmlr_CAT_Schndr_Delay=0

        
    global  total_point_eval_Schndr
    global  total_point_match_Schndr
    global  total_point_mismatch_Schndr

    

    #global len_list_DAHU_1_Schn
    len_list_DAHU_1_Schn = 0  # This variable tracksd the points which are being checked for DAHUs
    #global len_list_DAHU_2_Schn # This parameter tracks the points which are not Compliant
    len_list_DAHU_2_Schn = 0 
    #global len_list_DAHU_3_Schn # This parameter tracks the points which are Compliant 
    len_list_DAHU_3_Schn = 0 


    # Similiar to above for CAT Sensors

    

    # Reading Input Files,Currently DefiningtheFiePath on C Drive
    #path ='C:'
    #alarm_File1_path= os.path.join(path,'\Projects\Script'+alarmFile1)
    #alarm_File2_path= os.path.join(path,'\Projects\Script'+alarmFile2)
    #trend_File_path= os.path.join(path,'\Projects\Script'+trendFile1)
    #cde_point_list_path= os.path.join(path,'\Projects\Script'+cdepointList)
    #output_File_path= os.path.join(path,'\Projects\Script'+outputFile)

    # In the New Regions, DAHUs are being called OAHU, the point name is therefore inclusive of this string , the alarm and trend configuration for the OAHUs are otherwise same 

    try: 


        output_File_path = Path(output_folder) /"Reporttest.xlsx"
        alarm_File1_path= Path(alarmFile1)
        alarm_File2_path= Path(alarmFile2)
        trend_File_path = Path(trendFile1)
        cde_point_list_path= Path(cdepointList)
        chart_path = Path(output_folder) /"myplot2.png"
        #plot_file_path = Path(output_folder) /"myplot.png"
        #report_file_path = Path (output_folder)/"Overview_Report.xlsx"
        #plot_file_path = "./images/myplot.png"

        print(output_File_path)
        print( alarm_File1_path)
        print ( alarm_File2_path)
        print(trend_File_path)
        print(cde_point_list_path)
        # Creating Data Frames 
        # Creating Data Frames so it would be ablr to handle the different Format of the text file it generates
        # The Niagara Export Reports seem to have different Text File Format,
        # The Other Objective is to have the Consistent Column Names
        encoding_list = [ 'latin_1','utf_8', 'utf_8_sig']

        #path = 'C:'
        #path= os.path.join(path,'\Projects\Script'+'\Alarm_report_2_PDX.csv')
        #print(path)
        for encoding in encoding_list:
            worked = True
            try:
                df_Alarms_1 = pd.read_csv(alarm_File1_path, encoding=encoding)
            except:
                worked = False
            if worked:
                break
        # Add Code Later for the consistent Column Names

        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[0]: "Equipment"}) 
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[1]: "Point_Name"}) 
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[2]: "Alarm_Extension"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[3]: "Alarm_Class"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[4]: "Alarm_Type"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[5]: "Delay"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[6]: "offNormal_Text"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_1.columns)[7]: "Normal_Text"})

        # ['Equipment','Point_Name','high_Limit','low_Limit','Dead_Band','High_Limit_Text','low_Limit_Text']

        #df_Alarms_1=pd.read_csv(alarm_File1_path,error_bad_lines =False)
        df_Alarms_1['Point_Name_1'] =  df_Alarms_1['Equipment'] +'_'+ df_Alarms_1['Point_Name']
        #
        for encoding in encoding_list:
            worked = True
            try:
                df_Alarms_2 = pd.read_csv(alarm_File2_path, encoding=encoding)
            except:
                worked = False
            if worked:
                break

        # Column Names 

        df_Alarms_2 = df_Alarms_2.rename(columns={list(df_Alarms_2.columns)[0]: "Equipment"})
        df_Alarms_2 = df_Alarms_2.rename(columns={list(df_Alarms_2.columns)[1]: "Point_Name"}) 
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[2]: "high_Limit"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[3]: "low_Limit"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[4]: "Dead_Band"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[5]: "high_Limit_Text"})
        df_Alarms_1 = df_Alarms_1.rename(columns={list(df_Alarms_2.columns)[6]: "low_Limit_Text"})


        ## Need to standardize other column names

        df_Alarms_2['Point_Name_1'] =  df_Alarms_2['Equipment'] +'_'+ df_Alarms_2['Point_Name']
        df_Alarms_2 =df_Alarms_2.drop(columns =['Equipment'])

        # Merging Alarm DataFrames
        df_Alarms_All = pd.merge(df_Alarms_1, df_Alarms_2, left_on = 'Point_Name_1', right_on ='Point_Name_1', how = 'outer')
        df_Alarms_All =df_Alarms_All[['Point_Name_1','Equipment','Alarm_Extension', 'Alarm_Class','Alarm_Type', 'Delay', 'offNormal_Text', 
                                    'Normal_Text',  'Point_Name_y', 'high_Limit', 'low_Limit', 'Dead_Band','High_Limit_Text', 
                                    'low_Limit_Text']]
        #print(df_Alarms_All.head(5))
        # Combining with Trend Data
        for encoding in encoding_list:
            worked = True
            try:
                df_Trends_1 = pd.read_csv(trend_File_path, encoding=encoding)
            except:
                worked = False
            if worked:
                break
        # Standardizing the Column Names

        df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[0]: "Equipment"})
        df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[1]: "Point_Name"})
        df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[2]: "Trend_Type"}) 
        df_Trends_1 = df_Trends_1.rename(columns={list(df_Trends_1.columns)[3]: "Trend_Interval"})
        
        
        list1 = ['BooleanCov','NumericInterval','EnumCov','NumericCov', 'NumericInteval','NumericInteval_01','NumericInteval_05', 'NumericInteval_15']
        df_Trends_1 = df_Trends_1.loc[df_Trends_1['Trend_Type'].isin(list1)]
        df_Trends_1['Point_Name_1'] = df_Trends_1['Equipment'] +'_'+ df_Trends_1['Point_Name']
        # Merging Alarm and Trend Data
        df_Alrm_Trend = pd.merge(df_Alarms_All, df_Trends_1, left_on = 'Point_Name_1', right_on ='Point_Name_1', how = 'outer')
        df_Alrm_Trend= df_Alrm_Trend.drop (columns = ['Point_Name_y','Point_Name','Equipment_y'])
        df_Alrm_Trend =df_Alrm_Trend.rename(columns={"Point_Name_1": "Point_Name",'Equipment_x':'Equipment' })
        
        ## Doing Some Cleaning Based on the Errors 
        df_Alrm_Trend["high_Limit"] = df_Alrm_Trend["high_Limit"].astype(str).str.replace('Â°F','°F' )
        df_Alrm_Trend["low_Limit"] = df_Alrm_Trend["low_Limit"].astype(str).str.replace('Â°F','°F' )
        df_Alrm_Trend["Dead_Band"] = df_Alrm_Trend["Dead_Band"].astype(str).str.replace('Î','' )
        df_Alrm_Trend["Dead_Band"] = df_Alrm_Trend["Dead_Band"].astype(str).str.replace('Â°F','°F' )
        
        # Splitting the Data Frame 
        #df = df_Alrm_Trend
        #df1= df["high_Limit"].str.split('0',1)[0] 
        #print(df1)
        #sep = '...'
        #stripped = text.split(sep, 1)[0]
        ## Need to write more universal code 
        
        #Dealing with B-Formatting
        ### A commonly encountered B-Formatting is %alarmData.highlimit% and %alarmData.lowlimit%, checking for B- Formatting 
        ###to see if it is correctly formatted
        df_Alrm_Trend_1 = df_Alrm_Trend[['high_Limit','low_Limit','High_Limit_Text','low_Limit_Text' ]]
        list7=list(df_Alrm_Trend_1['low_Limit_Text'].str.contains('%alarmData.lowLimit%') == True)
        list8 = list(enumerate(list7))
        list9=[]
        for i in range (len(list8)):
            if(list8[i][1] == True):
                list9.append(list8[i][0])
        first_row_in_excel=2
        list10=[x+ first_row_in_excel for x in list9]
        string="N"
        list11= ["{}{}".format(string,i) for i in list10]
        ## Highlighting the Cell Containing B-Formatting as %alarmData.highLimit%, highlighting the rows 
        ##which contains this B-Formatted String
        list17=list(df_Alrm_Trend_1['High_Limit_Text'].str.contains('%alarmData.highLimit%') == True)
        list18 = list(enumerate(list7))
        list19=[]
        for i in range (len(list18)):
            if(list18[i][1] == True):
                list19.append(list8[i][0])
        first_row_in_excel=2
        list20=[x+ first_row_in_excel for x in list19]
        string="M"
        list21= ["{}{}".format(string,i) for i in list20]
        ## Transfering combined data to Excel
        writer = pd.ExcelWriter(output_File_path, engine="xlsxwriter")
        df_Alrm_Trend.to_excel(writer, sheet_name='Summary')
        writer.save()
        writer.close()
        ## Highlighting the Cells Containing B- formatting of %alarmData.lowLimit%
        wb = openpyxl.load_workbook(output_File_path)
        ws = wb['Summary'] #Name of the working sheet
        fill_cell1 = PatternFill(patternType='solid', fgColor='ffff00')
        for cell in list11:
            ws[cell].fill = fill_cell1
        wb.save(output_File_path)
        ## Highlighting the Cells Containing B- formatting of %alarmData.highLimit%
        wb = openpyxl.load_workbook(output_File_path)
        ws = wb['Summary'] #Name of the working sheet
        fill_cell1 = PatternFill(patternType='solid', fgColor='ffff00')
        for cell in list21:
            ws[cell].fill = fill_cell1
        wb.save(output_File_path)
        ## Adding Functionality of Uploading CDE Point List that can be used to compare Alarms
        ##importing Point List and renaming Columns

        # Importing the CDE Point

        df_Compare = pd.read_excel(cde_point_list_path,sheet_name ='BMS Points',skiprows=[0])
        df_Compare = df_Compare.rename(columns ={"FUNCTION":'Point_Description',df_Compare.columns[1]:'Point_Name', "DEFINITION":"Trend_interval",df_Compare.columns[17]:"Alarm",df_Compare.columns[18]:"Alarm LO/HI"}).drop(columns=['JITDC','OPTDC','Site Specific','Relinquish Default','NOTES'])
        df_Compare_1 = df_Compare[['Point_Description','Point_Name']]
        ## Splitting Point_Name in the DataFrame of the Point List
        df_Compare_2 =df_Compare_1['Point_Name'].str.split(r'\[(.*?)\]', expand = True).astype(object).mask(lambda x: x.isna(), None)
        df_Alrm_Trend = df_Alrm_Trend.reset_index()
        print('df_Alrm_Trend')
        print(df_Alrm_Trend)
        # df_Alrm_Trend.to_excel('C:\Projects\Script\AllAlarms.xlsx')
        # df_Alrm_Trend.to_excel('C:\Projects\Script\N4Script\Script_Demo\Test4\Alarms.xlsx')

        df_Compare_11 = df_Alrm_Trend[df_Alrm_Trend.Point_Name.str.contains('DAHU') | (df_Alrm_Trend.Point_Name.str.contains('OAHU'))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))] 
        print('df_DAHU_DataFrame')
        print(df_Compare_11)
        #df_Compare_11.to_excel('C:\Users\gUppal\Documents\Script\Dahudata.xlsx')
        df_Compare_11.to_excel('C:\Projects\Script\DAHUDATA.xlsx')
        print (" The Script is Evaluating AHUs !")
        print (' Is it Empty?')
        print (df_Compare_11.empty)

        if (df_Compare_11.empty == False):

            print (' The Script is evaluating DAHUs, please be patient, it may take several minutes !')
        
            data1 = [{'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG1][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG1][NO][FLW]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG2][NO][FLW]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG1][NO][FLW]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG2][NO][FLW]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},  
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG2][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG1][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG2][FLOW][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[PWR][UPS][STATUS]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High' }, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][MODE][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SAT][STPT][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SF][SPD][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][LOWER][ALM]','ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'Medium'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][UPPER][ALM]','ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[ECFAN][FAIL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][CMD]','ALARM DELAY' :'30 secs','NOTIFICATION LEVEL' : 'Medium'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][HIGH][WATER][ALM]', 'ALARM DELAY' :'5 sec','NOTIFICATION LEVEL' : 'High' },
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][LOW][WATER][ALM]', 'ALARM DELAY' :'5 sec','NOTIFICATION LEVEL' : 'High' },
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][LOWER][STS]', 'ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][UPPER][STS]', 'ALARM DELAY' :'2 minute', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FILDP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[PREFILTER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FINALFILTER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][LOWER][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][UPPER][ALM]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][HIGH][ALM]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][LOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T1]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T2]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT1][FLT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT2][FLT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAH7][FLT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAH1][FLT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAH2][FLT]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP][LO][LVL]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
             {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, ]
            
        
        
            df_dict = pd.DataFrame.from_dict(data1, orient='columns')
            df_dict =df_dict['Point_Name'].str.split(r'\[(.*?)\]', expand = True).astype(object).mask(lambda x: x.isna(), None)
        

            ##Creating a Data Frame to start comparing the DAHU Points in combined Report with CDE Point List 
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'DAHU'):
                    idx_start =index
                    list_ids.append(idx_start)
                elif (df_Compare.loc[index,'Point_Description'] == 'DAHU - Evap'):
                    idx_start =index
                    list_ids.append(idx_start)

            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            # Creating a Dictionary for DAHU Points
            df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            df_Compare_8  = pd.concat([df_Compare_8, df_dict],axis=0)
            print(df_Compare_8)
            df_Compare_8.fillna("",inplace=True)
            df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
            df_Compare_9['new_col'] = 'DAHU'
            '''
            for index, rows in df_Compare_8.iterrows():
                df_Compare_9.loc[index,'new_col'] = df_Compare_9.loc[list_ids[0],'new_col'] 
                df_Compare_10 = df_Compare_9.loc[list_ids[0]+1:list_ids[1], ]
            '''
            dist1 = {};
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_9.iterrows():
                key = df_Compare_9.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_9.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
        ## Creating an Index in Consolidated Alarm Trend Data Frame to track rows
        #df_Alrm_Trend = df_Alrm_Trend.reset_index()
        
        ## Filtering the Consolidated Alarm Data Frame for DAHU key in the Schneider Report generated
        


            #df_CAT_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('CAT') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            df_Compare_11.to_excel('C:\Projects\Script\SchniderDAHU.xlsx')
            #print(df_Compare_11)
            ## This Line of Code is Simens Specific, it is because Siemens uses both the Small and Cap Letter 
            ## for Point Name
            ##Also need to add rules to delete the numbers that come in between the alphabets
            df_Compare_11 = df_Compare_11.assign(Point_Name=df_Compare_11.Point_Name.str.upper())
            df = df_Compare_11
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df[4] = df[4].fillna('').astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            # Creating a List of Dictionaries to iterate over
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            # Creating the list of Data Frames
            list_df_DAHU=[]

            list_pts_1=[]
            for lst in list_pts:
                if not ((len(lst) == 1) & (lst[0] =='OAD')) |((len(lst) == 1) & (lst[0] =='BYPD')) | ((len(lst) == 1) & (lst[0] =='RAD')) |(((len(lst) == 1) & (lst[0] =='MAT') )|((len(lst) == 1) & (lst[0] =='SS')))|((len(lst) == 1) & (lst[0] =='EVAP'))|((len(lst) == 1) & (lst[0] =='SAT')):
                    list_pts_1.append(lst)
        
            print(list_pts_1)

            df.to_excel('C:\Projects\Script\SchniderData.xlsx')

            for lst in list_pts_1:
                list1 = ' '.join(lst)
                df_DAHU=df[df.Point_Name.str.contains(list1)]
                df_DAHU['Point_Name'] = df_DAHU['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_DAHU)
                list_df_DAHU.append(df_DAHU)
        # Creating Keys to Compare the Data Frames

        # First for the CDE Point List
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_Point.to_excel('C:\Projects\Script\cdedahu.xlsx')
            # Creating a key to Match in Siemens Report Generated
            list_merge_1=[]

            for x in range(0,len(list_df_DAHU),1):
                df_DAHU = list_df_DAHU[x]
                if df_DAHU.empty:
                    continue
                print('Raw')
                print(df_DAHU[['Point_Name']])

                if df_DAHU['Point_Name'].str.contains('DAHU').any():
                    print("DAHU IS THERE")
                    df_DAHU['key1'] = df_DAHU['Point_Name'].str.replace('DAHU', '')
                    # df_DAHU.loc[df_DAHU['Point_Name'].str.contains('DAHU'), 'Point_Name'] = df_DAHU['Point_Name'].str.replace('DAHU', '')
                
                elif df_DAHU['Point_Name'].str.contains('OAHUA').any():
                    print("OAHU IS THERE")
                    df_DAHU['key1'] = df_DAHU['Point_Name'].str.replace('OAHUA', '')
                
                


                    # df_DAHU.loc[df_DAHU['Point_Name'].str.contains('OAHU'), 'Point_Name'] = df_DAHU['Point_Name'].str.replace('OAHU', '')

                #df_DAHU['key1']=df_DAHU['Point_Name'].str.replace('DAHU','')
                #df_DAHU['key1']=df_DAHU['Point_Name'].str.replace('OAHUA','')
                print(df_DAHU)
                df_DAHU_Point = pd.merge(df_DAHU ,df_Compare_Point,left_on='key1',right_on='key2', how='inner')
                if (df_DAHU_Point.empty == False):
                     
                     list_merge_1.append(df_DAHU_Point)
                     print(df_DAHU_Point[['key1','key2']])
                     print(df_DAHU_Point)


            global  len_list_tot_Class_Schneider_DAHU
            global  len_list_tot_Delay_Schneider_DAHU
            global  len_list_smlr_Class_Schneider_DAHU
            global  len_list_smlr_Class_Schneider_Delay
            global  len_list_dissmlr_Class_Schneider_DAHU
            global  len_list_dissmlr_Delay_Schneider_DAHU
            global  len_list_tot_Text_Schneider
            global  len_list_smlr_Text_Schnneider_DAHU 
            global  len_list_dissmlr_Schneider_Text

            
            len_list_tot_Class_Schneider_DAHU =0
            len_list_tot_Delay_Schneider_DAHU =0
            len_list_smlr_Class_Schneider_DAHU =0
            len_list_smlr_Class_Schneider_Delay =0
            len_list_dissmlr_Class_Schneider_DAHU =0
            len_list_dissmlr_Delay_Schneider_DAHU=0
            len_list_tot_Text_Schneider =0 
            len_list_smlr_Text_Schnneider_DAHU =0
            len_list_dissmlr_Schneider_Text =0
            
            for x in range(0,len(list_merge_1),1):

                print('Merged DataFrame')
                print(list_merge_1[x])
                df_DAHU_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                print(df_DAHU_Alm_Class)

                df_diff_7=  df_DAHU_Alm_Class [df_DAHU_Alm_Class['Alarm_Class'].astype(str).values ==df_DAHU_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                #print(df_diff_7)
                list_wtht_diff=  df_DAHU_Alm_Class['index'].to_list()
                len_list_tot_Class_Schneider_DAHU = len_list_tot_Class_Schneider_DAHU + len (list_wtht_diff)
                print('len_list_tot_Class_Schneider_DAHU')
                print(len_list_tot_Class_Schneider_DAHU)
            
                list_smlr = df_diff_7['index'].to_list()
                len_list_smlr_Class_Schneider_DAHU = len_list_smlr_Class_Schneider_DAHU +  len (list_smlr)
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('len_list_smlr_Class_Schneider_DAHU')
                print(len_list_smlr_Class_Schneider_DAHU)

                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                len_list_dissmlr_Class_Schneider_DAHU = len_list_dissmlr_Class_Schneider_DAHU + len (list10)
            
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                #print(list_pts)
                ## Alarm Delay - Comparing Alarm Delay between the Point List and Report
                df_DAHU_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_DAHU_delay_Alarm['Delay'] = df_DAHU_delay_Alarm['Delay'].apply(str)
                df_DAHU_delay_Alarm['ALARM DELAY'] = df_DAHU_delay_Alarm['ALARM DELAY'].apply(str)
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('seconds', 'sec')
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('minutes', 'min')
                df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('1 minute', '60 sec')
                #df_DAHU_delay_Alarm['Delay'] =  df_DAHU_delay_Alarm.Delay.str.replace('1 minute', '60 sec')
                df_diff_8 = df_DAHU_delay_Alarm [df_DAHU_delay_Alarm['Delay'].astype(str).values ==df_DAHU_delay_Alarm['ALARM DELAY'].astype(str).values]
                list_wtht_diff=  df_DAHU_delay_Alarm['index'].to_list()
                len_list_tot_Delay_Schneider_DAHU = len_list_tot_Delay_Schneider_DAHU + len (list_wtht_diff)
                list_smlr = df_diff_8['index'].to_list()
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                len_list_smlr_Class_Schneider_Delay = len_list_smlr_Class_Schneider_Delay + len (list_smlr)
                #print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                len_list_dissmlr_Delay_Schneider_DAHU = len_list_dissmlr_Delay_Schneider_DAHU + len (list11)
                #Adding Code to Create the Chart showing the Total number of Alarms 
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                ## Alarm Text - Comparing Alarm Delay between the Point List and Report    
                df_Compare_text = list_merge_1[x][['index','Point_Name_x','Point_Name_y','ALARM TXT','offNormal_Text']]
    
                df_diff_9 = df_Compare_text [df_Compare_text['ALARM TXT'].astype(str).values ==df_Compare_text ['offNormal_Text'].astype(str).values]
                list_wtht_diff=  df_Compare_text['index'].to_list()
                len_list_tot_Text_Schneider = len_list_tot_Text_Schneider + len(list_wtht_diff)
                
                list_smlr = df_diff_9['index'].to_list()
                list12 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                len_list_dissmlr_Schneider_Text = len_list_dissmlr_Schneider_Text + len(list12)
                first_row_in_excel=2
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]

                len_list_smlr_Text_Schnneider_DAHU = len_list_smlr_Text_Schnneider_DAHU + len(list_smlr)
                highlight_rows_alarm_text=[x+ first_row_in_excel for x in list12]
                string="H"
                Cell_numbers_Difference_Highlighted_2= ["{}{}".format(string,i) for i in highlight_rows_alarm_text]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_2:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                ## High Low Limit Test  - Not Required for DAHU 
                ## Checking for Trends 
                df_Compare_Eqpt_Trnd = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Trend_Interval','Trend_interval']]
                df_Compare_Eqpt_Trnd['Trend_Interval']= df_Compare_Eqpt_Trnd['Trend_Interval'].str.replace('irregular', 'COV')
                df_diff_10 = df_Compare_Eqpt_Trnd [df_Compare_Eqpt_Trnd['Trend_Interval'].astype(str).values ==df_Compare_Eqpt_Trnd ['Trend_interval'].astype(str).values]
                list_wtht_diff=  df_Compare_Eqpt_Trnd['index'].to_list()
                list_smlr = df_diff_10['index'].to_list()
                first_row_in_excel=2
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                list13 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                highlight_rows_alarm_trend=[x+ first_row_in_excel for x in list12]
                string="P"
                Cell_numbers_Difference_Highlighted_3= ["{}{}".format(string,i) for i in highlight_rows_alarm_trend]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_3:
                    ws[cell].fill = fill_cell1
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb.save(output_File_path)
                # Highlighting Rows Which were not Validated
            DAHU_cons = pd.concat(list_df_DAHU)
            DAHU_cons_duplicate = DAHU_cons[DAHU_cons.index.duplicated()]
            DAHU_cons_1=DAHU_cons.drop_duplicates(subset = 'index')
            list_DAHU_1= df_Compare_11['index'].to_list()
            list_all_DAHU_found = DAHU_cons_1['index'].to_list()
            list_all_DAHU_not_found = [int(x) for x in list_DAHU_1 if x not in list_all_DAHU_found]
            first_row_in_excel=2
            highlight_rows_DAHU=[x+ first_row_in_excel for x in list_all_DAHU_not_found]
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_DAHU:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)
            
            len_list_DAHU_1_Schn = len_list_DAHU_1_Schn + len_list_tot_Class_Schneider_DAHU + len_list_tot_Delay_Schneider_DAHU + len_list_tot_Text_Schneider
            len_list_DAHU_2_Schn = len_list_DAHU_2_Schn + len_list_smlr_Class_Schneider_DAHU +  len_list_smlr_Class_Schneider_Delay + len_list_smlr_Text_Schnneider_DAHU
            len_list_DAHU_3_Schn = len_list_DAHU_3_Schn + len_list_dissmlr_Class_Schneider_DAHU + len_list_dissmlr_Delay_Schneider_DAHU + len_list_dissmlr_Schneider_Text  
            print ('Total Number of DAHU Points Evaluated : ')
            print (len_list_DAHU_1_Schn)
            print ('Total Number of DAHU Points Compliant : ')
            print (len_list_DAHU_2_Schn)
            print ('Total Number of DAHU Points Non-Compliant : ')
            print (len_list_DAHU_3_Schn)

        
        else :
            print (' The Script does not have any DAHUs !')
        
        # Checking AHUs JITDC

        df_Compare_ER = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("D1E1") |df_Alrm_Trend.Point_Name.str.contains("D1E2"))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        df_Compare_ER.to_excel('C:\Projects\Script\ElectricRoom.xlsx') # This is a Check Point Only and will not be part of the final implementation  
        if (df_Compare_ER.shape[0] >=1):
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Electric Rm / Catcher Rm  (Area Controller)'):
                    idx_start =index
                    list_ids.append(idx_start)
                
            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            print(df_Compare_8)
            df_Compare_8.fillna("",inplace=True)
            df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
            df_Compare_9['new_col'] = 'ER'
            df_Compare_9.to_excel('C:\Projects\Script\ERSchneider.xlsx')
            # Create a Dictionary 
            print ("The Script is evalauting Electric Room it may take several minutes, please be patient")
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_9.iterrows():
                key = df_Compare_9.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_9.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
                print ("The Script is evalauting AHUs it may take several minutes, please be patient")
            df = df_Compare_ER
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.assign(Point_Name=df.Point_Name.str.upper())
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\df_ER.xlsx') # This is a checkpoint only and not part of the final code 
            print(df)
            #df[4] = df[4].fillna('').astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            # Need to look at other Siemens sites to check if the above code is required and create a condition accordingly 
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            # Creating a List of Dictionaries to iterate over
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            print(list_pts)
            list_df_ER=[]
            for lst in list_pts:
                list1 = ' '.join(lst)
                df_ER=df[df.Point_Name.str.contains(list1)]
                df_ER['Point_Name'] = df_ER['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_ER)
                list_df_ER.append(df_ER)
             # First for the CDE Point List, create a key that will be matched to the Database 
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_Point.to_excel('C:\Projects\Script\df_Compare_Point_ER.xlsx')
            list_merge_1=[]
            for x in range(0,len(list_df_ER),1):
                df_ER = list_df_ER[x]
                if df_ER.empty:
                    continue
                print('Raw')
                #print(df_ER[['Point_Name']])
                df_ER['key1'] = df_ER ['Point_Name']
                print(df_ER)
                merged_dfs =[]
                for index, row in df_Compare_Point.iterrows():
                    filtered_df = df_ER[df_ER['key1'].str.contains(row['key2'], case=False, regex=False)]
                    if not filtered_df.empty:
                        #print('What are we merging')
                        #print(df_Compare_Point.loc[[index]])
                        df_matched = df_Compare_Point.loc[[index]]
                        df_matched = df_matched.reset_index(drop = True)
                        #print('df_matched')
                        #print(df_matched)
                        df_matched['match1']= ['dummy']
                        df_ER.index = df_ER.index.astype(str)
                        df_ER['match2'] = ['dummy'] * len(df_ER)
                        merged_df = pd.merge(df_ER, df_matched, left_on='match2', right_on='match1', how = 'inner')
                        print(merged_df)
                        df_ER_Point = merged_df
                        df_ER_Point.to_excel('C:\Projects\Script\SchneiderER.xlsx')
                        print(df_ER_Point[['key1','key2']])
                        list_merge_1.append(df_ER_Point)
                        print('The Length of List is')
                        print (len(list_merge_1))
            
            global len_list_ER_1,len_list_ER_2,len_list_ER_3
            len_list_ER_1 =0 # Tracks the total number of Electrical Room Points
            len_list_ER_2 =0
            len_list_ER_3 =0  # Traks the number of Electrical Room Points which are similiar 


            for x in range(0,len(list_merge_1),1):
                print ('The Length of List is')
                print (len(list_merge_1))
                print ("The Script is evalauting Electric Rom it may take several minutes, please be patient")
                print(list_merge_1[x])
                list_merge_1[x].to_excel('C:\Projects\Script\Test44.xlsx' )
                df_ER_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                #df_ER_Alm_Class = list_merge_1[x][['index','Point_Name','Alarm_Class','NOTIFICATION LEVEL']]
                print('df_ER_Alm_Class')
                print(df_ER_Alm_Class)
                len_list_ER_1 = len_list_ER_1 + len (df_ER_Alm_Class['index'].to_list())
                df_diff_7=  df_ER_Alm_Class [df_ER_Alm_Class['Alarm_Class'].astype(str).values ==df_ER_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                list_wtht_diff=  df_ER_Alm_Class['index'].to_list()
                len_list_ER_3 = len_list_ER_3 + len (list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print(list10)
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

                df_ER_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_ER_delay_Alarm['Delay'] = df_ER_delay_Alarm['Delay'].apply(str)
                df_ER_delay_Alarm['ALARM DELAY'] = df_ER_delay_Alarm['ALARM DELAY'].apply(str)
                print('df_ER_delay_Alarm')
                print(df_ER_delay_Alarm)
                print(df_ER_delay_Alarm.dtypes)
                len_list_ER_1 = len_list_ER_1 + len (df_ER_delay_Alarm['index'].to_list())
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('1 minute', '60 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('0.0', 'nan')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('30000', '30 sec')
                df_ER_delay_Alarm['Delay'] =  df_ER_delay_Alarm.Delay.astype(str).str.replace('10000', '10 sec')
                
                
                df_diff_8 = df_ER_delay_Alarm [df_ER_delay_Alarm['Delay'].astype(str).values ==df_ER_delay_Alarm['ALARM DELAY'].astype(str).values]
                list_wtht_diff=  df_ER_delay_Alarm['index'].to_list()

                list_smlr = df_diff_8['index'].to_list()
                print('similiar_Delay')
                len_list_ER_3 =len_list_ER_3 + len (list_smlr)
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('Not_similiar_Delay')
                print(list11)
                #len_list_AHU_4=len_list_AHU_4 + len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
            print ("The Script is evalauting ERs it may take several minutes, please be patient")
            ER_cons = pd.concat(list_df_ER)
            ER_cons_duplicate = ER_cons[ER_cons.index.duplicated()]
            ER_cons_1=ER_cons.drop_duplicates(subset = 'index')
            list_ER_1= df_Compare_ER['index'].to_list()
            list_all_ER_found = ER_cons_1['index'].to_list()
            list_all_ER_not_found = [int(x) for x in list_ER_1 if x not in list_all_ER_found]
            first_row_in_excel=2
            highlight_rows_ER=[x+ first_row_in_excel for x in list_all_ER_not_found]
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_ER:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)     




     # Evaluating CRAH

        df_Compare_CRAH = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("CRAH")) & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        df_Compare_CRAH.to_excel('C:\Projects\Script\CRAH.xlsx')
        if (df_Compare_CRAH.shape[0] >=1):
            list_ids=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'CRAH'):
                    idx_start =index
                    list_ids.append(idx_start)
            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
            print(df_Compare_8)
            df_Compare_8.fillna("",inplace=True)
            df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
            df_Compare_9['new_col'] = 'CRAH'
            df_Compare_9.to_excel('C:\Projects\Script\CRAHSchneider.xlsx')
            # Create a Dictionary 
            print ("The Script is evalauting Electric Room it may take several minutes, please be patient")
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_9.iterrows():
                key = df_Compare_9.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_9.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            print(list2)

            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
                print ("The Script is evalauting CRAHUs it may take several minutes, please be patient")
            df = df_Compare_CRAH
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.assign(Point_Name=df.Point_Name.str.upper())
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\df_CRAH.xlsx') # This is a checkpoint only and not part of the final codE
            print(df)
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            df.to_excel('C:\Projects\Script\df_CRAH_1.xlsx') # This is a checkpoint only and not part of the final codE
            # Creating a List of Dictionaries to iterate over
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            print(list_pts)
            list_df_CRAH=[]
            for lst in list_pts:
                list1 = ' '.join(lst)
                df_CRAH=df[df.Point_Name.str.contains(list1)]
                df_CRAH['Point_Name'] = df_CRAH['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_CRAH)
                list_df_CRAH.append(df_CRAH)
            # First for the CDE Point List, create a key that will be matched to the Database
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_Point.to_excel('C:\Projects\Script\df_Compare_Point_CRAH.xlsx')
            list_merge_1=[]
            for x in range(0,len(list_df_CRAH),1):
                df_CRAH = list_df_CRAH[x]
                if df_CRAH.empty:
                    continue
                print('Raw')
                #print(df_ER[['Point_Name']])
                df_CRAH['key1'] = df_CRAH ['Point_Name']
                print(df_CRAH)
                merged_dfs =[]
                for index, row in df_Compare_Point.iterrows():
                    filtered_df = df_CRAH[df_CRAH['key1'].str.contains(row['key2'], case=False, regex=False)]
                    if not filtered_df.empty:
                        #print('What are we merging')
                        #print(df_Compare_Point.loc[[index]])
                        df_matched = df_Compare_Point.loc[[index]]
                        df_matched = df_matched.reset_index(drop = True)
                        #print('df_matched')
                        #print(df_matched)
                        df_matched['match1']= ['dummy']
                        df_CRAH.index = df_CRAH.index.astype(str)
                        df_CRAH['match2'] = ['dummy'] * len(df_CRAH)
                        merged_df = pd.merge(df_CRAH, df_matched, left_on='match2', right_on='match1', how = 'inner')
                        print(merged_df)
                        df_CRAH_Point = merged_df
                        df_CRAH_Point.to_excel('C:\Projects\Script\SchneiderCRAH.xlsx')
                        print(df_CRAH_Point[['key1','key2']])
                        list_merge_1.append(df_CRAH_Point)
                        print('The Length of List is')
                        print (len(list_merge_1)) 
            
            global len_list_CRAH_1,len_list_CRAH_2,len_list_CRAH_3
            len_list_CRAH_1 =0 # Total Number of CRAHU Points
            len_list_CRAH_2 =0
            len_list_CRAH_3 =0 # Total Number of CRAHU Points which are similiar

            for x in range(0,len(list_merge_1),1):
                print ('The Length of List is')
                print (len(list_merge_1))
                print ("The Script is evalauting Electric Rom it may take several minutes, please be patient")
                print(list_merge_1[x])
                list_merge_1[x].to_excel('C:\Projects\Script\TestCRAH.xlsx' )
                df_CRAH_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                #df_ER_Alm_Class = list_merge_1[x][['index','Point_Name','Alarm_Class','NOTIFICATION LEVEL']]
                print('df_ER_Alm_Class')
                print(df_CRAH_Alm_Class)
                len_list_CRAH_1 = len_list_CRAH_1 + len (df_CRAH_Alm_Class['index'].to_list())
                df_diff_7=  df_CRAH_Alm_Class [df_CRAH_Alm_Class['Alarm_Class'].astype(str).values ==df_CRAH_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                list_wtht_diff=  df_CRAH_Alm_Class['index'].to_list()
                len_list_CRAH_3 = len_list_CRAH_3 + len (list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print(list10)
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

                df_CRAH_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_CRAH_delay_Alarm['Delay'] = df_CRAH_delay_Alarm['Delay'].apply(str)
                df_CRAH_delay_Alarm['ALARM DELAY'] = df_CRAH_delay_Alarm['ALARM DELAY'].apply(str)
                print('df_CRAH_delay_Alarm')
                print(df_CRAH_delay_Alarm)
                print(df_CRAH_delay_Alarm.dtypes)
                len_list_CRAH_1 = len_list_CRAH_1 + len (df_CRAH_delay_Alarm['index'].to_list())
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('1 minute', '60 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('0.0', 'nan')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('30000', '30 sec')
                df_CRAH_delay_Alarm['Delay'] =  df_CRAH_delay_Alarm.Delay.astype(str).str.replace('10000', '10 sec')

                df_CRAH_delay_Alarm.loc[(df_CRAH_delay_Alarm['Delay'] == '0.0') & (df_CRAH_delay_Alarm['ALARM DELAY'] == 'nan'), 'Delay'] = 'nan'
                #df_CRAH_delay_Alarm['ALARM DELAY'] = np.where((df_CRAH_delay_Alarm['Delay'] == 0.0) & (df_CRAH_delay_Alarm['ALARM DELAY'].isna()), 0.0, df_CRAH_delay_Alarm['ALARM DELAY'])

                df_diff_8 = df_CRAH_delay_Alarm [df_CRAH_delay_Alarm['Delay'].astype(str).values ==df_CRAH_delay_Alarm['ALARM DELAY'].astype(str).values]
                list_wtht_diff=  df_CRAH_delay_Alarm['index'].to_list()

                list_smlr = df_diff_8['index'].to_list()
                print('similiar_Delay')
                len_list_CRAH_3 =len_list_CRAH_3 + len (list_smlr)
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('Not_similiar_Delay')
                print(list11)
                #len_list_AHU_4=len_list_AHU_4 + len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
            print ("The Script is evalauting ERs it may take several minutes, please be patient")
            CRAHU_cons = pd.concat(list_df_CRAH)
            CRAHU_cons_duplicate = CRAHU_cons[CRAHU_cons.index.duplicated()]
            CRAHU_cons_1=CRAHU_cons.drop_duplicates(subset = 'index')
            list_CRAHU_1= df_Compare_CRAH['index'].to_list()
            list_all_CRAHU_found = CRAHU_cons_1['index'].to_list()
            list_all_CRAHU_not_found = [int(x) for x in list_CRAHU_1 if x not in list_all_CRAHU_found]
            first_row_in_excel=2
            highlight_rows_CRAHU=[x+ first_row_in_excel for x in list_all_CRAHU_not_found]
            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in highlight_rows_CRAHU:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)     



        if ((df_Compare_11.empty == True) & (df_Compare_ER.empty == True)) :
            global df_Compare_12
            df_Compare_12 = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("AHU") |df_Alrm_Trend.Point_Name.str.contains("Ahu"))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            print ('df_Compare_12')
            print(df_Compare_12)
            df_Compare_12.to_excel('C:\Projects\Script\SchneiderAHU.xlsx')
            
            # This code will execute only if the Data base has AHUs or OAHUs
            if(df_Compare_12.shape[0]>= 1):


                print ("The Script is evalauting AHUs it may take several minutes, please be patient")

                global len_list_AHU_1
                len_list_AHU_1 =0  # This parameter tracks the points which are being checked 
                global len_list_AHU_2 # Not used for now
                len_list_AHU_2 =0
                global len_list_AHU_3 # This parameter tracks the points which are not Compliant
                len_list_AHU_3 =0 
                global len_list_AHU_4 # Not used for now 
                len_list_AHU_4=0

                # Adding the code so nothing needs to be added so no manual addition is required 

                data1 = [{'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG1][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG2][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG1][FLOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG2][FLOW][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[PWR][UPS][STATUS]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High' }, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][MODE][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SAT][STPT][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SF][SPD][ALM]','ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][LOWER][ALM]','ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'Medium'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][UPPER][ALM]','ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[ECFAN][FAIL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][CMD]','ALARM DELAY' :'30 secs','NOTIFICATION LEVEL' : 'Medium'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][HIGH][WATER][ALM]', 'ALARM DELAY' :'5 sec','NOTIFICATION LEVEL' : 'High' },
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][LOW][WATER][ALM]', 'ALARM DELAY' :'5 sec','NOTIFICATION LEVEL' : 'High' },
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][LOWER][STS]', 'ALARM DELAY' :'60 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][UPPER][STS]', 'ALARM DELAY' :'2 minute', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FILDP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[PREFILTER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[FINALFILTER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][LOWER][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][UPPER][ALM]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][HIGH][ALM]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][LOW][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'Medium'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'}, 
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][HI][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][LOW][LVL][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][PMP][ALM]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T1]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T2]', 'ALARM DELAY' :'5 sec', 'NOTIFICATION LEVEL' : 'High'},
            {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT]', 'ALARM DELAY' :'30 sec', 'NOTIFICATION LEVEL' : 'High'},
            ]
            
            
                df_dict = pd.DataFrame.from_dict(data1, orient='columns')
                df_dict =df_dict['Point_Name'].str.split(r'\[(.*?)\]', expand = True).astype(object).mask(lambda x: x.isna(), None)

                
                # This Code looks at the CDE Point List and will Execuete Only if there is a 'AHU' in the Database

                list_ids=[]
                for index,rows in df_Compare.iterrows():
                    if (df_Compare.loc[index,'Point_Description'] == 'AHU'):
                        idx_start =index
                        list_ids.append(idx_start)

                df_Compare_final_ids = df_Compare.iloc[list_ids[0]:]
                df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
                list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
                df_Compare_8 =df_Compare_2.loc[list_ids[0]:list_ids[1], :]
                df_Compare_8  = pd.concat([df_Compare_8, df_dict],axis=0)
                print(df_Compare_8)
                df_Compare_8.fillna("",inplace=True)
                df_Compare_9 = df_Compare_8.assign(new_col=df_Compare_8[3].str.replace('##',''))
                df_Compare_9['new_col'] = 'AHU'
                df_Compare_9.to_excel('C:\Projects\Script\AHUSchneider.xlsx') # This is a Check Point Only and will not be part of the final implementatio  
                # Create a Dictionary 
                print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
                list1=[]
                list2=[]
                for index,rows in df_Compare_9.iterrows():
                    key = df_Compare_9.loc[index,'new_col']
                    for x in list_val:
                        list1.append(df_Compare_9.loc[index,x])
                    value =list1
                    dicT= {key : value} 
                    list1=[]
                    list2.append(dicT)
                list102=[]
                list101=[]
                for index in range(len(list2)):
                    for i,x in enumerate (list2[index][key]):
                        if (x == ''):
                            if(i <= len(list2[index][key])):
                                list101.append(i)
                    list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                    list101=[]
                    print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                df = df_Compare_12
                df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
                df = df.assign(Point_Name=df.Point_Name.str.upper())
                df = df.join(df.Point_Name.str.split(r' ',expand=True))
                df.to_excel('C:\Projects\Script\df.xlsx') # This is a checkpoint only and not part of the final code 
                print(df)
                #df[4] = df[4].fillna('').astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
                # Need to look at other Siemens sites to check if the above code is required and create a condition accordingly 
                df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
                # Creating a List of Dictionaries to iterate over
                list_pts=[]
                for dicts in list2:
                    for eqpt,points in dicts.items():
                        list_pts.append(points) 
                
                list_pts_1=[]
                for lst in list_pts:
                    if not ((len(lst) == 1) & (lst[0] =='OAD')) |((len(lst) == 1) & (lst[0] =='BYPD')) | ((len(lst) == 1) & (lst[0] =='RAD')) |(((len(lst) == 1) & (lst[0] =='MAT') )|((len(lst) == 1) & (lst[0] =='SS') |((len(lst) == 1) & (lst[0] =='EVAP')))|((len(lst) == 1) & (lst[0] =='SAT'))):
                        list_pts_1.append(lst)
        
                print(list_pts_1)
                df.to_excel('C:\Projects\Script\SchniderAHUData.xlsx')  

                list_df_AHU=[]

                for lst in list_pts_1:
                    list1 = ' '.join(lst)
                    df_AHU=df[df.Point_Name.str.contains(list1)]
                    df_AHU['Point_Name'] = df_AHU['Point_Name'].str.replace(r"\s+", "")
                    print(list1)
                    print(df_AHU)
                    list_df_AHU.append(df_AHU) 

                # First for the CDE Point List, create a key that will be matched to the Database 
                df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :] ## Manual???
                df_dict_1 = pd.DataFrame.from_dict(data1, orient='columns')
                df_Compare_Point  = pd.concat([df_Compare_Point, df_dict_1],axis=0,ignore_index = True)
                df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
                df_Compare_Point.to_excel('C:\Projects\Script\df_Compare_Point.xlsx')

                list_merge_1=[]

                for x in range(0,len(list_df_AHU),1):
                    df_AHU = list_df_AHU[x]
                    if df_AHU.empty:
                        continue
                    print('Raw')
                    print(df_AHU[['Point_Name']])

                    if df_AHU['Point_Name'].str.contains('AHU').any():
                        print("DAHU IS THERE")
                        df_AHU['key1'] = df_AHU['Point_Name'].str.replace('AHU', '')
                        df_AHU.to_excel('C:\Projects\Script\key_problem_1.xlsx')

                        # df_DAHU.loc[df_DAHU['Point_Name'].str.contains('DAHU'), 'Point_Name'] = df_DAHU['Point_Name'].str.replace('DAHU', '')

                    print(df_AHU)

                    # Changing a Logic a Bit

                    merged_dfs = []
                    #list_merge_1=[]

                    for index, row in df_Compare_Point.iterrows():
                        #print ('Trying to Merge')
                        #print(row['key2'])
                        filtered_df = df_AHU[df_AHU['key1'].str.contains(row['key2'], case=False, regex=False)]

                        if not filtered_df.empty:
                            #print('What are we merging')
                            #print(df_Compare_Point.loc[[index]])
                            df_matched = df_Compare_Point.loc[[index]]
                            df_matched = df_matched.reset_index(drop = True)
                            #print('df_matched')
                            #print(df_matched)
                            df_matched['match1']= ['dummy']

                            
                            #print('df_AHU_before_Key')
                            # print(df_AHU)
                            df_AHU.index = df_AHU.index.astype(str)
                            df_AHU['match2'] = ['dummy'] * len(df_AHU)
                            #df_AHU['match2']= ['dummy']
                            '''
                            print('Will it work?')
                            print('df_AHU_afeter_Key')
                            df_AHU = df_AHU. drop ( columns =['key1'])
                            print(df_AHU)
                            df_AHU = df_AHU.assign(key1 = df_matched['key2'])
                            print('adding key2 again')
                            print(df_AHU)
                            '''
                            #print(df_AHU, df_matched)
                        
                            #merged_df = df_AHU.merge(df_matched, left_index = True,right_on = 'key2')
                            #merged_df =pd.concat([df_AHU,df_matched])

                            merged_df = pd.merge(df_AHU, df_matched, left_on='match2', right_on='match1', how = 'inner')
                            print(merged_df)
                            #merged_df.to_excel('C:\Projects\Script\key_problem_3.xlsx')
                            df_AHU_Point = merged_df
                            print(df_AHU_Point[['key1','key2']])
                            list_merge_1.append(df_AHU_Point)
                            print('The Length of List is')
                            print (len(list_merge_1))

                            '''
                            
                            if (df_AHU_Point.empty == False):
                                print(df_AHU_Point)

                                list_merge_1.append(df_AHU_Point)

                                print(df_AHU_Point[['key1','key2']])
                                print(df_AHU_Point)

                            '''

                
        
                    # df_AHU_Point = pd.merge(df_AHU ,df_Compare_Point,left_on='key1',right_on='key2', how='inner')
                

                    
                for x in range(0,len(list_merge_1),1):
                    print ('The Length of List is')
                    print (len(list_merge_1))
                    print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                    print(list_merge_1[x])

                    
                    # This Code Checks the Configuration of the Notification Class against Cde Point List

                    df_AHU_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                    print('df_AHU_Alm_Class')
                    print(df_AHU_Alm_Class)
                    #print(df_DAHU_Alm_Class)
                    len_list_AHU_1 = len_list_AHU_1 + len (df_AHU_Alm_Class['index'].to_list())
                    df_diff_7=  df_AHU_Alm_Class [df_AHU_Alm_Class['Alarm_Class'].astype(str).values ==df_AHU_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                    #print(df_diff_7)
                    list_wtht_diff=  df_AHU_Alm_Class['index'].to_list()
                    # For AHUs, we are evaluating these configurations : 1. Notification Class 2. Alarm Delay
                    # Adding the Notification Class Files
                    len_list_AHU_3 = len_list_AHU_3 + len (list_wtht_diff)
                    list_smlr = df_diff_7['index'].to_list()
                    list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                    #len_list_AHU_3 =len_list_AHU_3 + len (list10)
                    first_row_in_excel=2
                    highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                    list_smlr= [x+ first_row_in_excel for x in list_smlr]
                    string="E"
                    Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                    print(Cell_numbers_Difference_Highlighted)
                    Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                    print(Cell_numbers_similiar)
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid', 
                                    fgColor='FC2C03')
                    for cell in Cell_numbers_Difference_Highlighted:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path) 
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                    for cell in Cell_numbers_similiar:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path) 

                    # This Code Checks the Configuration of the Notification Class against Cde Point List

                    df_AHU_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                    df_AHU_delay_Alarm['Delay'] = df_AHU_delay_Alarm['Delay'].apply(str)
                    df_AHU_delay_Alarm['ALARM DELAY'] = df_AHU_delay_Alarm['ALARM DELAY'].apply(str)
                    print('df_AHU_delay_Alarm')
                    print(df_AHU_delay_Alarm)
                    print(df_AHU_delay_Alarm.dtypes)
                    len_list_AHU_1 = len_list_AHU_1 + len (df_AHU_delay_Alarm['index'].to_list())
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.astype(str).str.replace('1 minute', '60 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                    df_AHU_delay_Alarm['Delay'] =  df_AHU_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                    df_diff_8 = df_AHU_delay_Alarm [df_AHU_delay_Alarm['Delay'].astype(str).values ==df_AHU_delay_Alarm['ALARM DELAY'].astype(str).values]
                    list_wtht_diff=  df_AHU_delay_Alarm['index'].to_list()

                    list_smlr = df_diff_8['index'].to_list()
                    print('similiar_Delay')
                    len_list_AHU_3 =len_list_AHU_3 + len (list_smlr)
                    list_smlr= [x+ first_row_in_excel for x in list_smlr]
                    print(list_smlr)
                    list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                    print('Not_similiar_Delay')
                    print(list11)
                    #len_list_AHU_4=len_list_AHU_4 + len (list11)
                    #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                    first_row_in_excel=2
                    highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                    #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                    #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                    string="G"
                    Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                    Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                    for cell in Cell_numbers_Difference_Highlighted_1:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path)
                    wb = openpyxl.load_workbook(output_File_path)
                    ws = wb['Summary'] #Name of the working sheet
                    fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                    for cell in Cell_numbers_similiar:
                        ws[cell].fill = fill_cell1
                    wb.save(output_File_path)

                
                print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                AHU_cons = pd.concat(list_df_AHU)
                AHU_cons_duplicate = AHU_cons[AHU_cons.index.duplicated()]
                AHU_cons_1=AHU_cons.drop_duplicates(subset = 'index')
                list_AHU_1= df_Compare_12['index'].to_list()
                list_all_AHU_found = AHU_cons_1['index'].to_list()
                list_all_AHU_not_found = [int(x) for x in list_AHU_1 if x not in list_all_AHU_found]
                first_row_in_excel=2
                highlight_rows_AHU=[x+ first_row_in_excel for x in list_all_AHU_not_found]
                with xw.App(visible=False)as app:
                    updated_wb= app.books.open(output_File_path)
                    updated_ws = updated_wb.sheets('Summary')
                    rng=updated_ws.used_range
                    print(rng.address) 
                    for row in rng.rows:
                        if row.row in highlight_rows_AHU:
                            row.color ='ffff00'
                    updated_wb.save(output_File_path) 
                
                global len_list_tot_AHU, len_list_wtht_diff_AHU,list_tot_AHU,list_dssmlr_AHU

                len_list_tot_AHU = len_list_AHU_1
                len_list_wtht_diff_AHU =  len_list_AHU_3
                list_tot_AHU=[]
                list_dssmlr_AHU = []
                list_tot_AHU.append(len_list_tot_AHU)
                list_dssmlr_AHU.append(len_list_wtht_diff_AHU)
                print('list_tot')
                print(list_tot_AHU)
                print('list_dssmlr')
                print(list_dssmlr_AHU)

                '''

                len_list_tot_AHU =0 
                len_list_wtht_diff_AHU=0
                list_dssmlr_AHU=0

                
                #len_list_tot_AHU = len_list_AHU_2 + len_list_AHU_4
                
                if(len_list_tot_AHU > 0) :
                    len_list_tot_AHU = len_list_AHU_1
                    len_list_wtht_diff_AHU =  len_list_AHU_3
                    list_tot_AHU=[]
                    list_dssmlr_AHU = []
                    list_tot_AHU.append(len_list_tot_AHU)
                    list_dssmlr_AHU.append(len_list_wtht_diff_AHU)
                    print('list_tot')
                    print(list_tot_AHU)
                    print('list_dssmlr')
                    print(list_dssmlr_AHU) 
            '''
    

        df_Compare_13 = df_Alrm_Trend[(df_Alrm_Trend.Point_Name.str.contains("IW") |df_Alrm_Trend.Point_Name.str.contains("Iw"))  & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
        print ('df_Compare_13')
        print(df_Compare_13)
        df_Compare_13.to_excel('C:\Projects\Script\SchneiderIW.xlsx') 

        
        if(df_Compare_13.shape[0] >=1):
            global len_list_IW_1
            global len_list_IW_3
            len_list_IW_1 =0
            len_list_IW_3 =0
            list_ids=[]
            
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'INDUSTRIAL WATER SYSTEM'):
                    idx_start =index
                    list_ids.append(idx_start)
            df_Compare_final_ids = df_Compare.iloc[list_ids[0]:] 
            df_Compare_final_ids_null = df_Compare_final_ids[df_Compare_final_ids['Point_Name'].isnull()]
            list_ids.append(((df_Compare_final_ids_null.index.to_list())[0])-1)
            df_Compare_IW_CDE =df_Compare_2.loc[list_ids[0]:list_ids[1], :] 
            print(df_Compare_IW_CDE) 
            df_Compare_IW_CDE.fillna("",inplace=True)
            df_Compare_IW_CDE_1 = df_Compare_IW_CDE.assign(new_col=df_Compare_IW_CDE[3].str.replace('##',''))
            df_Compare_IW_CDE_1['new_col'] = 'IW'
            df_Compare_IW_CDE_1.to_excel('C:\Projects\Script\IWSchneider.xlsx')
            
                    # This Code Checks the Configuration of the Noti
        
            list_val =[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10] # Is this hard-coded??
            list1=[]
            list2=[]
            for index,rows in df_Compare_IW_CDE_1.iterrows():
                key = df_Compare_IW_CDE_1.loc[index,'new_col']
                for x in list_val:
                    list1.append(df_Compare_IW_CDE_1.loc[index,x])
                value =list1
                dicT= {key : value} 
                list1=[]
                list2.append(dicT)
            list102=[]
            list101=[]
            for index in range(len(list2)):
                for i,x in enumerate (list2[index][key]):
                    if (x == ''):
                        if(i <= len(list2[index][key])):
                            list101.append(i)
                list2[index][key] = np.delete((list2[index][key]),list101).tolist()
                list101=[]
            print(list2)
            df = df_Compare_13
            df = df.assign(Point_Name=df.Point_Name.str.replace('_',' '))
            df = df.assign(Point_Name=df.Point_Name.str.upper())
            df = df.join(df.Point_Name.str.split(r' ',expand=True))
            df.to_excel('C:\Projects\Script\dfIW.xlsx') # This is a checkpoint only and not part of the final code 
            print(df)
            df['Point_Name']=df['Point_Name'].astype(str).str.replace(r'[^A-Za-z ]', '', regex=True).replace('', np.nan, regex=False)
            list_pts=[]
            for dicts in list2:
                for eqpt,points in dicts.items():
                    list_pts.append(points) 
            print(list_pts)
            list_pts_1=[]
            for lst in list_pts:
                if not ((len(lst) == 1) & (lst[0] =='PRESS')) |((len(lst) == 1) & (lst[0] =='MODE')) | ((len(lst) == 1) & (lst[0] =='IW')) |(((len(lst) == 1) & (lst[0] =='MAT') )|((len(lst) == 1) & (lst[0] =='SS') |((len(lst) == 1) & (lst[0] =='EVAP')))|((len(lst) == 1) & (lst[0] =='SAT'))):
                    list_pts_1.append(lst)

            print(list_pts_1)

            list_df_IW=[]

            for lst in list_pts_1:
                list1 = ' '.join(lst)
                df_IW=df[df.Point_Name.str.contains(list1)]
                df_IW['Point_Name'] = df_IW['Point_Name'].str.replace(r"\s+", "")
                print(list1)
                print(df_IW)
                list_df_IW.append(df_IW)
            df_Compare_Point=df_Compare.loc[list_ids[0]+1:list_ids[1], :]
            df_Compare_Point['key2']= df_Compare_Point.Point_Name.str.replace('[','').str.replace(']','')
            df_Compare_Point.to_excel('C:\Projects\Script\df_Compare_Point_IW.xlsx')

            list_merge_1=[]
            for x in range(0,len(list_df_IW),1):
                df_IWS = list_df_IW[x]
                if df_IWS.empty:
                    continue
                print('Raw')
                print(df_IWS[['Point_Name']])

                if df_IWS['Point_Name'].str.contains('IW').any():
                    print("IW IS THERE")
                    df_IWS['key1'] = df_IWS['Point_Name'].str.replace('IW', '')
                    df_IWS.to_excel('C:\Projects\Script\key_problem_IW.xlsx')
                merged_dfs = []
                for index, row in df_Compare_Point.iterrows():
                    #print ('Trying to Merge')
                    #print(row['key2'])
                    filtered_df = df_IWS[df_IWS['key1'].str.contains(row['key2'], case=False, regex=False)]
                    if not filtered_df.empty:
                        #print('What are we merging')
                        #print(df_Compare_Point.loc[[index]])
                        df_matched = df_Compare_Point.loc[[index]]
                        df_matched = df_matched.reset_index(drop = True)
                        #print('df_matched')
                        #print(df_matched)
                        df_matched['match1']= ['dummy']
                        #print('df_AHU_before_Key')
                        # print(df_AHU)
                        df_IWS.index = df_IWS.index.astype(str)
                        df_IWS['match2'] = ['dummy'] * len(df_IWS)
                        #df_AHU['match2']= ['dummy']
                        merged_df = pd.merge(df_IWS, df_matched, left_on='match2', right_on='match1', how = 'inner')
                        print(merged_df)
                        merged_df.to_excel('C:\Projects\Script\key_problem_3.xlsx')
                        df_IW_Point = merged_df
                        print(df_IW_Point[['key1','key2']])
                        list_merge_1.append(df_IW_Point)
                        print('The Length of List is')
                        print (len(list_merge_1))

            for x in range(0,len(list_merge_1),1):
                print ('The Length of List is')
                print (len(list_merge_1))
                print ("The Script is evalauting AHUs it may take several minutes, please be patient")
                print(list_merge_1[x])
                

                
                # This Code Checks the Configuration of the Notification Class against Cde Point List

                df_IW_Alm_Class = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Alarm_Class','NOTIFICATION LEVEL']]
                print('df_IW_Alm_Class')
                print(df_IW_Alm_Class)
                #print(df_DAHU_Alm_Class)
                len_list_IW_1 = len_list_IW_1 + len (df_IW_Alm_Class['index'].to_list())
                df_diff_7=  df_IW_Alm_Class [df_IW_Alm_Class['Alarm_Class'].astype(str).values ==df_IW_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
                #print(df_diff_7)
                list_wtht_diff=  df_IW_Alm_Class['index'].to_list()
                # For AHUs, we are evaluating these configurations : 1. Notification Class 2. Alarm Delay
                # Adding the Notification Class Files
                len_list_IW_3 = len_list_IW_3 + len (list_wtht_diff)
                list_smlr = df_diff_7['index'].to_list()
                list10 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                #len_list_AHU_3 =len_list_AHU_3 + len (list10)
                first_row_in_excel=2
                highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list10]
                list_smlr= [x+ first_row_in_excel for x in list_smlr]
                string="E"
                Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
                print(Cell_numbers_Difference_Highlighted)
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                print(Cell_numbers_similiar)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', 
                                fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path) 

                # This Code Checks the Configuration of the Notification Class against Cde Point List

                df_IW_delay_Alarm = list_merge_1[x][['index','Point_Name_x','Point_Name_y','Delay','ALARM DELAY' ]]
                df_IW_delay_Alarm['Delay'] = df_IW_delay_Alarm['Delay'].apply(str)
                df_IW_delay_Alarm['ALARM DELAY'] = df_IW_delay_Alarm['ALARM DELAY'].apply(str)
                print(df_IW_delay_Alarm.dtypes)
                len_list_IW_1 = len_list_IW_1 + len (df_IW_delay_Alarm['index'].to_list())
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('seconds', 'sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('minutes', 'min')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('1 minute', '60 sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('30secs', '30 sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('5secs', '5 sec')
                df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('30 seconds', '30 sec')
                df_IW_delay_Alarm['Delay'] = df_IW_delay_Alarm['Delay'].astype(str).str.replace('0 ms', 'nan')
                print('df_IW_delay_Alarm')
                print(df_IW_delay_Alarm)
                #df_IW_delay_Alarm['Delay'] =  df_IW_delay_Alarm.Delay.astype(str).str.replace('0 ms', '')
                list_wtht_diff=  df_IW_delay_Alarm['index'].to_list()
                print('list_wtht_diff')
                print(list_wtht_diff)
                df_diff_8 = df_IW_delay_Alarm [df_IW_delay_Alarm['Delay'].astype(str).values ==df_IW_delay_Alarm['ALARM DELAY'].astype(str).values]
                # list_wtht_diff=  df_IW_delay_Alarm['index'].to_list()
                list_smlr = df_diff_8['index'].to_list()
                print('similiar_Delay')
                len_list_IW_3 =len_list_IW_3 + len (list_smlr)
                # list_smlr= [x+ first_row_in_excel for x in list_smlr]
                print(list_smlr)
                list11 = [int(x) for x in list_wtht_diff if x not in list_smlr]
                print('Not_similiar_Delay')
                print(list11)
                #len_list_AHU_4=len_list_AHU_4 + len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                first_row_in_excel=2
                highlight_rows_alarm_delay=[x+ first_row_in_excel for x in list11]
                #len_list_AHU_4 =len_list_AHU_4+ len (list11)
                #len_list_AHU_3 =len_list_AHU_3 + len (list11)
                string="G"
                Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay]
                Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
                for cell in Cell_numbers_Difference_Highlighted_1:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws = wb['Summary'] #Name of the working sheet
                fill_cell1 = PatternFill(patternType='solid',fgColor='00FFFFFF')
                for cell in Cell_numbers_similiar:
                    ws[cell].fill = fill_cell1
                wb.save(output_File_path)   


       
            # Creating List 

            #len_list_1 =0 
            #len_list_2 =0
            #len_list_3 =0 
            #len_list_4=0

    # Adding the code for checking the Area Controller 
    # For the data hall area controller 
    # Starting with CAT Sensors 

        df_Schneider_Area = df_Alrm_Trend
        df_Schneider_Area.to_excel ('C:\Projects\Script\Area\Test7.xlsx')
        list_sensors = ['CAT_A', 'CAT_B', 'CAT_C','CAT_D','CAT_E', 'CAT_F']
        list_aisles =  ['IO_SYS','AISLE01','_AISLE01','AISLE02','AISLE03', 'AISLE04', 'AISLE05','AISLE06', 'AISLE07', 'AISLE08', 'AISLE09','AISLE10' ]
        list_CAT_1 =[]
        list_CAT_2=[]
        list_CAT_3 =[]

        for x in list_sensors:
            list_CAT_1.append(x)
            for y in list_aisles:
                list_CAT_2.append (y)
                list_CAT_4 = list (zip (list_CAT_1, list_CAT_2))
                list_CAT_3.append(list_CAT_4)
                list_CAT_2=[]
            list_CAT_1=[]
        
        print(list_CAT_3)

        # Creating a data frame of the Cold Aisle Sensors 
        df_Schneider_CAT_Aisle = pd.DataFrame()
        for lsts in list_CAT_3:
            for x, y in lsts:
                df_Schneider_Area_3 = df_Schneider_Area [(df_Schneider_Area['Point_Name'].str.contains(x)) &(df_Schneider_Area['Equipment'].str.contains(y))]
                df_Schneider_CAT_Aisle = pd.concat([df_Schneider_CAT_Aisle,df_Schneider_Area_3])

        # df_Schneider_CAT_Aisle = df_Schneider_CAT_Aisle[df_Schneider_CAT_Aisle['Alarm_Extension'].str.contains ('OutOfRangeAlarmExt')]

        print ('df_Schneider_CAT_Aisle')
        print(df_Schneider_CAT_Aisle)
        df_Schneider_CAT_Aisle.to_excel('C:\Projects\Script\Area\CATsensors.xlsx')
        
        # Comparing CAT Configuration with CDE Point List

        #df_CDE_CAT = pd.read_excel(cde_point_list_path) 
        #df_CDE_CAT = df_CDE_CAT.rename(columns ={"FUNCTION":'Point_Description',df_CDE_CAT.columns[1]:'Point_Name', "DEFINITION":"Trend_interval",df_CDE_CAT.columns[17]:"Alarm",df_CDE_CAT[18]:"Alarm LO/HI"}).drop(columns=['JITDC','OPTDC','Site Specific','Relinquish Default','NOTES'])
        
        if (df_Schneider_CAT_Aisle.shape[0] >= 1 ):

            list_id_CAT=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Cold Aisle Temp'):
                    idx_start =index
                    list_id_CAT.append(idx_start)

            df_Compare_final_ids = df_Compare.iloc[[list_id_CAT[0]]]
            df_Compare_final_ids.to_excel('C:\Projects\Script\Cat_Original.xlsx')

            df_CAT =df_Compare_final_ids [['Point_Description','Point_Name','Alarm LO/HI','ALARM DELAY', 'NOTIFICATION LEVEL', 'ALARM TXT', 'DEAD BAND' ]]
            df_CAT = df_CAT.rename(columns ={df_CAT.columns[2]:'Alarm_Range'})
            df_CAT.to_excel('C:\Projects\Script\Area\CATsensors.xlsx')

            df_CAT['Alarm_Range'] = df_CAT['Alarm_Range'].str.replace(r"\(.*\)","")
            df_CAT [['Alarm_Low','Alarm_High']] = df_CAT ['Alarm_Range'].str.split ( '/', expand = True, regex=False) 
            df_CAT.to_excel('C:\Projects\Script\Area\CATsensors1.xlsx')
            df_CAT_1 = df_CAT
            df_CAT_1 = df_CAT_1.drop (columns = 'Point_Name')
            df_CAT_1 ['key2'] = 'CAT'

            df_Schneider_CAT_Aisle['key1'] = 'CAT'
            
            ## REVIST THIS PART OF THE CODE
            # Adding the below line of codes to remove characters after dots in strings
            #df_Schneider_CAT_Aisle ['high_Limit'] = df_Schneider_CAT_Aisle ['high_Limit'].str.split('.').str[0]
            #df_Schneider_CAT_Aisle ['low_Limit'] = df_Schneider_CAT_Aisle ['low_Limit'].str.split('.').str[0]
            # Adding the below line of codes to ensure that the time formatting 
            #df_Schneider_CAT_Aisle ['Delay'] = df_Schneider_CAT_Aisle.Delay.str.replace('seconds', 'sec')
            #df_Schneider_CAT_Aisle ['Delay'] = df_Schneider_CAT_Aisle.Delay.str.replace('minutes', 'min')
            #df_Schneider_CAT_Aisle ['Delay'] = df_Schneider_CAT_Aisle  .Delay.str.replace('1 minute', '60 sec')
                
            # Merging the Data Frame

            df_CAT_compare = pd.merge(df_Schneider_CAT_Aisle,df_CAT_1,left_on = 'key1', right_on ='key2', how = 'inner')

            # Total number of CAT attributes checked
            len_list_tot_CAT_Schndr =0
            len_list_tot_CAT_Schndr = len_list_tot_CAT_Schndr +  5 * (len (df_CAT_compare['index'].to_list()))
            # Converting the units 
            df_CAT_compare['Delay'] = df_CAT_compare['Delay'].apply(str)
    
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('seconds', 'sec')
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('minutes', 'min')
            df_CAT_compare['Delay'] =  df_CAT_compare.Delay.str.replace('1 minute', '60 sec')

            df_CAT_compare['DEAD BAND'] = df_CAT_compare['DEAD BAND'].str.replace(r"\(.*\)","")

            df_CAT_compare.to_excel ('C:\Projects\Script\Area\Merge.xlsx')
            

            # Creating this error to Test the code 
            #df_CAT_compare.loc[0,'Alarm_Class'] = 'Low'
            # Converting the error in the code 

            #df_CAT_compare.loc[0,'Delay'] = 1000


            df_CAT_compare.to_excel('C:\Projects\Script\Area\Testpoint.xlsx')

            list_CAT=  df_CAT_compare['index'].to_list()
            print("CAT List")
            print(list_CAT)
            list_CAT= [ x for x in list_CAT if ~np.isnan(x)]
            print(list_CAT)

            #Comparing the Low Temperature Range betwwen the N4 Report and the Point List 

            

            df_CAT_compare['low_Limit'] = df_CAT_compare['low_Limit'].str.split(' ').str[0]
            df_CAT_compare['high_Limit'] =df_CAT_compare['high_Limit'].str.split(' ').str[0]
            df_CAT_compare['Dead_Band'] =df_CAT_compare['Dead_Band'].str.split(' ').str[0]
            #df_DP['DEAD BAND'] = df_DP['DEAD BAND'].str.replace(r"\(.*\)","")

            df_CAT_compare.to_excel ('C:\Projects\Script\Area\Farenheit.xlsx')

        

            df_CAT_Comp_Low =   df_CAT_compare [['index', 'Equipment','low_Limit','Alarm_Low' ]]
            len_list_tot_CAT_Schndr_Low = len_list_tot_CAT_Schndr_Low + len(df_CAT_Comp_Low['index'].to_list())

            df_CAT_Comp_Low_1 = df_CAT_Comp_Low[df_CAT_Comp_Low['low_Limit'].astype(float).values == df_CAT_Comp_Low['Alarm_Low'].astype(float).values]
            list_smlr_CAT= df_CAT_Comp_Low_1['index'].to_list()
            list_smlr_CAT= [ x for x in list_smlr_CAT if ~np.isnan(x)]
            len_list_smlr_CAT_Schndr_Low = len_list_smlr_CAT_Schndr_Low + len (list_smlr_CAT)
            print(list_smlr_CAT)
            list_not_smlr_CAT = [int(x) for x in list_CAT if x not in list_smlr_CAT]
            len_list_dissmlr_CAT_Schndr_Low = len_list_dissmlr_CAT_Schndr_Low + len (list_not_smlr_CAT)
            print (list_not_smlr_CAT)

            # Writing to SpreadsHEET Where there are differences between the Low Limit 

            # Writing Back to Excel sheet where the Low Pressure limit does not match between the Point List and N4 Report 
            # Writing to Column K to the SpreadSheet

            first_row_in_excel=2
            highlight_rows_CAT_low_mismatch=[x+ first_row_in_excel for x in list_not_smlr_CAT]
            string="K"
            Cell_numbers_Difference_Highlighted_Low_CAT= ["{}{}".format(string,i) for i in highlight_rows_CAT_low_mismatch]
            print(Cell_numbers_Difference_Highlighted_Low_CAT)
        

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Low_CAT:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Comparing the High Limit betwwen the N4 Report and the Point List 

            df_CAT_Comp_High=   df_CAT_compare [['index', 'Equipment','high_Limit','Alarm_High' ]]
            len_list_tot_CAT_Schndr_High = len_list_tot_CAT_Schndr_High + len(df_CAT_Comp_High['index'].to_list())
            df_CAT_compare_High_1 = df_CAT_Comp_High[df_CAT_Comp_High['high_Limit'].astype(float).values == df_CAT_Comp_High['Alarm_High'].astype(float).values]
            list_smlr_CAT_1 = df_CAT_compare_High_1['index'].to_list()
            list_smlr_CAT_1= [ x for x in list_smlr_CAT_1 if ~np.isnan(x)]
            len_list_smlr_CAT_Schndr_High = len_list_smlr_CAT_Schndr_High + len (list_smlr_CAT)
            print(list_smlr_CAT_1)
            list_not_smlr_CAT_1 = [int(x) for x in list_CAT if x not in list_smlr_CAT_1]
            len_list_dissmlr_CAT_Schndr_High = len_list_dissmlr_CAT_Schndr_High+ len(list_not_smlr_CAT_1)
            print (list_not_smlr_CAT_1)


            first_row_in_excel=2
            highlight_rows_alarm_CAT_high_mismatch=[x+ first_row_in_excel for x in list_not_smlr_CAT_1]
            string="J"
            Cell_numbers_Difference_Highlighted_High_CAT= ["{}{}".format(string,i) for i in highlight_rows_alarm_CAT_high_mismatch]
            print(Cell_numbers_Difference_Highlighted_High_CAT)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_High_CAT:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Comparing the Dead Band between the CDE Point List and configured as shown by N4 Report 

            df_CAT_DdBnd = df_CAT_compare [['index', 'Equipment','Dead_Band','DEAD BAND' ]]
            len_list_tot_CAT_Schndr_Dbnd = len_list_tot_CAT_Schndr_Dbnd + len(df_CAT_DdBnd ['index'].to_list())

            df_CAT_DdBnd_1 = df_CAT_DdBnd[df_CAT_DdBnd['Dead_Band'].astype(float).values == df_CAT_DdBnd['DEAD BAND'].astype(float).values]
            lst_smlr_CAT_DdBnd =df_CAT_DdBnd_1 ['index'].to_list()
            lst_smlr_CAT_DdBnd_1=[ x for x in lst_smlr_CAT_DdBnd   if ~np.isnan(x)]
            len_list_smlr_CAT_Schndr_Dbnd = len_list_smlr_CAT_Schndr_Dbnd + len(lst_smlr_CAT_DdBnd_1)
            lst_smlr_CAT_DdBnd_2=[int(x) for x in list_CAT if x not in lst_smlr_CAT_DdBnd_1]
            len_list_dissmlr_CAT_Schndr_Dbnd = len_list_dissmlr_CAT_Schndr_Dbnd + len(lst_smlr_CAT_DdBnd_2)

            first_row_in_excel=2
            highlight_rows_alarm_CAT_dbnd=[x+ first_row_in_excel for x in lst_smlr_CAT_DdBnd_2]
            string="L"
            Cell_numbers_Difference_Highlighted_Dnd_CAT= ["{}{}".format(string,i) for i in highlight_rows_alarm_CAT_dbnd]
            print(Cell_numbers_Difference_Highlighted_Dnd_CAT)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Dnd_CAT:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)


            # Comparing the Notification Class of the Cold Aisle Sensors
        #global  len_list_tot_CAT_Schndr_Class
            
            len_list_tot_CAT_Schndr_Class = len_list_tot_CAT_Schndr_Dbnd

            df_diff_7_CAT=  df_CAT_compare [df_CAT_compare['Alarm_Class'].astype(str).values ==df_CAT_compare ['NOTIFICATION LEVEL'].astype(str).values]

            # Highlighlighting Cells where there is a diffwnce in the Notificatiob clasas


            list_smlr = df_diff_7_CAT['index'].to_list()
            list_smlr= [ x for x in list_smlr if ~np.isnan(x)]
            len_list_smlr_CAT_Schndr_Class = len_list_smlr_CAT_Schndr_Class + len(list_smlr)

            list100 = [int(x) for x in list_CAT if x not in list_smlr]
            len_list_dissmlr_CAT_Schndr_Class = len_list_dissmlr_CAT_Schndr_Class + len(list100)
            #cleanedList = [x for x in countries if x != 'nan']
            print(list100)

            # Writing back to excel file to highlight where CAT notification class do not match 

            
            first_row_in_excel=2
            highlight_rows_alarm_class_mismatch=[x+ first_row_in_excel for x in list100]
            list_smlr= [x+ first_row_in_excel for x in list_smlr]
            string="E"
            Cell_numbers_Difference_Highlighted= ["{}{}".format(string,i) for i in highlight_rows_alarm_class_mismatch]
            print(Cell_numbers_Difference_Highlighted)
            Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr]


            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            len_list_tot_CAT_Schndr_Delay = len_list_tot_CAT_Schndr_Dbnd

            ## Alarm Delay - Comparing Alarm Delay between the Point List and Report
            df_diff_8_CAT=  df_CAT_compare [df_CAT_compare['Delay'].astype(str).values ==df_CAT_compare ['ALARM DELAY'].astype(str).values]
            df_diff_8_CAT.to_excel('C:\Projects\Script\CATDelay.xlsx')
            list_smlr_1 = df_diff_8_CAT['index'].to_list()
            len_list_smlr_CAT_Schndr_Delay= len_list_smlr_CAT_Schndr_Delay+ len (list_smlr_1)
            list_smlr_1= [ x for x in list_smlr_1 if ~np.isnan(x)]
            list200 = [int(x) for x in list_CAT if x not in list_smlr_1]
            len_list_dissmlr_CAT_Schndr_Delay = len_list_dissmlr_CAT_Schndr_Delay + len (list200)
            print('list200')
            print(list200)

            # Writing back to excel sheet

            first_row_in_excel=2
            highlight_rows_alarm_delay_mismatch=[x+ first_row_in_excel for x in list200]
            list_smlr_1= [x+ first_row_in_excel for x in list_smlr_1]
            string="G"
            Cell_numbers_Difference_Highlighted_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_delay_mismatch]
            print(Cell_numbers_Difference_Highlighted_1)
            Cell_numbers_similiar=["{}{}".format(string,i) for i in list_smlr_1]

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_1:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Misiing for CAT checks the code on comparing the low and high limits of the Alarm and also Deabd
            # Here the Code is becoming REPETIVITE, the next iteration to create a Python Function and call it with different Data Frames, this will result in more organized and better code 

            # Adding code to call out the cat alarms not part of cde point list 
            # Missing eadband 

            df_CAT_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('CAT') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            list_all_CAT= df_CAT_all ['index'].to_list()
            list_CAT_aisle_sensors = df_Schneider_CAT_Aisle['index'].to_list()
            list_CAT_not_comp= [int(x) for x in list_all_CAT if x not in list_CAT_aisle_sensors]
            first_row_in_excel=2
            list_CAT_not_comp= [x+ first_row_in_excel for x in list_CAT_not_comp]

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_CAT_not_comp:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)

        
        # Adding the functionality of the pressure transmitters


        list_DP_Sensors = ['IO_SYS','_SAPDP01', '_SAPDP02', '_SAPDP03','_SAPDP04','_SAPDP05', '_SAPDP06']
        list_Areas =  ['D1A1','D1A2','D1A3', 'D1A4', 'D1A5','D1A6', 'D2A1','D2A2','D2A3', 'D2A4', 'D2A5','D2A6']
        list9=[]
        list10=[]
        list11 =[]
        for x in list_Areas:
            list9.append(x)
            #print (list9)
            for y in list_DP_Sensors:
                list10.append(y)
                list12 = list(zip(list9,list10))
                #print(list12)
                list11.append(list12)
                #list11 =list11.append(list12)
                list10 =[] 
            list9=[]


        list12=[]
        for lsts in list11:
            for x, y in lsts:
                list12.append(x+y)
        
        print (list12)

        df_Schneider_Area_DP_2 = pd.DataFrame()
        for x in list12:
            df_Schneider_Area_DP_1 = df_Schneider_Area [(df_Schneider_Area['Point_Name'].str.contains(x)) &(df_Schneider_Area['Equipment'].str.contains('DAHU_SYS'))]
            df_Schneider_Area_DP_2 = pd.concat([df_Schneider_Area_DP_2,df_Schneider_Area_DP_1])
        
        df_Schneider_Area_DP_2.to_excel ('C:\Projects\Script\Area\Pressure.xlsx')

        
        if (df_Schneider_Area_DP_2.shape[0] >= 1 ):


        # Creating the Pressure data frame
            df_pressure = df_Schneider_Area_DP_2
            df_pressure = df_pressure.drop (columns = 'Point_Name')
            df_pressure['key1'] = 'SHDP'


            # Saving the differential pressure data frame to excel file

            list_id_DP=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Supply Air Header to OA DP'):
                    idx_start =index
                    list_id_DP.append(idx_start)

            df_Compare_final_ids_dP = df_Compare.iloc[[list_id_DP[0]]]
        

            df_DP =df_Compare_final_ids_dP [['Point_Description','Point_Name','Alarm LO/HI','ALARM DELAY', 'NOTIFICATION LEVEL', 'ALARM TXT', 'DEAD BAND']]
            df_DP.to_excel ('C:\Projects\Script\Area\PointListNew.xlsx')
            df_DP = df_DP.rename(columns ={df_DP.columns[2]:'Alarm_Range'})
            df_DP['Alarm_Range'] = df_DP['Alarm_Range'].str.replace(r"\(.*\)","")
            df_DP['DEAD BAND'] = df_DP['DEAD BAND'].str.replace(r"\(.*\)","")
            df_DP [['Alarm_Low','Alarm_High']] = df_DP ['Alarm_Range'].str.split ( '/', expand = True) 
            df_DP_1 = df_DP
            df_DP_1 = df_DP_1.drop (columns = 'Point_Name')
            df_DP_1 ['key2'] = 'SHDP'

            # Merging the Dtata Frame
            df_dp_compare = pd.merge(df_pressure,df_DP_1,left_on = 'key1', right_on ='key2', how = 'inner')
            list_DP_Comp = df_dp_compare ['index'].to_list()


            df_dp_compare.to_excel ('C:\Projects\Script\Area\Pressure2.xlsx')
            
            # Filtering and Comparing the Pressure Sensor attributes between the Point List and N4 Report
            
            # Comparing the Pressure Sensors attributes 
            # Comparing the low and high Pressure Limits
            # LOW LIMIT

            df_DP_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('SAPDP') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            df_DP_all.to_excel ('C:\Projects\Script\Area\PressureAll.xlsx')
            list_all_DP= df_DP_all ['index'].to_list()

            print(list_all_DP)

            # Creating an Exception here - writing all the rows back to white, this is because some previous codfe have made these rows yellow  

            first_row_in_excel=2
            list_all_DP_1= [x+ first_row_in_excel for x in list_all_DP]


            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_all_DP_1:
                        row.color ='ffffff'
                updated_wb.save(output_File_path)


            # Highlighting Rows for SAPDP where the comparision was not made 

            list_DP_not_comp = [int(x) for x in list_all_DP if x not in list_DP_Comp ]
            first_row_in_excel=2
            list_DP_not_comp_1= [x+ first_row_in_excel for x in list_DP_not_comp]

            print(list_DP_not_comp_1)

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_DP_not_comp_1:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)


            # Comparing the Low Pressure Range betwwen the N4 Report and the Point List 

            df_Pressure_Comp_Low =   df_dp_compare [['index', 'Equipment','low_Limit','Alarm_Low' ]]
            df_Pressure_Comp_Low['Alarm_Low'] = df_Pressure_Comp_Low['Alarm_Low'].replace('"','',regex=True).astype(float)
            df_Pressure_Comp_Low_1 = df_Pressure_Comp_Low[df_Pressure_Comp_Low['low_Limit'].astype(float).values == df_Pressure_Comp_Low['Alarm_Low'].astype(float).values]
            list_smlr_DP = df_Pressure_Comp_Low_1['index'].to_list()
            list_smlr_DP= [ x for x in list_smlr_DP if ~np.isnan(x)]
            print(list_smlr_DP)
            list_not_smlr_DP = [int(x) for x in list_DP_Comp if x not in list_smlr_DP]
            print (list_not_smlr_DP)
            
            # Writing Back to Excel sheet where the Low Pressure limit does not match between the Point List and N4 Report 
            # Writing to Column K to the SpreadSheet

            first_row_in_excel=2
            highlight_rows_alarm_low_mismatch=[x+ first_row_in_excel for x in list_not_smlr_DP]
            string="K"
            Cell_numbers_Difference_Highlighted_Low_DP= ["{}{}".format(string,i) for i in highlight_rows_alarm_low_mismatch]
            print(Cell_numbers_Difference_Highlighted_Low_DP)
        

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Low_DP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

        # Comparing the High Pressure Range betwwen the N4 Report and the Point List 

            df_Pressure_Comp_High=   df_dp_compare [['index', 'Equipment','high_Limit','Alarm_High' ]]
            df_Pressure_Comp_High['Alarm_High'] = df_Pressure_Comp_High['Alarm_High'].replace('"','',regex=True).astype(float)
            df_Pressure_Comp_High_1 = df_Pressure_Comp_High[df_Pressure_Comp_High['high_Limit'].astype(float).values == df_Pressure_Comp_High['Alarm_High'].astype(float).values]
            list_smlr_DP_1 = df_Pressure_Comp_High_1['index'].to_list()
            list_smlr_DP_1= [ x for x in list_smlr_DP_1 if ~np.isnan(x)]
            print(list_smlr_DP_1)
            list_not_smlr_DP_1 = [int(x) for x in list_DP_Comp if x not in list_smlr_DP_1]
            print (list_not_smlr_DP_1)


            first_row_in_excel=2
            highlight_rows_alarm_high_mismatch=[x+ first_row_in_excel for x in list_not_smlr_DP_1]
            string="J"
            Cell_numbers_Difference_Highlighted_High_DP= ["{}{}".format(string,i) for i in highlight_rows_alarm_high_mismatch]
            print(Cell_numbers_Difference_Highlighted_High_DP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_High_DP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

        # Comparing the notification class 

            df_Pressure_Alm_Class = df_dp_compare [['index', 'Equipment','Alarm_Class','NOTIFICATION LEVEL' ]]
            df_Pressure_Class=df_Pressure_Alm_Class [df_Pressure_Alm_Class['Alarm_Class'].astype(str).values ==df_Pressure_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
            list_smlr_Pressure = df_Pressure_Class ['index'].to_list()
            list_smlr_Pressure_1=[ x for x in list_smlr_Pressure if ~np.isnan(x)]
            list_smlr_Pressure_2=[int(x) for x in list_DP_Comp if x not in list_smlr_Pressure_1]

            first_row_in_excel=2
            highlight_rows_alarm_Class_mismatch=[x+ first_row_in_excel for x in list_smlr_Pressure_2]
            string="E"
            Cell_numbers_Difference_Highlighted_Class_DP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Class_mismatch]
            print(Cell_numbers_Difference_Highlighted_Class_DP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Class_DP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

        # Comparing the Delay time
            
            df_Pressure_Delay = df_dp_compare [['index', 'Equipment','Delay','ALARM DELAY' ]]
            df_Pressure_Delay['Delay'] = df_Pressure_Delay['Delay'].apply(str)
            df_Pressure_Delay['Delay'] =  df_Pressure_Delay.Delay.str.replace('seconds', 'sec')
            df_Pressure_Delay['Delay'] =  df_Pressure_Delay.Delay.str.replace('minutes', 'min')
            df_Pressure_Delay['Delay'] =  df_Pressure_Delay.Delay.str.replace('1 minute', '60 sec')

            df_Pressure_Delay.to_excel('C:\Projects\Script\Pressure7.xlsx')

            df_delay_DP=  df_Pressure_Delay [df_Pressure_Delay['Delay'].astype(str).values ==df_Pressure_Delay ['ALARM DELAY'].astype(str).values]
            
            df_delay_DP.to_excel('C:\Projects\Script\Pressure9.xlsx')


            # Highlighlighting Cells where there is a diffwnce in the Notificatiob clasas
            lst_smlr_Pressure_Delay = df_delay_DP['index'].to_list()
            lst_smlr_Pressure_Delay_1=[ x for x in lst_smlr_Pressure_Delay if ~np.isnan(x)]
            print(lst_smlr_Pressure_Delay_1)
            lst_smlr_Pressure_Delay_2=[int(x) for x in list_DP_Comp if x not in lst_smlr_Pressure_Delay_1]
            print(lst_smlr_Pressure_Delay_2)

        
            first_row_in_excel=2
            highlight_rows_alarm_Delay_mismatch=[x+ first_row_in_excel for x in lst_smlr_Pressure_Delay_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Delay_DP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Delay_mismatch]
            #print(Cell_numbers_Difference_Highlighted_Delay_DP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Delay_DP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

        # Comparing the Dead Band , the Deadband is an important variable to check for analog type Alarms 

            df_Pressure_DdBnd = df_dp_compare [['index', 'Equipment','Dead_Band','DEAD BAND' ]]
            df_Pressure_DdBnd_1 = df_Pressure_DdBnd[df_Pressure_DdBnd['Dead_Band'].astype(float).values == df_Pressure_DdBnd['DEAD BAND'].astype(float).values]
            lst_smlr_Pressure_DdBnd =df_Pressure_DdBnd_1 ['index'].to_list()
            lst_smlr_Pressure_DdBnd_1=[ x for x in lst_smlr_Pressure_DdBnd if ~np.isnan(x)]
            lst_smlr_Pressure_DdBnd_2=[int(x) for x in list_DP_Comp if x not in lst_smlr_Pressure_DdBnd_1]

            first_row_in_excel=2
            highlight_rows_alarm_Delay_dbnd=[x+ first_row_in_excel for x in lst_smlr_Pressure_DdBnd_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Dnd_DP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Delay_dbnd]
            print(Cell_numbers_Difference_Highlighted_Dnd_DP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Dnd_DP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

    # ADDING THE CODE FOR CHECKING CADP ALARMS

        list_CADP_Sensors = ['IO_SYS','_CADP01', '_CADP02', '_CADP03', '_CADP15','_CADP16' ]
        list_Areas =  ['D1A1','D1A2','D1A3', 'D1A4', 'D1A5','D1A6', 'D2A1','D2A2','D2A3', 'D2A4', 'D2A5','D2A6']
        list99=[]
        list100=[]
        list110 =[]
        for x in list_Areas:
            list99.append(x)
            #print (list9)
            for y in list_CADP_Sensors:
                list100.append(y)
                list120 = list(zip(list99,list100))
                #print(list12)
                list110.append(list120)
                #list11 =list11.append(list12)
                list100 =[] 
            list99=[]
        
        list120=[]
        for lsts in list110:
            for x, y in lsts:
                list120.append(x+y)
        print(list120)

        # Checking the different attributes for the Alarm configuration using, creating a Data frame

        df_Schneider_Area_DP_4 = pd.DataFrame()
        for x in list120:
            df_Schneider_Area_DP_3 = df_Schneider_Area [(df_Schneider_Area['Point_Name'].str.contains(x)) &(df_Schneider_Area['Equipment'].str.contains('AUXILIARY_SYS'))]
            df_Schneider_Area_DP_4 = pd.concat([df_Schneider_Area_DP_4,df_Schneider_Area_DP_3])

        df_Schneider_Area_DP_4.to_excel('C:\Projects\Script\Area\CADP.xlsx')

        
        if (df_Schneider_Area_DP_4.shape[0] >= 1 ):


            df_Cold_DP = df_Schneider_Area_DP_4
            df_Cold_DP = df_Cold_DP.drop (columns = 'Point_Name')
            df_Cold_DP['key1'] = 'CADP'


            list_id_CADP=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Cold Aisle Differential Pressure'):
                    idx_start =index
                    list_id_CADP.append(idx_start)
            
            df_Compare_final_ids_CAdP = df_Compare.iloc[[list_id_CADP[0]]]
            df_Compare_final_ids_CAdP.to_excel ('C:\Projects\Script\Area\CADPCDE.xlsx')

            # Creating a Merged DataFrame

            df_CADP =df_Compare_final_ids_CAdP [['Point_Description','Point_Name','Alarm LO/HI','ALARM DELAY', 'NOTIFICATION LEVEL', 'ALARM TXT', 'DEAD BAND']]
            df_CADP.to_excel ('C:\Projects\Script\Area\PointListNew1.xlsx')
            df_CADP = df_CADP.rename(columns ={df_CADP.columns[2]:'Alarm_Range'})
            df_CADP['Alarm_Range'] = df_CADP['Alarm_Range'].str.replace(r"\(.*\)","")
            df_CADP['DEAD BAND'] = df_CADP['DEAD BAND'].str.replace(r"\(.*\)","")
            df_CADP [['Alarm_Low','Alarm_High']] = df_CADP ['Alarm_Range'].str.split ( '/', expand = True) 
            df_CADP_1 = df_CADP
            df_CADP_1 = df_CADP_1.drop (columns = 'Point_Name')
            df_CADP_1 ['key2'] = 'CADP'

            # Merging the DataFrames

            df_CADP_compare = pd.merge(df_Cold_DP,df_CADP_1,left_on = 'key1', right_on ='key2', how = 'inner')
            df_CADP_compare.to_excel('C:\Projects\Script\Area\CADPCompare.xlsx')
            list_CADP_Compare = df_CADP_compare['index'].to_list()

            # Start Comparing Alarm Attributes 

            #Filtering and Comparing the Pressure Sensor attributes between the Point List and N4 Report
            
            # Comparing the Pressure Sensors attributes 
            # Comparing the low and high Pressure Limits
            # LOW LIMIT

            df_CADP_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('CADP') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            df_CADP_all.to_excel ('C:\Projects\Script\Area\CADPAll.xlsx')
            list_all_CADP= df_CADP_all ['index'].to_list()
            print(list_all_CADP)

            # Creating an Exception here - writing all the rows back to white, this is because some previous codfe have made these rows yellow  

            # Highlighting Rows for SAPDP where the comparision was not made 

            list_CADP_not_comp = [int(x) for x in list_all_CADP if x not in list_CADP_Compare]
            first_row_in_excel=2
            list_CADP_not_comp_1= [x+ first_row_in_excel for x in list_CADP_not_comp]

            print(list_CADP_not_comp_1)

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_CADP_not_comp_1:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)
            
            # Comparing Alarm Attributes 

            df_Pressure_CADP_Low =   df_CADP_compare [['index', 'Equipment','low_Limit','Alarm_Low' ]]
            df_Pressure_CADP_Low['Alarm_High'] = df_Pressure_CADP_Low['Alarm_Low'].replace('"','',regex=True).astype(float)
            df_Pressure_CADP_Low_1 = df_Pressure_CADP_Low[df_Pressure_CADP_Low['low_Limit'].astype(float).values == df_Pressure_CADP_Low['Alarm_Low'].astype(float).values]
            list_smlr_CADP = df_Pressure_CADP_Low_1['index'].to_list()
            list_smlr_CADP= [ x for x in list_smlr_CADP if ~np.isnan(x)]
            print(list_smlr_CADP)
            list_not_smlr_CADP = [int(x) for x in list_CADP_Compare if x not in list_smlr_CADP]
            print (list_not_smlr_CADP)

            # Writing Back to Excel sheet where the Low Pressure limit does not match between the Point List and N4 Report 
            # Writing to Column K to the SpreadSheet

            first_row_in_excel=2
            highlight_rows_alarm_low_CADP_mismatch=[x+ first_row_in_excel for x in list_not_smlr_CADP]
            string="K"
            Cell_numbers_Difference_Highlighted_Low_CADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_low_CADP_mismatch]
            print(Cell_numbers_Difference_Highlighted_Low_CADP)
        
            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Low_CADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Comparing Alarm Attributes - Comparing the High Limit for CADP Alarm  

            
            df_Pressure_CADP_High=   df_CADP_compare [['index', 'Equipment','high_Limit','Alarm_High' ]]
            df_Pressure_CADP_High['Alarm_High'] = df_Pressure_CADP_High['Alarm_High'].replace('"','',regex=True).astype(float)
            df_Pressure_CADP_High_1 = df_Pressure_CADP_High[df_Pressure_CADP_High['high_Limit'].astype(float).values == df_Pressure_CADP_High['Alarm_High'].astype(float).values]
            list_smlr_CADP_1 = df_Pressure_CADP_High_1['index'].to_list()
            list_smlr_CADP_1= [ x for x in list_smlr_CADP_1 if ~np.isnan(x)]
            print(list_smlr_CADP_1)
            list_not_smlr_CADP_1 = [int(x) for x in list_CADP_Compare if x not in list_smlr_CADP]
            print (list_not_smlr_CADP_1)

            first_row_in_excel=2
            highlight_rows_alarm_high_mismatch_1=[x+ first_row_in_excel for x in list_not_smlr_CADP_1]
            string="J"
            Cell_numbers_Difference_Highlighted_High_DP_1= ["{}{}".format(string,i) for i in highlight_rows_alarm_high_mismatch_1]
            print(Cell_numbers_Difference_Highlighted_High_DP_1)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_High_DP_1:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)



            # Comparing Alarm Attributes - Comparing the High Limit for CADP Alarm  

            df_CADP_Alm_Class = df_CADP_compare [['index', 'Equipment','Alarm_Class','NOTIFICATION LEVEL' ]]
            df_CADP_Alm_Class_1=df_CADP_Alm_Class [df_CADP_Alm_Class['Alarm_Class'].astype(str).values ==df_CADP_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
            list_smlr_CADP_Class = df_CADP_Alm_Class_1 ['index'].to_list()
            list_smlr_CADP_Class=[ x for x in list_smlr_CADP_Class if ~np.isnan(x)]
            
            list_not_smlr_CADP_class=[int(x) for x in list_CADP_Compare if x not in list_smlr_CADP_Class]

            first_row_in_excel=2
            highlight_rows_alarm_Class_mismatch_CADP=[x+ first_row_in_excel for x in list_not_smlr_CADP_class]
            string="E"
            Cell_numbers_Difference_Highlighted_Class_CADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Class_mismatch_CADP]
            print(Cell_numbers_Difference_Highlighted_Class_CADP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Class_CADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)


        # Comparing the Delay time

            df_CADP_Delay = df_CADP_compare [['index', 'Equipment','Delay','ALARM DELAY' ]]
            df_CADP_Delay['Delay'] = df_CADP_Delay['Delay'].apply(str)
            df_CADP_Delay['ALARM DELAY'] = df_CADP_Delay['ALARM DELAY'].apply(str)
            df_CADP_Delay['Delay'] =  df_CADP_Delay.Delay.str.replace('seconds', 'sec')
            df_CADP_Delay['Delay'] =  df_CADP_Delay.Delay.str.replace('minutes', 'min')
            df_CADP_Delay['Delay'] =  df_CADP_Delay.Delay.str.replace('1 minute', '60 sec')

            df_delay_CADP=  df_CADP_Delay [df_CADP_Delay['Delay'].astype(str).values ==df_CADP_Delay ['ALARM DELAY'].astype(str).values]

            # Highlighlighting Cells where there is a diffwnce in the Notificatiob clasas
            lst_smlr_CADP_Delay = df_delay_CADP['index'].to_list()
            lst_smlr_CADP_Delay_1=[ x for x in lst_smlr_CADP_Delay if ~np.isnan(x)]
            lst_smlr_CADP_Delay_2=[int(x) for x in list_CADP_Compare if x not in lst_smlr_CADP_Delay_1]

        
            first_row_in_excel=2
            highlight_rows_alarm_Delay_CADP=[x+ first_row_in_excel for x in lst_smlr_CADP_Delay_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Delay_CADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Delay_CADP]
            print(Cell_numbers_Difference_Highlighted_Delay_CADP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Delay_CADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

        # Comparing the Dead Band , the Deadband is an important variable to check for analog type Alarms 

            df_CADP_DdBnd = df_CADP_compare [['index', 'Equipment','Dead_Band','DEAD BAND' ]]
            df_CADP_DdBnd_1 = df_CADP_DdBnd[df_CADP_DdBnd['Dead_Band'].astype(float).values == df_CADP_DdBnd['DEAD BAND'].astype(float).values]
            lst_smlr_CADP_DdBnd =df_CADP_DdBnd_1 ['index'].to_list()
            lst_smlr_CADP_DdBnd_1=[ x for x in lst_smlr_CADP_DdBnd if ~np.isnan(x)]
            lst_smlr_CADP_DdBnd_2=[int(x) for x in  list_CADP_Compare if x not in lst_smlr_CADP_DdBnd_1]

            first_row_in_excel=2
            highlight_rows_alarm_CADP_dbnd=[x+ first_row_in_excel for x in lst_smlr_CADP_DdBnd_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Dnd_CADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_CADP_dbnd]
            print(Cell_numbers_Difference_Highlighted_Dnd_CADP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Dnd_CADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)


    # ADDING THE CODE FOR CHECKING HADP ALARMS

        list_HADP_Sensors = ['IO_SYS','_HADP01', '_HADP02', '_HADP03','_HADP15','_HADP16']
        list_Areas_1 =  ['D1A1','D1A2','D1A3', 'D1A4', 'D1A5','D1A6', 'D2A1','D2A2','D2A3', 'D2A4', 'D2A5','D2A6']
        list990=[]
        list1000=[]
        list1100 =[]
        for x in list_Areas_1:
            list990.append(x)
            #print (list9)
            for y in list_HADP_Sensors:
                list1000.append(y)
                list1200 = list(zip(list990,list1000))
                #print(list12)
                list1100.append(list1200)
                #list11 =list11.append(list12)
                list1000 =[] 
            list990=[]

        list1200=[]
        for lsts in list1100:
            for x, y in lsts:
                list1200.append(x+y)

        df_Schneider_Area_HADP_2 = pd.DataFrame()
        for x in list1200:
            df_Schneider_Area_HADP_1 = df_Schneider_Area [(df_Schneider_Area['Point_Name'].str.contains(x)) &(df_Schneider_Area['Equipment'].str.contains('EF_SYS'))]
            df_Schneider_Area_HADP_2 = pd.concat([df_Schneider_Area_HADP_2,df_Schneider_Area_HADP_1])
        
        df_Schneider_Area_HADP_2.to_excel ('C:\Projects\Script\Area\HADP.xlsx')

        
        if (df_Schneider_Area_HADP_2.shape[0] >= 1 ):


            df_Hot_HADP = df_Schneider_Area_HADP_2
            df_Hot_HADP = df_Hot_HADP.drop (columns = 'Point_Name')
            df_Hot_HADP['key1'] = 'HADP'

            list_id_HADP=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Hot Aisle Differential Pressure'):
                    idx_start =index
                    list_id_HADP.append(idx_start)
            
            df_Compare_final_ids_HAdP = df_Compare.iloc[[list_id_HADP[0]]]
            df_Compare_final_ids_HAdP.to_excel ('C:\Projects\Script\Area\HADPCDE.xlsx')

            df_HADP =df_Compare_final_ids_HAdP [['Point_Description','Point_Name','Alarm LO/HI','ALARM DELAY', 'NOTIFICATION LEVEL', 'ALARM TXT', 'DEAD BAND' ]]

            df_HADP = df_HADP.rename(columns ={df_HADP.columns[2]:'Alarm_Range'})
            df_HADP['Alarm_Range'] = df_HADP['Alarm_Range'].astype(str).str.replace(r"\(.*\)","")
            df_HADP.to_excel('C:\Projects\Script\Area\Test.xlsx')
            df_HADP [['Alarm_Low','Alarm_High']] = df_HADP ['Alarm_Range'].str.split ( '/', expand = True) 
            df_HADP_1 = df_HADP
            df_HADP_1 = df_HADP_1.drop (columns = 'Point_Name')
            df_HADP_1 ['key2'] = 'HADP'


            df_HADP_compare = pd.merge(df_Hot_HADP,df_HADP_1,left_on = 'key1', right_on ='key2', how = 'inner')
            df_HADP_compare.to_excel('C:\Projects\Script\Area\HADPCompare.xlsx')
            list_HADP_Compare = df_HADP_compare['index'].to_list()
            print(list_HADP_Compare)

            # Calling Out those rows which can not be compared 

            df_HADP_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('HADP') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            df_HADP_all.to_excel ('C:\Projects\Script\Area\HADPAll.xlsx')
            list_all_HADP= df_HADP_all ['index'].to_list()
            print(list_all_HADP)

            # Creating a List of all HADP Sensors which can not be compared .

            list_HADP_not_comp = [int(x) for x in list_all_HADP if x not in list_HADP_Compare]
            first_row_in_excel=2
            list_HADP_not_comp_1= [x+ first_row_in_excel for x in list_HADP_not_comp]

            print(list_CADP_not_comp_1)

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_HADP_not_comp_1:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)

            # Highlighting those rows as yellow which have not been checked 
            # Comparing different Alarm attributes 
            # Low Limit of HADP Alarm 
            df_HADP_Comp_Low =   df_HADP_compare [['index', 'Equipment','low_Limit','Alarm_Low' ]]
            df_HADP_Comp_Low['Alarm_Low'] = df_HADP_Comp_Low['Alarm_Low'].replace('"','',regex=True).astype(float)
            df_HADP_Comp_Low['low_Limit'] = df_HADP_Comp_Low['low_Limit'].replace('in/wc','',regex=True).astype(float)
            df_Pressure_HADP_Low_1 = df_HADP_Comp_Low[df_HADP_Comp_Low['low_Limit'].astype(float).values == df_HADP_Comp_Low['Alarm_Low'].astype(float).values]
            list_smlr_HADP = df_Pressure_HADP_Low_1['index'].to_list()
            list_smlr_HADP= [ x for x in list_smlr_HADP if ~np.isnan(x)]
            print(list_smlr_HADP)
            list_not_smlr_HADP = [int(x) for x in list_HADP_Compare if x not in list_smlr_HADP]
            print (list_not_smlr_HADP)
            
            # Writing Back to Excel sheet where the Low Pressure limit does not match between the Point List and N4 Report 
            # Writing to Column K to the SpreadSheet

            first_row_in_excel=2
            highlight_rows_alarm_low_HADP_mismatch=[x+ first_row_in_excel for x in list_not_smlr_HADP]
            string="K"
            Cell_numbers_Difference_Highlighted_Low_HADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_low_HADP_mismatch]
            print(Cell_numbers_Difference_Highlighted_Low_HADP)
        

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Low_HADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # High Limit of HADP Alarm

            df_Pressure_HADP_High=   df_HADP_compare [['index', 'Equipment','high_Limit','Alarm_High' ]]
            df_Pressure_HADP_High['Alarm_High'] = df_Pressure_HADP_High['Alarm_High'].replace('"','',regex=True).astype(float)
            df_Pressure_HADP_High['high_Limit'] = df_Pressure_HADP_High['high_Limit'].replace('in/wc','',regex=True).astype(float)
            df_Pressure_HADP_High_1 = df_Pressure_HADP_High[df_Pressure_HADP_High['high_Limit'].astype(float).values == df_Pressure_HADP_High['Alarm_High'].astype(float).values]
            list_smlr_HADP_1 = df_Pressure_HADP_High_1['index'].to_list()
            list_smlr_HADP_1= [ x for x in list_smlr_HADP_1 if ~np.isnan(x)]
            print(list_smlr_HADP_1)
            list_not_smlr_HADP_1 = [int(x) for x in list_HADP_Compare if x not in list_smlr_HADP_1]
            print (list_not_smlr_HADP_1)


            first_row_in_excel=2
            highlight_rows_alarm_high_mismatch_2=[x+ first_row_in_excel for x in list_not_smlr_HADP_1]
            string="J"
            Cell_numbers_Difference_Highlighted_High_DP_2= ["{}{}".format(string,i) for i in highlight_rows_alarm_high_mismatch_2]
            print(Cell_numbers_Difference_Highlighted_High_DP_2)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_High_DP_2:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)



            # Comparing Alarm Attributes - Comparing the Alarm Notification 

            df_HADP_Alm_Class = df_HADP_compare [['index', 'Equipment','Alarm_Class','NOTIFICATION LEVEL' ]]
            df_HADP_Alm_Class_1=df_HADP_Alm_Class [df_HADP_Alm_Class['Alarm_Class'].astype(str).values ==df_HADP_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
            list_smlr_HADP_Class = df_HADP_Alm_Class_1 ['index'].to_list()
            list_smlr_HADP_Class=[ x for x in list_smlr_HADP_Class if ~np.isnan(x)]
            
            list_not_smlr_HADP_class=[int(x) for x in list_HADP_Compare if x not in list_smlr_HADP_Class]

            first_row_in_excel=2
            highlight_rows_alarm_Class_mismatch_HADP=[x+ first_row_in_excel for x in list_not_smlr_HADP_class]
            string="E"
            Cell_numbers_Difference_Highlighted_Class_HADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Class_mismatch_HADP]
            print(Cell_numbers_Difference_Highlighted_Class_HADP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Class_HADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            #Comparing the Delay time

            df_HADP_Delay = df_HADP_compare [['index', 'Equipment','Delay','ALARM DELAY' ]]
            df_HADP_Delay['Delay'] = df_HADP_Delay['Delay'].apply(str)
            df_HADP_Delay['Delay'] =  df_HADP_Delay.Delay.str.replace('seconds', 'sec')
            df_HADP_Delay['Delay'] =  df_HADP_Delay.Delay.str.replace('minutes', 'min')
            df_HADP_Delay['Delay'] =  df_HADP_Delay.Delay.str.replace('1 minute', '60 sec')

            df_delay_HADP=  df_HADP_Delay [df_HADP_Delay['Delay'].astype(str).values ==df_HADP_Delay ['ALARM DELAY'].astype(str).values]

            # Highlighlighting Cells where there is a diffwnce in the Notificatiob clasas
            lst_smlr_HADP_Delay = df_delay_HADP['index'].to_list()
            lst_smlr_HADP_Delay_1=[ x for x in lst_smlr_HADP_Delay if ~np.isnan(x)]
            lst_smlr_HADP_Delay_2=[int(x) for x in list_HADP_Compare if x not in lst_smlr_HADP_Delay_1]

        
            first_row_in_excel=2
            highlight_rows_alarm_Delay_HADP=[x+ first_row_in_excel for x in lst_smlr_HADP_Delay_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Delay_HADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_Delay_HADP]
            print(Cell_numbers_Difference_Highlighted_Delay_HADP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Delay_HADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # deadbabnd ??


            df_HADP_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('HADP') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            list_all_HADP= df_HADP_all ['index'].to_list()
            list_HADP_aisle_sensors = df_HADP_compare['index'].to_list()
            list_HADP_not_comp= [int(x) for x in list_all_HADP if x not in list_HADP_aisle_sensors ]
            first_row_in_excel=2
            list_HADP_not_comp= [x+ first_row_in_excel for x in list_HADP_not_comp]

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_HADP_not_comp:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)

            # Comparing the Deadband


            # Comparing the Dead Band , the Deadband is an important variable to check for analog type Alarms 

            df_HADP_DdBnd = df_HADP_compare [['index', 'Equipment','Dead_Band','DEAD BAND' ]]
            df_HADP_DdBnd['Dead_Band'] = df_HADP_DdBnd['Dead_Band'].replace('in/wc','',regex=True).astype(float)
            df_HADP_DdBnd['DEAD BAND'] = df_HADP_DdBnd['DEAD BAND'].str.replace(r"\(.*\)","")

            df_HADP_DdBnd_1 = df_HADP_DdBnd[df_HADP_DdBnd['Dead_Band'].astype(float).values == df_HADP_DdBnd['DEAD BAND'].astype(float).values]

            lst_smlr_HADP_DdBnd =df_HADP_DdBnd_1 ['index'].to_list()
            lst_smlr_HADP_DdBnd_1=[ x for x in lst_smlr_HADP_DdBnd if ~np.isnan(x)]
            lst_smlr_HADP_DdBnd_2=[int(x) for x in list_HADP_Compare if x not in lst_smlr_HADP_DdBnd_1]

            first_row_in_excel=2
            highlight_rows_alarm_HADP_dbnd=[x+ first_row_in_excel for x in lst_smlr_HADP_DdBnd_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Dnd_HADP= ["{}{}".format(string,i) for i in highlight_rows_alarm_HADP_dbnd]
            print(Cell_numbers_Difference_Highlighted_Dnd_HADP)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Dnd_HADP:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)


        # Comparing Alarms at RSD, RSD NEEDS TO CHECK 

        list_RSD = ['IO_SYS','_RSD_FBK_A', '_RSD_FBK_B']
        list_Aisles =  ['_AISLE01','_AISLE02','_AISLE03', '_AISLE04', '_AISLE05']
        list_RSD_1 =[]
        list_RSD_2=[]
        list_RSD_3 =[]

        for x in list_RSD:
            list_RSD_1.append(x)
            for y in list_Aisles:
                list_RSD_2.append (y)
                list_RSD_4 = list (zip (list_RSD_1, list_RSD_2))
                list_RSD_3.append(list_RSD_4)
                list_RSD_2=[]
            list_RSD_1=[]

        print(list_RSD_3)

        list12000=[]
        for lsts in list_RSD_3:
            for x, y in lsts:
                list12000.append(y+x)

        # Creating a Data Frame to compare RSD 

        df_Schneider_Area_RSD_2 = pd.DataFrame()
        for x in list12000:
            df_Schneider_Area_RSD_1 = df_Schneider_Area [(df_Schneider_Area['Point_Name'].str.contains(x)) &(df_Schneider_Area['Equipment'].str.contains('Alarms')) ]
            df_Schneider_Area_RSD_2 = pd.concat([df_Schneider_Area_RSD_2,df_Schneider_Area_RSD_1])

        df_Schneider_Area_RSD_2 = df_Schneider_Area_RSD_2[df_Schneider_Area_RSD_2["Point_Name"].apply(lambda x: 'FORCED_ALM' not in x)]

        
        df_Schneider_Area_RSD_2.to_excel('C:\Projects\Script\Area\RSDCompare.xlsx')

        if (df_Schneider_Area_RSD_2.shape[0] >= 1 ):

            df_RSD_Aisle =  df_Schneider_Area_RSD_2
            df_RSD_Aisle['key1'] = 'RSD'

            # CDE Comparision

            list_id_RSD=[]
            for index,rows in df_Compare.iterrows():
                if (df_Compare.loc[index,'Point_Description'] == 'Row Supply Damper Feedback'):
                    idx_start =index
                    list_id_RSD.append(idx_start)
            
            df_Compare_final_ids_RSD = df_Compare.iloc[[list_id_RSD[0]]]

            df_Compare_final_ids_RSD.to_excel('C:\Projects\Script\Area\RSDCDE.xlsx')

            df_RSD =df_Compare_final_ids_RSD [['Point_Description','Point_Name','Alarm LO/HI','ALARM DELAY', 'NOTIFICATION LEVEL', 'ALARM TXT', 'DEAD BAND' ]]
            df_RSD_1 = df_RSD.drop (columns = 'Point_Name')
            df_RSD_1 ['key2'] = 'RSD'

            # Merging the Data Frames to Compare 

            df_RSD_compare = pd.merge(df_RSD_Aisle,df_RSD_1,left_on = 'key1', right_on ='key2', how = 'inner')
            df_RSD_compare.to_excel('C:\Projects\Script\Area\RSD_compare.xlsx')
            list_RSD_Compare = df_RSD_compare['index'].to_list()

            df_RSD_Alm_Class = df_RSD_compare [['index', 'Equipment','Alarm_Class','NOTIFICATION LEVEL' ]]

            df_RSD_Alm_Class_1=df_RSD_Alm_Class [df_RSD_Alm_Class['Alarm_Class'].astype(str).values ==df_RSD_Alm_Class ['NOTIFICATION LEVEL'].astype(str).values]
            list_smlr_RSD_Class = df_RSD_Alm_Class_1 ['index'].to_list()
            list_smlr_RSD_Class=[ x for x in list_smlr_RSD_Class if ~np.isnan(x)]
            
            list_not_smlr_RSD_class=[int(x) for x in list_RSD_Compare if x not in list_smlr_RSD_Class]

            first_row_in_excel=2
            highlight_rows_alarm_Class_mismatch_RSD=[x+ first_row_in_excel for x in list_not_smlr_RSD_class]
            string="E"
            Cell_numbers_Difference_Highlighted_Class_RSD =["{}{}".format(string,i) for i in highlight_rows_alarm_Class_mismatch_RSD]
            print(Cell_numbers_Difference_Highlighted_Class_RSD)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Class_RSD:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Comparing the Delay Time

            df_RSD_Delay = df_RSD_compare [['index', 'Equipment','Delay','ALARM DELAY' ]]
            df_RSD_Delay['Delay'] = df_RSD_Delay['Delay'].astype(str)
            df_RSD_Delay['Delay'] =  df_RSD_Delay.Delay.str.replace('seconds', 'sec')
            df_RSD_Delay['Delay'] =  df_RSD_Delay.Delay.str.replace('minutes', 'min')
            df_RSD_Delay['Delay'] =  df_RSD_Delay.Delay.str.replace('1 minute', '60 sec')

            df_delay_RSD=  df_RSD_Delay [df_RSD_Delay['Delay'].astype(str).values ==df_RSD_Delay ['ALARM DELAY'].astype(str).values]

            # Highlighlighting Cells where there is a diffwnce in the Notificatiob clasas
            lst_smlr_RSD_Delay = df_delay_RSD['index'].to_list()
            lst_smlr_RSD_Delay_1=[ x for x in lst_smlr_RSD_Delay if ~np.isnan(x)]
            lst_smlr_RSD_Delay_2=[int(x) for x in list_RSD_Compare if x not in lst_smlr_RSD_Delay_1]

        
            first_row_in_excel=2
            highlight_rows_alarm_Delay_RSD=[x+ first_row_in_excel for x in lst_smlr_RSD_Delay_2]
            string="G"
            Cell_numbers_Difference_Highlighted_Delay_RSD= ["{}{}".format(string,i) for i in highlight_rows_alarm_Delay_RSD]
            print(Cell_numbers_Difference_Highlighted_Delay_RSD)

            wb = openpyxl.load_workbook(output_File_path)
            ws = wb['Summary'] #Name of the working sheet
            fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
            for cell in Cell_numbers_Difference_Highlighted_Delay_RSD:
                ws[cell].fill = fill_cell1
            wb.save(output_File_path)

            # Highlighting all Rows Containing RSD 

            df_RSD_all =  df_Alrm_Trend[ df_Alrm_Trend['Point_Name'].str.contains('RSD') & (( df_Alrm_Trend['Alarm_Extension'].str.contains('OutOfRangeAlarmExt'))|( df_Alrm_Trend['Alarm_Extension'].str.contains('BooleanChangeOfStateAlarmExt')))]
            list_all_RSD= df_RSD_all ['index'].to_list()
            list_RSD_not_comp= [int(x) for x in list_all_RSD if x not in list_RSD_Compare ]
            first_row_in_excel=2
            list_RSD_not_comp= [x+ first_row_in_excel for x in list_RSD_not_comp]

            with xw.App(visible=False)as app:
                updated_wb= app.books.open(output_File_path)
                updated_ws = updated_wb.sheets('Summary')
                rng=updated_ws.used_range
                print(rng.address) 
                for row in rng.rows:
                    if row.row in list_RSD_not_comp:
                        row.color ='ffff00'
                updated_wb.save(output_File_path)
            
            # D
        
        # Comparing RSD to
        '''
        if ((df_Compare_11.empty == False) | (df_Compare_ER.empty == False)) :
            len_list_tot_AHU=0
            len_list_wtht_diff_AHU =0
        '''
        try:

            if (df_Compare_12.empty == True) :
                len_list_tot_AHU =0 # Tracks the total number of Electrical Room Points
                len_list_wtht_diff_AHU =0
                #len_list_ER_3 =0 
        except:
                len_list_tot_AHU =0 # Tracks the total number of Electrical Room Points
                len_list_wtht_diff_AHU =0




        
        # Resetting the totals to zeros for ER 

        if (df_Compare_ER.empty == True) :
            len_list_ER_1 =0 # Tracks the total number of Electrical Room Points
            len_list_ER_2 =0
            len_list_ER_3 =0 

               
        # Resetting the totals to zeros for CRAHs

        if (df_Compare_CRAH.empty == True) :
            len_list_CRAH_1 =0 # Total Number of CRAHU Points
            len_list_CRAH_2 =0
            len_list_CRAH_3 =0 # Total Number of CRAHU Points which are similiar
    
        if(df_Compare_13.empty == True):
            len_list_IW_1 =0
            len_list_IW_3 =0
       

                
        #len_list_tot_CAT_Schndr = len_list_tot_CAT_Schndr + len_list_tot_CAT_Schndr_High + len_list_tot_CAT_Schndr_Low + len_list_tot_CAT_Schndr_Dbnd + len_list_tot_CAT_Schndr_Class +len_list_tot_CAT_Schndr_Delay +len_list_DAHU_1_Schn 
        # len_list_tot_CAT_Schndr is already calculated
        len_list_smlr_CAT_Schndr =  len_list_smlr_CAT_Schndr_High + len_list_smlr_CAT_Schndr_Low + len_list_smlr_CAT_Schndr_Dbnd+ len_list_smlr_CAT_Schndr_Class +len_list_smlr_CAT_Schndr_Delay
        len_list_dissmlr_CAT_Schndr =len_list_dissmlr_CAT_Schndr_High + len_list_dissmlr_CAT_Schndr_Low + len_list_dissmlr_CAT_Schndr_Dbnd+ len_list_dissmlr_CAT_Schndr_Class +len_list_dissmlr_CAT_Schndr_Delay+(len_list_tot_AHU -len_list_wtht_diff_AHU)+(len_list_IW_1-len_list_IW_3)
        total_point_eval_Schndr = len_list_ER_1+len_list_CRAH_1+len_list_tot_CAT_Schndr + len_list_DAHU_1_Schn+len_list_tot_AHU +  len_list_ER_1+len_list_IW_1
        total_point_match_Schndr =len_list_ER_3+len_list_CRAH_3+ len_list_smlr_CAT_Schndr + len_list_DAHU_2_Schn+len_list_wtht_diff_AHU+len_list_IW_3

        print ('len_list_smlr_CAT_Schndr')
        print (len_list_smlr_CAT_Schndr)
        print ('len_list_dismlr_CAT_Schndr')
        print(len_list_dissmlr_CAT_Schndr)
        print('totalpointeval')
        print (total_point_eval_Schndr)
        print('totalpointmatch')
        print(total_point_match_Schndr)

        # Creating a Report 

        try:

            list_tot_pts_1=[]
            list_smlr_pts_1 = []
            list_tot_pts_1.append(total_point_eval_Schndr)
            list_smlr_pts_1.append(total_point_match_Schndr)
            list_report = list(zip(list_tot_pts_1,list_smlr_pts_1))
            df_report_1 = pd.DataFrame (list_report, columns = ['Total_Points','Compliant_Points'])
            df_report_1['Non_Compliant_Points'] = df_report_1['Total_Points'] - df_report_1['Compliant_Points']
            df_report_1['Percent_Compliant'] = (df_report_1['Compliant_Points']/df_report_1['Total_Points'])*100
            print (' df_report_1')
            print (df_report_1)


        except:
            
            df_report_1 = pd. DataFrame()
            # df_notfound= pd.DataFrame()
            print (' No Points are evaluated!')
        
        df_notfound_1 = pd.DataFrame()
        workbook = openpyxl.load_workbook(output_File_path)
        workbook.create_sheet('Alarms_not_found')
        workbook.create_sheet('Chart')
        #df_notfound.to_excel(writer, sheet_name='Alarms_not_found') 
        workbook.save(output_File_path)
        app = xw.App(visible=False)
        wb = xw.Book(output_File_path)  
        ws = wb.sheets['Alarms_not_found']
        # Update the Workbook
        if(df_notfound_1.empty == False):
            ws.range('A2').options(index=False).value = df_notfound_1
        ws = wb.sheets['Chart']
        if (df_report_1.empty == False):
            print ("Is it Execueting??")
            ws.range('A2').options(index=False).value = df_report_1
        wb.save()
        wb.close()
        app.quit()

        if (df_report_1.empty == False):
            print (' It should draw the Chart')
        
        try:
        
            if (df_report_1.empty == False):
                print ('Drawing the Chart')
                df_report_2 = df_report_1[['Compliant_Points','Non_Compliant_Points']]
                df_report_3 = df_report_2.T
                df_report_4 = df_report_3.rename(columns = {df_report_3.columns[0]:"Point_Distribution"})
                fig = plt.figure()
                plt.pie(df_report_4['Point_Distribution'], labels = ['Compliant_Points','Non_Compliant_Points'], colors =['Green','Red'])
                plt.savefig('C:\Projects\Script\myplot1.png', dpi=150)
                plt.savefig(chart_path)
                wb = openpyxl.load_workbook(output_File_path)
                ws=wb['Chart']
                img = openpyxl.drawing.image.Image(chart_path)
                ws.add_image(img, "A5")
                wb.save(output_File_path)
        except: 
                print ('There is a NaN Value!')



        



        # Load Workbook



            
        sg.popup(" Done! :)")

    except RuntimeError :
        pass

    pass

# ------- GUI Related Functions ------- #

# Displaying Files which are Uploaded, the purpose of this function is that the individual files can be visualized before uploading

def display_first_file(displayalarmFile1):
    df = pd.read_csv(displayalarmFile1)
    filename = Path.name
    sg.popup_scrolled (df.dtypes, "="*50, df, title= filename )
    pass

def display_second_file(displayalarmFile2):
    df = pd.read_csv(displayalarmFile2)
    filename = Path.name
    sg.popup_scrolled (df.dtypes, "="*50, df, title= filename )
    pass

def display_third_file(displayalarmFile3):
    df = pd.read_csv(displayalarmFile3)
    filename = Path.name
    sg.popup_scrolled (df.dtypes, "="*50, df, title= filename )
    pass

def display_fourth_file(displayalarmFile4):
    df = pd.read_excel(displayalarmFile4)
    filename = Path.name
    sg.popup_scrolled (df.dtypes, "="*50, df, title= filename )    
    pass

def is_valid_path(filepath):
    if filepath and Path(filepath).exists():
        return True
    sg.popup_error("Filepath not correct")
    return False

# ---- GUI Definition ---- # 

layout = [ [sg.Text("Alarm File 1 (Boolean Alarms):"), sg.Input (key = "-IN1-"), sg.FileBrowse(file_types =(("csv Files","*.csv"), ))],
           [sg.Text("Alarm File 2 (Numeric Alarms):"), sg.Input (key = "-IN2-"), sg.FileBrowse(file_types =(("csv Files","*.csv"), ))],
           [sg.Text("Trend File 1 (Trend Configurations):"), sg.Input (key = "-IN3-"), sg.FileBrowse(file_types =(("csv Files","*.csv"), ))],
           [sg.Text("Controls Point List (CPL):"), sg.Input (key = "-IN4-"), sg.FileBrowse(file_types =(("Excel Files","*.xlsx"), ))],
           [sg.Text ("Output Folder:"), sg.Input (key = "-OUT-"), sg.FolderBrowse()],[sg.Exit(), sg.Button ("Point Validation Schneider"), sg.Button ("Point Validation Siemens")], 
            [sg.Button('Display Alarm File 1'), sg.Button('Display Alarm File 2')],[sg.Button('Display Trend File'),sg.Button('Point List')],
           [sg.Text('Progress Bar')],[sg.ProgressBar(1000, orientation='h', size=(60, 20), key='progress')]]

window = sg.Window ("N4 Point validation tool", layout)
#status, step = window['Status'], window['Step']
progress_bar = window['progress']


while True:
    event, values = window.read ()
    #print (event, values)
   
    if event == "Point Validation Schneider":
        alarmFile1 =values["-IN1-"]
        alarmFile2 =values["-IN2-"]
        trendFile1 = values["-IN3-"]
        cdepointList = values["-IN4-"]
        output_folder = values ["-OUT-"]
        #x = threading.Thread(target =pointcheckoutSchneider(alarmFile1 =values["-IN1-"],alarmFile2 =values["-IN2-"] ,trendFile1 = values["-IN3-"],cdepointList = values["-IN4-"], output_folder = values ["-OUT-"]), daemon=True )
        if (is_valid_path(alarmFile1) and is_valid_path(alarmFile2) and is_valid_path(trendFile1) and is_valid_path(cdepointList) and is_valid_path(output_folder)) :
            x=threading.Thread(target=pointcheckoutSchneider, args=[alarmFile1, alarmFile2,trendFile1,cdepointList,output_folder])
            x.daemon = True
            x.start()
            time.sleep(1)

            for i in range(1000):
                event, values = window.read(timeout=600)
                if event == 'Exit' or event == sg.WIN_CLOSED or x.is_alive() == False:
                    break
                progress_bar.UpdateBar(i + 1)

    if event == "Point Validation Siemens":
        alarmFile1 =values["-IN1-"]
        alarmFile2 =values["-IN2-"]
        trendFile1 = values["-IN3-"]
        cdepointList = values["-IN4-"]
        output_folder = values ["-OUT-"]
        #x = threading.Thread(target =pointcheckoutSchneider(alarmFile1 =values["-IN1-"],alarmFile2 =values["-IN2-"] ,trendFile1 = values["-IN3-"],cdepointList = values["-IN4-"], output_folder = values ["-OUT-"]), daemon=True )
        if (is_valid_path(alarmFile1) and is_valid_path(alarmFile2) and is_valid_path(trendFile1) and is_valid_path(cdepointList) and is_valid_path(output_folder)) :
            x=threading.Thread(target=pointcheckoutSiemens, args=[alarmFile1, alarmFile2,trendFile1,cdepointList,output_folder])
            x.daemon = True
            x.start()
            time.sleep(1)

            for i in range(1000):
                event, values = window.read(timeout=600)
                if event == 'Exit' or event == sg.WIN_CLOSED or x.is_alive() == False:
                    break
                progress_bar.UpdateBar(i + 1)
    
    


    if event == "Display Alarm File 1":
       displayalarmFile1 =values["-IN1-"]
       if is_valid_path (displayalarmFile1):
           display_first_file(displayalarmFile1)
    
    
    if event == "Display Alarm File 2":
         displayalarmFile2 =values["-IN2-"]
         if is_valid_path (displayalarmFile2):
             display_second_file(displayalarmFile2)
    
    if event == "Display Trend File":
         displayalarmFile3 =values["-IN3-"]
         if is_valid_path (displayalarmFile3):
             display_third_file(displayalarmFile3)

    if event == "Point List":
        displayalarmFile4 =values["-IN4-"]
        if is_valid_path (displayalarmFile4):
            display_fourth_file(displayalarmFile4)
    
    

    #sg.popup_no_titlebar(" Done! :)")
    if event in (sg.WINDOW_CLOSED,"Exit"):
        try:
            x.terminate()
        except AttributeError or PermissionError or NameError:
            window.close()
            pass
        break 
    if event == sg.WINDOW_CLOSED :
        try:
            x.terminate()
        except PermissionError or NameError:
            window.close()
            pass
        break
    
            
window.close()

# Creating a chart to show the overall results 