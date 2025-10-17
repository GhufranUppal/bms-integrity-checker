import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Original Evap Cooler data
data1 = [
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG1][FLOW][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP1][STG2][FLOW][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG1][FLOW][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP2][STG2][FLOW][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[PWR][UPS][STATUS]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][MODE][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SAT][STPT][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[BMS][SF][SPD][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][LOWER][ALM]', 'ALARM DELAY': '60 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[BYPD][UPPER][ALM]', 'ALARM DELAY': '60 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[ECFAN][FAIL][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][CMD]', 'ALARM DELAY': '30 secs', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][HIGH][WATER][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[EVAP][LOW][WATER][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][LOWER][STS]', 'ALARM DELAY': '60 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[FEC][UPPER][STS]', 'ALARM DELAY': '2 minute', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[FILDP][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[PREFILTER]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[FINALFILTER]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][LOWER][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAD][UPPER][ALM]', 'ALARM DELAY': '30 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][HIGH][ALM]', 'ALARM DELAY': '30 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][LOW][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'Medium'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][HI][LVL][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][LOW][LVL][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP1][PMP][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][HI][LVL][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][LOW][LVL][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SUMP2][PMP][ALM]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[UPS][POWER]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T1]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT][T2]', 'ALARM DELAY': '5 sec', 'NOTIFICATION LEVEL': 'High'},
    {'Point_Description': 'Evap Cooler', 'Point_Name': '[SAT]', 'ALARM DELAY': '30 sec', 'NOTIFICATION LEVEL': 'High'},
]

# Function to improve description based on Point_Name
def improve_description(point_name):
    parts = point_name.strip("[]").split("][")
    description_map = {
        "EVAP": "Evaporative Cooler",
        "EVAP1": "Evaporative Cooler 1",
        "EVAP2": "Evaporative Cooler 2",
        "STG1": "Stage 1",
        "STG2": "Stage 2",
        "FLOW": "Flow",
        "ALM": "Alarm",
        "UPS": "UPS",
        "PWR": "Power",
        "BMS": "Building Management System",
        "MODE": "Mode",
        "SAT": "Supply Air Temperature",
        "STPT": "Setpoint",
        "SF": "Supply Fan",
        "SPD": "Speed",
        "BYPD": "Bypass Damper",
        "LOWER": "Lower",
        "UPPER": "Upper",
        "ECFAN": "Evaporative Cooler Fan",
        "FAIL": "Failure",
        "FEC": "Fan Exhaust Cooler",
        "CMD": "Command",
        "HIGH": "High",
        "LOW": "Low",
        "WATER": "Water",
        "STS": "Status",
        "FILDP": "Filter Differential Pressure",
        "PREFILTER": "Pre-Filter",
        "FINALFILTER": "Final Filter",
        "SAD": "Supply Air Damper",
        "SUMP1": "Sump 1",
        "SUMP2": "Sump 2",
        "PMP": "Pump",
        "T1": "Temperature Sensor 1",
        "T2": "Temperature Sensor 2"
    }
    description = " ".join([description_map.get(part, part) for part in parts])
    return description

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Evap Cooler DAHU WATER Points"

# Write headers
headers = ["Point_Description", "Point_Name", "ALARM DELAY", "NOTIFICATION LEVEL"]
ws.append(headers)

# Add improved Evap Cooler points
for item in data1:
    improved_desc = improve_description(item['Point_Name'])
    ws.append([improved_desc, item['Point_Name'], item['ALARM DELAY'], item['NOTIFICATION LEVEL']])

# Add DAHU heading
ws.append(["DAHU", "", "", ""])

# Add DAHU points
dahu_points = [
    ("Dedicated AHU Supply Fan Alarm", "[DAHU][SF][ALM]", "5 sec", "High"),
    ("Dedicated AHU Filter Differential Pressure Alarm", "[DAHU][FILDP][ALM]", "5 sec", "Medium"),
    ("Dedicated AHU Supply Air Temp High Alarm", "[DAHU][SAT][HIGH][ALM]", "30 sec", "Medium"),
    ("Dedicated AHU Supply Air Temp Low Alarm", "[DAHU][SAT][LOW][ALM]", "5 sec", "Medium"),
    ("Dedicated AHU Mode Alarm", "[DAHU][MODE][ALM]", "5 sec", "High")
]
for point in dahu_points:
    ws.append(point)

# Add blank row after DAHU points
ws.append(["", "", "", ""])

# Add INDUSTRIAL WATER SYSTEM heading
ws.append(["INDUSTRIAL WATER SYSTEM", "", "", ""])

# Add Industrial Water System points
industrial_water_points = [
    ("Industrial Water Pump Failure Alarm", "[IWS][PUMP][FAIL][ALM]", "5 sec", "High"),
    ("Industrial Water Tank High Level Alarm", "[IWS][TANK][HIGH][LVL][ALM]", "5 sec", "High"),
    ("Industrial Water Tank Low Level Alarm", "[IWS][TANK][LOW][LVL][ALM]", "5 sec", "High"),
    ("Industrial Water Flow Alarm", "[IWS][FLOW][ALM]", "5 sec", "Medium"),
    ("Industrial Water Filter Differential Pressure Alarm", "[IWS][FILDP][ALM]", "5 sec", "Medium"),
    ("Industrial Water Supply Pressure Low Alarm", "[IWS][SUPPLY][PRESSURE][LOW][ALM]", "5 sec", "High")
]
for point in industrial_water_points:
    ws.append(point)

# Adjust column widths
for col in range(1, ws.max_column + 1):
    max_length = 0
    col_letter = get_column_letter(col)
    for cell in ws[col_letter]:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Save the workbook
wb.save("C:\GHUFRAN\Old\PythonScripting\Niagara\Evap_Cooler_DAHU_WATER_Points.xlsx")
print("Excel file 'Evap_Cooler_DAHU_WATER_Points.xlsx' has been created successfully.")



electrical_points = [
    ("Main Breaker Trip Alarm", "[ELEC][MAIN][BRKR][TRIP][ALM]", "5 sec", "High"),
    ("Panel Overheat Alarm", "[ELEC][PNL][OVERHEAT][ALM]", "10 sec", "High"),
    ("UPS Failure Alarm", "[ELEC][UPS][FAIL][ALM]", "5 sec", "High"),
    ("Room Temperature High Alarm", "[ELEC][ROOM][TEMP][HIGH][ALM]", "30 sec", "Medium"),
    ("Smoke Detection Alarm", "[ELEC][SMOKE][ALM]", "5 sec", "High"),
    ("Power Loss Alarm", "[ELEC][POWER][LOSS][ALM]", "5 sec", "High"),
    ("Generator Running Status", "[ELEC][GEN][RUN][STS]", "5 sec", "Medium"),
    ("Battery Low Alarm", "[ELEC][BATTERY][LOW][ALM]", "10 sec", "High"),
    ("Surge Protection Device Alarm", "[ELEC][SPD][ALM]", "5 sec", "Medium"),
    ("Circuit Breaker Status", "[ELEC][CB][STS]", "5 sec", "Medium")
]
